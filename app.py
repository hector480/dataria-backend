"""
Dataria · Backend de orquestación en vivo (archivo único)
==========================================================
Versión combinada para despliegue simple: dpo + assembler + main en un solo app.py.
Endpoints:  GET /health   ·   POST /api/zona/analyze
"""

# ╔══════════════ SECCIÓN 1: DERIVACIÓN DIGO/DPO (dpo) ══════════════╗
import statistics
import asyncio as _aio
import math
import random
import re
import os
import time as _time
import datetime as _dt
from pathlib import Path
from typing import Optional, List, Dict, Any

# Banda de plausibilidad física (m²)
M2_MIN, M2_MAX = 25, 500
CV_BARRERA = 0.30      # umbral de "barrera amplia" de percepción de valor
COBERTURA_MIN = 0.20   # 20% mínimo de cobertura de isócrona sobre clúster alto valor


# ──────────────────────── Helpers numéricos N/D-safe ────────────────────────
def _num(v) -> Optional[float]:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # -1 es marcador de "no disponible" en los datasets de origen
        if v == -1:
            return None
        return float(v)
    return None


def _price(v) -> Optional[float]:
    """Precio de unidad plausible: > 100k. Filtra -1, 0 y placeholders."""
    n = _num(v)
    return n if (n is not None and n > 100000) else None


def _pm2(v) -> Optional[float]:
    """Precio por m² plausible: > 1000."""
    n = _num(v)
    return n if (n is not None and n > 1000) else None


def _sum(rows: List[Dict], key: str) -> Optional[float]:
    vals = [_num(r.get(key)) for r in rows]
    vals = [v for v in vals if v is not None]
    return sum(vals) if vals else None


def _wavg(rows: List[Dict], key: str, weight: str) -> Optional[float]:
    num = 0.0
    den = 0.0
    for r in rows:
        v = _num(r.get(key)); w = _num(r.get(weight))
        if v is not None and w is not None and w > 0:
            num += v * w; den += w
    return (num / den) if den > 0 else None


# ════════════ RES-2 · ESTADÍSTICA ROBUSTA (estándar de toda la herramienta) ════════════
# Regla de Héctor (coincide con el estándar): NUNCA promedios simples. Lo "normal" se
# describe con MEDIANA + MAD; las bandas con percentiles; los "anormales" (outliers) se
# detectan con la regla de Tukey sobre el IQR sin que distorsionen la lectura.
def _percentil(vals, p: float):
    """Percentil p (0-1) con interpolación lineal. None si no hay datos."""
    v = sorted(x for x in vals if x is not None)
    if not v:
        return None
    if len(v) == 1:
        return v[0]
    k = (len(v) - 1) * p
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return v[int(k)]
    return v[f] + (v[c] - v[f]) * (k - f)


def _stats_robustas(vals, k_iqr: float = 1.5) -> Dict[str, Any]:
    """Describe una muestra con métricas ROBUSTAS (catálogo universal):
      • mediana, mad (desviación absoluta mediana; mad×1.4826 ≈ σ equivalente),
      • cv_robusto = (mad×1.4826)/mediana,
      • p10/p25/p75/p90 (bandas), iqr,
      • outliers por Tukey: fuera de [p25 − k·IQR, p75 + k·IQR] (k=1.5 estándar).
    Sin datos → n=0 y None en todo (N/D legítimo, jamás inventar)."""
    v = sorted(x for x in vals if x is not None)
    if not v:
        return {"n": 0, "mediana": None, "mad": None, "cv_robusto": None,
                "p10": None, "p25": None, "p75": None, "p90": None, "iqr": None,
                "outliers_n": 0, "outliers": [], "min": None, "max": None}
    med = statistics.median(v)
    mad = statistics.median([abs(x - med) for x in v]) if len(v) > 1 else 0.0
    p10 = _percentil(v, 0.10); p25 = _percentil(v, 0.25)
    p75 = _percentil(v, 0.75); p90 = _percentil(v, 0.90)
    iqr = (p75 - p25) if (p75 is not None and p25 is not None) else None
    lo = (p25 - k_iqr * iqr) if iqr is not None else None
    hi = (p75 + k_iqr * iqr) if iqr is not None else None
    outliers = [x for x in v if (lo is not None and x < lo) or (hi is not None and x > hi)]
    return {"n": len(v), "mediana": round(med, 2), "mad": round(mad, 2),
            "cv_robusto": round((1.4826 * mad) / med, 4) if med else None,
            "p10": round(p10, 2), "p25": round(p25, 2), "p75": round(p75, 2), "p90": round(p90, 2),
            "iqr": round(iqr, 2) if iqr is not None else None,
            "outliers_n": len(outliers), "outliers": [round(x, 2) for x in outliers[:12]],
            "min": round(v[0], 2), "max": round(v[-1], 2)}


# ──────────────────────── Isócrona por tamaño/uso ────────────────────────
def isochrone_profile(predio_m2: Optional[float], comercial: bool) -> Dict[str, Any]:
    """Replica exacta de la regla del tablero (sección Zona de análisis)."""
    if comercial or (predio_m2 is not None and predio_m2 > 35000):
        return {"key": "p24seg", "label": "24 min (segmentada 8/14)", "minutos": [8, 14, 24], "principal": 24}
    if predio_m2 is None:
        # Sin tamaño y sin comercial: default conservador 8 min
        return {"key": "p8", "label": "8 min", "minutos": [8], "principal": 8}
    if predio_m2 < 1500:
        return {"key": "p8", "label": "8 min", "minutos": [8], "principal": 8}
    if predio_m2 < 15000:
        return {"key": "p14", "label": "14 min", "minutos": [8, 14], "principal": 14}
    if predio_m2 <= 35000:
        return {"key": "p18", "label": "18 min", "minutos": [8, 18], "principal": 18}
    return {"key": "p24seg", "label": "24 min (segmentada 8/14)", "minutos": [8, 14, 24], "principal": 24}


# ──────────────────────── Geometría: punto en anillo ────────────────────────
def _point_in_ring(lng: float, lat: float, ring: List[List[float]]) -> bool:
    inside = False
    n = len(ring)
    j = n - 1
    for i in range(n):
        xi, yi = ring[i][0], ring[i][1]
        xj, yj = ring[j][0], ring[j][1]
        if ((yi > lat) != (yj > lat)) and (lng < (xj - xi) * (lat - yi) / (yj - yi) + xi):
            inside = not inside
        j = i
    return inside


def _row_coords(row: Dict):
    """Extrae (lng, lat) de un registro VVV (geometry o X_coor/Y_coor)."""
    a = row.get("attributes", {})
    g = row.get("geometry", {})
    lat = g.get("y") if isinstance(g.get("y"), (int, float)) else a.get("Y_coor")
    lng = g.get("x") if isinstance(g.get("x"), (int, float)) else a.get("X_coor")
    if isinstance(lat, (int, float)) and isinstance(lng, (int, float)):
        return (lng, lat)
    return (None, None)


def filter_vvv_by_polygon(vvv: Dict, ring: List[List[float]]) -> Dict:
    """
    Filtra los datasets de VVV (resumen, ft, pagos) dejando SOLO los registros cuyo
    punto cae DENTRO del polígono de la isócrona. Garantía de integridad espacial:
    el backend de ArcGIS puede usar bounding box; aquí imponemos el polígono exacto.
    Universal para cualquier zona de México.
    """
    if not vvv or "datasets" not in vvv:
        return vvv
    ds = vvv["datasets"]
    resumen = ds.get("resumen", [])
    ft = ds.get("ft", [])
    pagos = ds.get("pagos", [])

    # 1. Proyectos (resumen) dentro del polígono → set de nombres válidos
    resumen_in = []
    proyectos_validos = set()
    for row in resumen:
        lng, lat = _row_coords(row)
        if lng is not None and _point_in_ring(lng, lat, ring):
            resumen_in.append(row)
            nombre = row.get("attributes", {}).get("PROYECTO")
            if nombre:
                proyectos_validos.add(nombre)

    # 2. ft (tipologías): dentro del polígono por coords propias O pertenece a proyecto válido
    ft_in = []
    for t in ft:
        lng, lat = _row_coords(t)
        nombre = t.get("attributes", {}).get("PROYECTO")
        if lng is not None and _point_in_ring(lng, lat, ring):
            ft_in.append(t)
        elif nombre and nombre in proyectos_validos:
            ft_in.append(t)

    # 3. pagos: solo de proyectos válidos. Integridad espacial: sin proyectos válidos dentro
    # del polígono NO se dejan pasar pagos huérfanos (antes se devolvían todos sin filtrar).
    pagos_in = [p for p in pagos if p.get("attributes", {}).get("PROYECTO") in proyectos_validos] \
        if proyectos_validos else []

    out = dict(vvv)
    out["datasets"] = {"resumen": resumen_in, "ft": ft_in, "pagos": pagos_in}
    out["_spatial_filter"] = {
        "resumen_in": len(resumen_in), "resumen_total": len(resumen),
        "ft_in": len(ft_in), "ft_total": len(ft),
    }
    return out


# ──────────────────────── Envolvente convexa (hull) ────────────────────────
def _convex_hull(points: List[List[float]]) -> List[List[float]]:
    """Envolvente convexa (monotone chain) sobre puntos [[lng, lat], ...].

    Devuelve el anillo CERRADO (primer punto repetido al final) en formato
    [[lng, lat], ...]. Si hay menos de 3 puntos distintos, devuelve [] (no se
    puede formar un polígono → integridad: sin invención).
    """
    pts = sorted(set((float(p[0]), float(p[1])) for p in points
                     if p and p[0] is not None and p[1] is not None))
    if len(pts) < 3:
        return []

    def cross(o, a, b):
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])

    lower = []
    for p in pts:
        while len(lower) >= 2 and cross(lower[-2], lower[-1], p) <= 0:
            lower.pop()
        lower.append(p)
    upper = []
    for p in reversed(pts):
        while len(upper) >= 2 and cross(upper[-2], upper[-1], p) <= 0:
            upper.pop()
        upper.append(p)
    hull = lower[:-1] + upper[:-1]
    if len(hull) < 3:
        return []
    ring = [[x, y] for (x, y) in hull]
    ring.append(ring[0])  # cerrar el anillo
    return ring


# ════════════════════ Detección universal de mercados (clustering) ════════════════════
#
# REGLA UNIVERSAL DE ZONA DE INFLUENCIA REAL
# ------------------------------------------
# Una isócrona puede contener DOS (o más) mercados distintos pegados, separados por una
# barrera —física (río con pocos puentes), de NSE del AGEB, o de percepción de valor—.
# El CV global puede ser moderado y aun así ocultar la frontera: lo que la delata es que
# al PARTIR los proyectos en grupos espacialmente contiguos, cada grupo es internamente
# homogéneo (CV interno bajo) y sus medias son muy distintas.
#
# No buscamos la barrera física; buscamos su EFECTO en los datos. Así la regla es universal,
# exista o no un río. El método:
#   1. Vector por proyecto con señales normalizadas y ponderadas:
#        ticket/unidad (0.40) · precio $/m² (0.30) · posición vs pin (0.30)
#        [+ NSE del AGEB (gancho enchufable) cuando exista geometría por CVEGEO]
#   2. k-means con k AUTOMÁTICO (k=1,2,3): se elige la partición con mejor separabilidad.
#   3. Validación robusta de que hay >1 mercado: gap de medias normalizado ≥ 0.85 Y
#      varianza explicada ≥ 0.35, con masa mínima por cluster (≥ 3 proyectos).
#   4. La zona de influencia real = cluster MÁS COMPATIBLE CON EL PIN (centroide = pin).
#
# Pesos de las señales (renormalizados si alguna señal no está disponible → integridad).
# nse ACTIVADO (pendiente prioritario #2 · 8 jul 2026): el nse_rank por proyecto se toma del
# AGEB georreferenciado más cercano (agebs_geo del KMZ del DI). Sin geometría, el peso se
# renormaliza solo (integridad). La barrera de mercado ahora ve las 4 señales de la regla:
# física (pos), percepción ($/m² y ticket) y NSE.
SIGNAL_WEIGHTS = {"ticket": 0.40, "pm2": 0.30, "pos": 0.30, "nse": 0.15}
# Soporte mínimo de muestra para detección de mercados: con menos proyectos el clúster
# degenera (zona de pocos vértices). Si el anillo base trae menos, la percepción se evalúa
# sobre el anillo MAYOR del perfil y se declara (el recorte al mercado del pin sigue).
ZONA_MUESTRA_MIN = int(os.environ.get("DATARIA_ZONA_MUESTRA_MIN", "15"))
CLUSTER_MASA_MIN = 3       # proyectos mínimos por cluster para considerarlo un mercado
GAP_MIN = 0.85             # separación de medias normalizada mínima
VAR_EXPLAINED_MIN = 0.35   # proporción de varianza explicada por la partición mínima
# Salvaguardas contra falsos positivos (zonas realmente uniformes):
CV_GLOBAL_MIN = 0.12       # si el CV global de la zona es menor, NO hay dos mercados que separar
SEP_MEDIAS_MIN = 0.18      # diferencia mínima de medias entre mercados (proporción), p.ej. 18%
SOLAPE_ESPACIAL_MAX = 0.60 # solape máximo permitido de los rangos espaciales de los clusters
KMEANS_RESTARTS = 12
KMEANS_ITERS = 60


def _zscore_params(vals: List[float]):
    """(media, desviación) ignorando None; desviación 1.0 si no hay dispersión."""
    v = [x for x in vals if x is not None]
    if len(v) < 2:
        return (statistics.mean(v) if v else 0.0), 1.0
    m = statistics.mean(v)
    s = statistics.pstdev(v) or 1.0
    return m, s


def fetch_ageb_geometria(cvegeo_list: List[str]) -> Dict[str, Dict[str, float]]:
    """GANCHO ENCHUFABLE · georreferenciación de AGEB por CVEGEO.

    Devuelve {cvegeo: {"lng": x, "lat": y}} con el CENTROIDE de cada AGEB para usar
    el NSE del AGEB como 4ª señal del clustering. Hoy PRSP no expone (aún) un endpoint
    documentado de geometría por CVEGEO, por lo que devuelve {} (NSE = N/D, no se inventa).

    Cuando exista la documentación del endpoint PRSP de geometría de AGEB:
      • Implementar aquí la consulta (POST {PRSP_BASE}/<ruta>, body con cvegeo_list).
      • Parsear el centroide/geometría de cada CVEGEO de la respuesta.
      • Subir SIGNAL_WEIGHTS["nse"] (p. ej. a 0.15) y renormalizar los demás pesos.
    El resto del clustering ya está preparado para incorporar la señal sin reescribirse.
    """
    return {}


def _build_feature_vectors(proyectos: List[Dict], pin_lng: float, pin_lat: float):
    """Vectores ponderados y normalizados por proyecto. Renormaliza pesos según
    las señales realmente disponibles (integridad: no se inventan señales faltantes)."""
    tickets = [p.get("ticket") for p in proyectos]
    pm2s = [p.get("pm2") for p in proyectos]
    lngs = [p.get("lng") for p in proyectos]
    lats = [p.get("lat") for p in proyectos]

    has_ticket = sum(1 for t in tickets if t is not None) >= max(3, len(proyectos) // 2)
    has_nse = any(p.get("nse_rank") is not None for p in proyectos)  # gancho NSE

    # Pesos activos según disponibilidad real de señales
    w = dict(SIGNAL_WEIGHTS)
    if not has_ticket:
        w["ticket"] = 0.0
    if not has_nse:
        w["nse"] = 0.0
    total_w = sum(w.values()) or 1.0
    w = {k: v / total_w for k, v in w.items()}  # renormalizar

    m_tk, s_tk = _zscore_params(tickets)
    m_pm, s_pm = _zscore_params(pm2s)
    m_ln, s_ln = _zscore_params(lngs)
    m_lt, s_lt = _zscore_params(lats)
    nse_ranks = [p.get("nse_rank") for p in proyectos]
    m_ns, s_ns = _zscore_params(nse_ranks)

    X = []
    for p in proyectos:
        f = []
        # ticket (si falta en un proyecto puntual, proxy con pm² para no perderlo)
        if w["ticket"] > 0:
            tval = p.get("ticket")
            if tval is not None:
                f.append(w["ticket"] * ((tval - m_tk) / s_tk))
            else:
                f.append(w["ticket"] * ((p.get("pm2", m_pm) - m_pm) / s_pm))
        if w["pm2"] > 0:
            f.append(w["pm2"] * ((p.get("pm2", m_pm) - m_pm) / s_pm))
        if w["pos"] > 0:
            # posición relativa al PIN (no al centro de masa): lng pesa completo, lat mitad
            f.append(w["pos"] * ((p["lng"] - pin_lng) / s_ln))
            f.append(w["pos"] * 0.5 * ((p["lat"] - pin_lat) / s_lt))
        if w["nse"] > 0 and p.get("nse_rank") is not None:
            f.append(w["nse"] * ((p["nse_rank"] - m_ns) / s_ns))
        X.append(f)
    return X, w


def _euclid(a, b):
    return math.sqrt(sum((a[i] - b[i]) ** 2 for i in range(len(a))))


def _kmeans(X: List[List[float]], k: int, seed: int = 42):
    """k-means con múltiples reinicios. Devuelve (labels, inertia)."""
    if k <= 1 or len(X) <= k:
        return [0] * len(X), 0.0
    rnd = random.Random(seed)
    best = None
    for _ in range(KMEANS_RESTARTS):
        idx = rnd.sample(range(len(X)), k)
        centroids = [X[i][:] for i in idx]
        labels = [0] * len(X)
        for _ in range(KMEANS_ITERS):
            for i, x in enumerate(X):
                labels[i] = min(range(k), key=lambda c: _euclid(x, centroids[c]))
            moved = False
            for c in range(k):
                grp = [X[i] for i in range(len(X)) if labels[i] == c]
                if grp:
                    nc = [sum(g[j] for g in grp) / len(grp) for j in range(len(X[0]))]
                    if nc != centroids[c]:
                        centroids[c] = nc
                        moved = True
            if not moved:
                break
        inertia = sum(_euclid(X[i], centroids[labels[i]]) ** 2 for i in range(len(X)))
        if best is None or inertia < best[1]:
            best = (labels[:], inertia)
    return best


def _partition_metrics(proyectos: List[Dict], labels: List[int], k: int):
    """Métricas de separabilidad sobre pm² entre los k grupos.
    Devuelve (gap_min_normalizado, varianza_explicada, masa_ok, stats_por_grupo)."""
    precios_all = [p["pm2"] for p in proyectos]
    if len(precios_all) < 2:
        return 0.0, 0.0, False, []
    M = statistics.mean(precios_all)
    var_total = statistics.pvariance(precios_all) or 1e-9

    grupos = []
    for c in range(k):
        gp = [proyectos[i]["pm2"] for i in range(len(proyectos)) if labels[i] == c]
        if gp:
            grupos.append({
                "c": c, "n": len(gp),
                "media": statistics.mean(gp),
                "sd": statistics.pstdev(gp),
                "cv": (statistics.pstdev(gp) / statistics.mean(gp)) if statistics.mean(gp) else None,
            })
    if len(grupos) < 2:
        return 0.0, 0.0, False, grupos

    # masa mínima por grupo
    masa_ok = all(g["n"] >= CLUSTER_MASA_MIN for g in grupos)

    # gap normalizado: para cada par adyacente por media, (Δmedias)/(sd_i+sd_j); tomamos el mínimo
    gs = sorted(grupos, key=lambda g: g["media"])
    gaps = []
    for i in range(len(gs) - 1):
        denom = (gs[i]["sd"] + gs[i + 1]["sd"]) or 1.0
        gaps.append(abs(gs[i + 1]["media"] - gs[i]["media"]) / denom)
    gap_min = min(gaps) if gaps else 0.0

    # varianza explicada = varianza entre medias / varianza total
    n_tot = sum(g["n"] for g in grupos)
    between = sum(g["n"] * (g["media"] - M) ** 2 for g in grupos) / n_tot
    var_explained = between / var_total

    return gap_min, var_explained, masa_ok, grupos


def _separacion_espacial(proyectos: List[Dict], labels: List[int], c_a: int, c_b: int) -> float:
    """Mide qué tan separados ESPACIALMENTE están dos clusters (0=encimados, 1=disjuntos).

    Proyecta cada proyecto sobre el eje que une los centroides de los dos clusters y
    calcula el solape de los rangos proyectados. Devuelve 1 - solape: alto = frontera
    espacial real (mercados en lados distintos), bajo = grupos entremezclados (falso
    positivo de un mercado uniforme partido artificialmente)."""
    A = [proyectos[i] for i in range(len(proyectos)) if labels[i] == c_a]
    B = [proyectos[i] for i in range(len(proyectos)) if labels[i] == c_b]
    if not A or not B:
        return 0.0
    ca = (statistics.mean(p["lng"] for p in A), statistics.mean(p["lat"] for p in A))
    cb = (statistics.mean(p["lng"] for p in B), statistics.mean(p["lat"] for p in B))
    vx, vy = cb[0] - ca[0], cb[1] - ca[1]
    norm = math.hypot(vx, vy) or 1e-9
    vx, vy = vx / norm, vy / norm  # eje unitario centroide A→B

    def proj_range(grp):
        vals = [p["lng"] * vx + p["lat"] * vy for p in grp]
        return min(vals), max(vals)
    a_lo, a_hi = proj_range(A)
    b_lo, b_hi = proj_range(B)
    # solape de los dos intervalos sobre el eje
    inter = max(0.0, min(a_hi, b_hi) - max(a_lo, b_lo))
    union = max(a_hi, b_hi) - min(a_lo, b_lo)
    solape = inter / union if union > 0 else 1.0
    return 1.0 - solape


def detect_markets(proyectos: List[Dict], pin_lng: float, pin_lat: float) -> Dict[str, Any]:
    """Detección universal de mercados dentro de la isócrona.

    Entrada: proyectos [{name, pm2, lat, lng, ticket?, nse_rank?}] (ya dentro de la isócrona).
    Salida: dict con la partición elegida y la asignación del PIN.
      • barrera (bool): True si se detecta >1 mercado robusto.
      • k_elegido, labels, grupos, gap, var_explained.
      • pin_cluster: índice del cluster donde cae el pin (mercado compatible con el predio).
      • cluster_proyectos: lista de proyectos del cluster del pin.
      • pesos: pesos de señales realmente aplicados (auditoría/integridad).

    Salvaguardas contra falsos positivos (zonas realmente uniformes):
      1. CV global de la zona ≥ CV_GLOBAL_MIN (si la zona ya es uniforme, no se parte).
      2. Diferencia de medias entre mercados ≥ SEP_MEDIAS_MIN (separación REAL, no relativa).
      3. Separación espacial de los clusters ≥ (1 - SOLAPE_ESPACIAL_MAX) (frontera real).
    """
    out = {
        "barrera": False, "k_elegido": 1, "labels": [0] * len(proyectos),
        "grupos": [], "gap": None, "var_explained": None,
        "pin_cluster": 0, "cluster_proyectos": proyectos, "pesos": None,
        "motivo": None,
    }
    n = len(proyectos)
    if n < 2 * CLUSTER_MASA_MIN:
        out["motivo"] = f"Muestra insuficiente para detectar mercados (n={n} < {2*CLUSTER_MASA_MIN})"
        return out

    # SALVAGUARDA 1: si la zona ya es uniforme (CV global bajo), no hay mercados que separar.
    precios = [p["pm2"] for p in proyectos]
    media_global = statistics.mean(precios)
    cv_global = (statistics.pstdev(precios) / media_global) if media_global else 0.0
    if cv_global < CV_GLOBAL_MIN:
        out["motivo"] = (f"Zona uniforme (CV global {cv_global:.2f} < {CV_GLOBAL_MIN}) · "
                         f"un solo mercado (zona = isócrona)")
        return out

    X, pesos = _build_feature_vectors(proyectos, pin_lng, pin_lat)
    out["pesos"] = pesos

    # k automático: probar k=2 y k=3, validar separabilidad, elegir el mejor robusto
    candidatos = []
    for k in (2, 3):
        if n < k * CLUSTER_MASA_MIN:
            continue
        labels, _ = _kmeans(X, k, seed=42)
        gap, var_exp, masa_ok, grupos = _partition_metrics(proyectos, labels, k)

        # SALVAGUARDA 2: separación de medias absoluta (real) entre los dos mercados extremos
        if len(grupos) >= 2:
            ms = sorted(g["media"] for g in grupos)
            sep_medias = (ms[-1] - ms[0]) / ms[-1] if ms[-1] else 0.0
        else:
            sep_medias = 0.0

        # SALVAGUARDA 3: separación espacial entre los dos clusters más distintos en media
        if len(grupos) >= 2:
            gs = sorted(grupos, key=lambda g: g["media"])
            sep_esp = _separacion_espacial(proyectos, labels, gs[0]["c"], gs[-1]["c"])
        else:
            sep_esp = 0.0

        es_robusto = (gap >= GAP_MIN and var_exp >= VAR_EXPLAINED_MIN and masa_ok
                      and sep_medias >= SEP_MEDIAS_MIN
                      and sep_esp >= (1.0 - SOLAPE_ESPACIAL_MAX))
        # puntaje: voto combinado de gap + varianza (la métrica más robusta)
        score = (gap / GAP_MIN) + (var_exp / VAR_EXPLAINED_MIN)
        candidatos.append({"k": k, "labels": labels, "gap": gap, "var": var_exp,
                           "masa_ok": masa_ok, "robusto": es_robusto, "score": score,
                           "grupos": grupos, "sep_medias": sep_medias, "sep_esp": sep_esp})

    robustos = [c for c in candidatos if c["robusto"]]
    if not robustos:
        out["motivo"] = "Sin partición robusta · un solo mercado (zona = isócrona)"
        return out

    # Elegir la partición robusta de mayor score
    mejor = max(robustos, key=lambda c: c["score"])
    out.update(barrera=True, k_elegido=mejor["k"], labels=mejor["labels"],
               grupos=mejor["grupos"], gap=round(mejor["gap"], 3),
               var_explained=round(mejor["var"], 3))

    # Asignar el PIN a su mercado: cluster cuyo CENTROIDE ESPACIAL está más cerca del pin
    k = mejor["k"]; labels = mejor["labels"]
    centroides = []
    for c in range(k):
        pts = [proyectos[i] for i in range(n) if labels[i] == c]
        if not pts:
            centroides.append((c, float("inf")))
            continue
        clng = statistics.mean(p["lng"] for p in pts)
        clat = statistics.mean(p["lat"] for p in pts)
        d = math.hypot(clng - pin_lng, clat - pin_lat)
        centroides.append((c, d))
    pin_cluster = min(centroides, key=lambda t: t[1])[0]
    out["pin_cluster"] = pin_cluster
    out["cluster_proyectos"] = [proyectos[i] for i in range(n) if labels[i] == pin_cluster]
    return out


# ──────────────────────── Percepción de valor + ajuste de zona ────────────────────────
# ──────────────────────── Valor de zona en CASCADA (regla de negocio) ────────────────────────
def _valor_zona_cascada(competidores: Dict, perception: Dict, demografia: Dict,
                        segments: List[Dict], modo: str = "vivienda_vertical") -> Dict[str, Any]:
    """SIEMPRE establece un VALOR DE ZONA (pm² y ticket de referencia), exista o no oferta.

    Regla de negocio del tablero (vertical y horizontal): el valor de la zona se determina en
    CASCADA, tomando la primera fuente disponible:
      1) Competidores DIRECTOS (mercado del pin) → pm² mediano observado.
      2) Competidores PRIMARIOS (isócrona 8 min) → pm² mediano.
      3) Competidores SECUNDARIOS (isócrona ampliada) → pm² mediano.
      4) PERCEPCIÓN DE VALOR por NSE/capacidad de pago de la demanda real de la zona
         (cuando no hay NINGÚN comparable: el valor lo define el poder adquisitivo + el producto
         que la metodología recomienda para ese NSE).
    Devuelve siempre {pm2, ticket_ref_M, fuente, rigidez, nota}. 'rigidez' indica si el valor
    es un precio de mercado establecido (hay comparables) o un valor indicativo (sin comparables,
    lo afinará la demanda y la calidad del proyecto). NUNCA devuelve None: la zona siempre tiene
    un valor de referencia. Integridad: cada nivel usa datos reales; si un nivel no tiene datos,
    baja al siguiente sin inventar.
    """
    def _mediana_pm2(items):
        vals = [c.get("pm2") for c in (items or []) if c.get("pm2") and c["pm2"] > 1000]
        return round(statistics.median(vals)) if vals else None

    # ── Niveles 1-3: comparables observados (directos → primarios → secundarios) ──
    cascada = [
        ("competidores_directos", competidores.get("directos")),
        ("competidores_primarios", competidores.get("primarios")),
        ("competidores_secundarios", competidores.get("secundarios")),
    ]
    for fuente, items in cascada:
        pm2 = _mediana_pm2(items)
        if pm2:
            # ticket de referencia = pm² × tamaño típico observado de ese set (o del producto recomendado)
            m2_ref = None
            m2_vals = [c.get("m2") for c in (items or []) if c.get("m2")]
            if m2_vals:
                m2_ref = round(statistics.median(m2_vals))
            return {
                "pm2": pm2,
                "m2_ref": m2_ref,
                "ticket_ref_M": round(pm2 * m2_ref / 1e6, 2) if m2_ref else None,
                "fuente": fuente,
                "n_comparables": len([c for c in items if c.get("pm2")]),
                # ZA-7 · descripción robusta del set que define el valor de zona
                "stats_robustas": _stats_robustas([c.get("pm2") for c in (items or [])
                                                   if c.get("pm2") and c["pm2"] > 1000]),
                "rigidez": "mercado_establecido",
                "_need_m2": m2_ref is None,   # señal: completar m2/ticket con el producto recomendado
                "nota": (f"Valor de zona anclado a {len([c for c in items if c.get('pm2')])} "
                         f"comparables ({fuente.replace('competidores_','')}). Precio de mercado observado."),
            }

    # ── Nivel 4: PERCEPCIÓN DE VALOR por NSE / capacidad de pago (sin comparables) ──
    # El valor lo define el poder adquisitivo real de la zona y el producto que la metodología
    # recomienda para ese NSE. Se toma del producto recomendado (que ya está anclado a la
    # capacidad de pago real de la demanda) o, en su defecto, del IXH de la zona.
    prod_ref = None
    for p in (segments or []):
        pass  # segments no traen pm2; se usa el producto derivado abajo vía perception
    # El producto recomendado (featured) ya viene anclado a capacidad de pago; usar su pm²/ticket.
    pm2_pv = None
    ticket_pv = None
    m2_pv = None
    return {
        "pm2": pm2_pv,                 # se completa en zona_procesar con el producto recomendado
        "m2_ref": m2_pv,
        "ticket_ref_M": ticket_pv,
        "fuente": "percepcion_valor_nse",
        "n_comparables": 0,
        "rigidez": "indicativo",
        "nota": ("La zona no tiene un valor de mercado preestablecido rígido (sin oferta "
                 "comparable). El valor se establece por percepción de valor del NSE y la "
                 "capacidad de pago de la demanda; la calidad del proyecto definirá el precio final."),
    }


def value_perception_adjust(resumen: List[Dict], base_ring: List[List[float]],
                            pin_lng: Optional[float] = None,
                            pin_lat: Optional[float] = None,
                            agebs_geo: Optional[List[Dict]] = None) -> Dict[str, Any]:
    """
    Capa de percepción de valor sobre la isócrona base (8 min) y ajuste de
    ZONA DE INFLUENCIA REAL mediante detección universal de mercados (clustering).

    La isócrona define quién llega; los datos de oferta (VVV) definen si dentro de
    esa isócrona hay UNO o VARIOS mercados. Si hay una barrera (física, de NSE o de
    percepción de valor), la zona de influencia real es el mercado MÁS COMPATIBLE
    CON EL PIN, no toda la isócrona. Ver `detect_markets` para el método universal.

    Devuelve `zona_poligono` (polígono GeoJSON · lista de anillos) para el mapa morado:
      • Sin barrera (un mercado)       → zona = isócrona de 8 min completa.
      • Con barrera (varios mercados)  → zona = hull del cluster del pin.
      • Sin datos suficientes          → isócrona (o None si no hay anillo).
    """
    proyectos = []
    for row in resumen:
        a = row.get("attributes", {})
        g = row.get("geometry", {})
        pm2 = _num(a.get("F__M2_PROM"))
        ticket = _num(a.get("F__UD_PROM"))   # ticket por unidad (señal de mayor peso)
        la = _num(g.get("y")) if g.get("y") is not None else _num(a.get("Y_coor"))
        ln = _num(g.get("x")) if g.get("x") is not None else _num(a.get("X_coor"))
        if pm2 is not None and pm2 > 1000 and la is not None and ln is not None:
            proyectos.append({"name": a.get("PROYECTO") or a.get("Nombre") or "N/D",
                              "pm2": pm2, "ticket": ticket, "lat": la, "lng": ln,
                              "nse_rank": None})  # nse_rank: gancho NSE (N/D hasta tener geometría)

    dentro = [p for p in proyectos if _point_in_ring(p["lng"], p["lat"], base_ring)]
    # NSE DEL ENTORNO por proyecto (señal de barrera de mercado · pendiente #2): nse_rank del
    # AGEB georreferenciado MÁS CERCANO (dato real del KMZ del DI). Sin geometría → None y
    # el peso de la señal se renormaliza en _build_feature_vectors (no se inventa).
    _geo_pts = [g for g in (agebs_geo or []) if g.get("nse_rank") is not None
                and g.get("lng") is not None and g.get("lat") is not None]
    if _geo_pts:
        for p in dentro:
            g = min(_geo_pts,
                    key=lambda q: (q["lng"] - p["lng"]) ** 2 + (q["lat"] - p["lat"]) ** 2)
            p["nse_rank"] = g["nse_rank"]
    precios = [p["pm2"] for p in dentro]

    out = {
        "n_total": len(dentro), "n_zona": len(dentro),
        "media": None, "sd": None, "cv": None,
        "barrera": False, "metodo": "isocrona", "cobertura_pct": 100,
        "motivo": None, "proyectos": dentro,
        "cluster_names": None,
        "zona_poligono": [base_ring] if base_ring else None,
        # VALOR PERCIBIDO: ¿la zona tiene un precio de mercado preestablecido (oferta comparable)?
        # Si no hay comparables, no se fuerza un precio rígido: la demanda y la calidad del
        # proyecto definirán producto y precio. El front muestra esta nota en ese caso.
        "sin_valor_percibido": len(dentro) == 0,
        "nota_valor": None,
        # Auditoría del clustering (integridad/transparencia)
        "mercados": {"detectado": False, "k": 1, "gap": None, "var_explained": None,
                     "pesos": None, "pin_cluster": None},
    }
    if len(dentro) == 0:
        out["nota_valor"] = ("La zona no tiene un valor preestablecido rígido: no hay oferta "
                             "comparable en el área de influencia. La demanda y la calidad del "
                             "proyecto establecerán la posibilidad de producto y precio.")
    if len(precios) < 3:
        out["motivo"] = "Muestra insuficiente para clúster · se mantiene la isócrona"
        return out

    media = statistics.mean(precios)
    sd = statistics.pstdev(precios)
    cv = sd / media if media > 0 else None
    out.update(media=media, sd=sd, cv=cv)
    # ZA-7/RES-2 · descripción ROBUSTA de la zona (mediana/MAD/percentiles/outliers);
    # media/sd/cv se conservan por compatibilidad (catálogo: no se cambia lo existente).
    out["stats_robustas"] = _stats_robustas(precios)

    # Pin: si no se pasó, usar el centroide de masa de los proyectos como referencia.
    p_lng = pin_lng if pin_lng is not None else statistics.mean(p["lng"] for p in dentro)
    p_lat = pin_lat if pin_lat is not None else statistics.mean(p["lat"] for p in dentro)

    # ── DETECCIÓN UNIVERSAL DE MERCADOS (clustering k-auto sobre señales ponderadas) ──
    mk = detect_markets(dentro, p_lng, p_lat)
    out["mercados"] = {
        "detectado": mk["barrera"], "k": mk["k_elegido"],
        "gap": mk["gap"], "var_explained": mk["var_explained"],
        "pesos": mk["pesos"], "pin_cluster": mk["pin_cluster"],
        "motivo": mk.get("motivo"),
    }

    if not mk["barrera"]:
        # La OFERTA no detecta barrera. Antes de devolver la isócrona completa, evaluar si
        # la DEMOGRAFÍA define una barrera de NSE (frontera espacial de niveles distintos).
        # Esto cubre zonas sin oferta vertical donde el cambio de NSE/IXH sí define el submercado.
        nse_zona = _zona_por_barrera_nse(agebs_geo or [], p_lng, p_lat, base_ring)
        if nse_zona["barrera"] and nse_zona.get("zona_poligono"):
            out["barrera"] = True
            out["metodo"] = "barrera_nse"
            out["zona_poligono"] = nse_zona["zona_poligono"]
            out["cobertura_pct"] = round(nse_zona["n_bloque"] / max(len(agebs_geo or [1]), 1) * 100)
            out["motivo"] = nse_zona["motivo"]
            out["mercados"]["barrera_nse"] = True
            return out
        # Sin barrera de oferta ni de NSE → zona = isócrona completa (regla confirmada).
        out["motivo"] = mk.get("motivo") or "Un solo mercado · zona = isócrona de 8 min"
        return out

    # Hay barrera: la zona de influencia real es el MERCADO DEL PIN.
    cluster = mk["cluster_proyectos"]
    if len(cluster) < CLUSTER_MASA_MIN:
        out["motivo"] = "Mercado del pin sin masa suficiente · se mantiene la isócrona"
        return out

    cprecios = [p["pm2"] for p in cluster]
    cmedia = statistics.mean(cprecios)
    out["barrera"] = True
    out["metodo"] = "mercado_del_pin"
    out["n_zona"] = len(cluster)
    out["proyectos"] = cluster
    out["media"] = cmedia
    out["sd"] = statistics.pstdev(cprecios)
    out["cv"] = (out["sd"] / cmedia) if cmedia else None
    out["stats_robustas"] = _stats_robustas(cprecios)   # robustas del MERCADO DEL PIN
    out["cobertura_pct"] = round(len(cluster) / max(len(dentro), 1) * 100)
    out["cluster_names"] = set(p["name"] for p in cluster if p.get("name") and p["name"] != "N/D")
    out["motivo"] = (f"{mk['k_elegido']} mercados detectados (gap={mk['gap']}, "
                     f"var={mk['var_explained']}) · zona = mercado del pin")

    # Zona de influencia real = hull del mercado del pin. Si no se forma (<3 puntos),
    # se mantiene la isócrona (integridad: no se inventa polígono).
    hull = _convex_hull([[p["lng"], p["lat"]] for p in cluster])
    if hull:
        out["zona_poligono"] = [hull]
    return out


# ──────────────────────── Clasificación de competidores en 3 zonas ────────────────────────
def clasificar_competidores(resumen: List[Dict],
                            base_ring: List[List[float]],
                            outer_ring: List[List[float]],
                            zona_poligono: Optional[List[List[List[float]]]]) -> Dict[str, Any]:
    """Clasifica cada proyecto de oferta (VVV) en uno de 3 sets de competidores:

      • DIRECTO    → dentro de la zona de influencia real (morado · mercado del pin).
                     Set de oferta competidora PRIMARIA a comparar (mismo mercado de valor).
      • PRIMARIO   → dentro de la isócrona de 8 min (azul · zona primaria), no directo.
      • SECUNDARIO → entre la isócrona azul y la mayor (verde · zona secundaria).

    Las zonas azul + verde son las GENERADORAS DE DEMANDA. El morado es independiente
    (puede caer en azul, cruzar a verde, o parcialmente fuera): se evalúa por geometría real.
    Cada proyecto devuelto lleva su etiqueta `set_competidor`.
    """
    morado_ring = zona_poligono[0] if (zona_poligono and zona_poligono[0]) else None

    def coords(row):
        a = row.get("attributes", {}); g = row.get("geometry", {})
        lng = _num(g.get("x")) if g.get("x") is not None else _num(a.get("X_coor"))
        lat = _num(g.get("y")) if g.get("y") is not None else _num(a.get("Y_coor"))
        return lng, lat

    directos, primarios, secundarios = [], [], []
    for row in resumen:
        lng, lat = coords(row)
        if lng is None or lat is None:
            continue
        a = row.get("attributes", {})
        item = {"name": a.get("PROYECTO") or a.get("Nombre") or "N/D",
                "pm2": _num(a.get("F__M2_PROM")), "ticket": _num(a.get("F__UD_PROM")),
                "lat": lat, "lng": lng}
        # DIRECTO: dentro del polígono morado (si existe)
        if morado_ring and _point_in_ring(lng, lat, morado_ring):
            item["set_competidor"] = "directo"
            directos.append(item)
        # PRIMARIO: dentro de la isócrona azul (8 min)
        elif base_ring and _point_in_ring(lng, lat, base_ring):
            item["set_competidor"] = "primario"
            primarios.append(item)
        # SECUNDARIO: dentro de la isócrona mayor (verde) pero fuera de la azul
        elif outer_ring and _point_in_ring(lng, lat, outer_ring):
            item["set_competidor"] = "secundario"
            secundarios.append(item)
        # (si no cae en ninguna, no se incluye — fuera del universo de análisis)

    return {
        "directos": directos, "primarios": primarios, "secundarios": secundarios,
        "n_directos": len(directos), "n_primarios": len(primarios), "n_secundarios": len(secundarios),
    }


# ──────────────────────── Demografía base (de AGEBs) ────────────────────────
# Mapeo de NSE INEGI a las clases del tablero
NSE_MAP = {"A": "A", "B": "B", "C+": "C+", "C": "C", "D+": "D+", "D": "D", "E": "E"}

# Rangos de ingreso MENSUAL por hogar (estándar AMAI/DIGO · MXN). Universal para todo México.
NSE_INCOME_BANDS = [
    ("A",  200000, None),
    ("B",   90000, 199999),
    ("C+",  40000, 89999),
    ("C",   17000, 39999),
    ("D+",  10000, 16999),
    ("D",    4000, 9999),
    ("E",       0, 3999),
]

# TCA (tasa de crecimiento anual de hogares, %) canónica por NSE — FUENTE ÚNICA del backend.
# La usan derive_nse_dim y derive_segments; el front la LEE del payload (nunca la recalcula).
NSE_TCA = {"A": 1.12, "B": 1.0, "C+": 1.39, "C": 1.03, "D+": 0.74, "D": 0.37, "E": 0}


def _ageb_ixh_mensual(r: Dict) -> Optional[float]:
    """Ingreso por hogar MENSUAL del AGEB.

    IMPORTANTE: el campo IXH ya viene MENSUAL por hogar (verificado contra datos reales:
    p. ej. NSE C+ → IXH ≈ $89,745/mes; NSE B → ≈ $401,847/mes; coherente con AMAI). NO se
    divide entre 12. El fallback `Ingresos totales / hogares` también es mensual por hogar.
    """
    ixh = _num(r.get("IXH"))
    if ixh is not None and ixh > 0:
        return ixh
    ing = _num(r.get("Ingresos totales 2026"))
    hog = _num(r.get("Hogares totales 2026")) or _num(r.get("Hogares totales 2020"))
    if ing is not None and hog and hog > 0:
        return ing / hog
    return None


def _ageb_vcasa(r: Dict) -> Optional[float]:
    """Valor de vivienda del AGEB (campo 'V CASAS')."""
    return _num(r.get("V CASAS")) or _num(r.get("V_CASAS")) or _num(r.get("V_CASA"))


def _ageb_prop_casa(r: Dict) -> float:
    """Proporción de viviendas tipo CASA (horizontal) sobre el total de viviendas del AGEB.

    Para vivienda HORIZONTAL: la demanda genérica de vivienda del AGEB se pondera por esta
    proporción, de modo que en zonas de casas la demanda horizontal es alta y en zonas de
    departamentos es baja. Universal: usa el conteo real de tipo de vivienda de la base.
    Integridad: si no hay dato de tipo de vivienda, devuelve 1.0 (no se descarta demanda;
    se asume que aplica al modo, evitando perder demanda por falta de desglose).
    """
    casa = _num(r.get("Casa"))
    depto = _num(r.get("Departamento o edificio"))
    vecindad = _num(r.get("Vivienda en vecidad o cuartería ")) or _num(r.get("Vivienda en vecidad o cuartería"))
    otro = _num(r.get("Otro tipo de vivienda"))
    if casa is None:
        return 1.0   # sin desglose → no se pondera (integridad: no perder demanda)
    total = sum(v for v in (casa, depto, vecindad, otro) if v is not None)
    if not total:
        return 1.0
    return max(0.0, min(1.0, casa / total))


def _ageb_prop_vertical(r: Dict) -> float:
    """Proporción de viviendas tipo DEPARTAMENTO (vertical) sobre el total del AGEB.
    Complemento conceptual de _ageb_prop_casa para el modo vertical. Hoy el modo vertical
    usa la demanda genérica (no se pondera) para no alterar el comportamiento validado;
    este helper queda disponible para refinar el modo vertical en el futuro."""
    casa = _num(r.get("Casa"))
    depto = _num(r.get("Departamento o edificio"))
    vecindad = _num(r.get("Vivienda en vecidad o cuartería ")) or _num(r.get("Vivienda en vecidad o cuartería"))
    otro = _num(r.get("Otro tipo de vivienda"))
    if depto is None:
        return 1.0
    total = sum(v for v in (casa, depto, vecindad, otro) if v is not None)
    if not total:
        return 1.0
    return max(0.0, min(1.0, depto / total))


def classify_nse_by_income(ixh_mensual: Optional[float]) -> Optional[str]:
    """Clasifica el NSE de un AGEB por su ingreso mensual real (AMAI). Universal."""
    if ixh_mensual is None:
        return None
    for nse, lo, hi in NSE_INCOME_BANDS:
        if ixh_mensual >= lo and (hi is None or ixh_mensual <= hi):
            return nse
    return "E"


def ageb_nse(r: Dict) -> Optional[str]:
    """
    NSE del AGEB tomado DIRECTAMENTE de la capa de NSE de ArcGIS (base Prosperia):
    campo 'NSE PER' (NSE por persona) con respaldo 'XI_Nivel socioeconómico por ingreso'.
    Estas capas ya vienen clasificadas por la fuente; no se recalcula por bandas de ingreso.
    Universal para todo México: la clasificación la provee ArcGIS, no un heurístico local.
    """
    cls = r.get("NSE PER") or r.get("XI_Nivel socioeconómico por ingreso")
    if cls:
        cls = str(cls).strip()
        # Normalizar variantes de etiqueta a las claves canónicas
        if cls in NSE_MAP:
            return cls
        # variantes comunes: "C plus" → "C+", "D plus" → "D+"
        cls2 = cls.replace(" plus", "+").replace("Plus", "+").replace(" ", "")
        if cls2 in NSE_MAP:
            return cls2
    return None


def _nse_percepcion_valor(ft: List[Dict], pm2_min: int = 20000) -> Optional[str]:
    """
    PERCEPCIÓN DE VALOR de la zona = NSE del PRECIO MEDIANO de la oferta vertical real (VVV).
    Regla DIGO: la isócrona define la ZONA, pero la percepción de valor define el VALOR del
    producto. El precio mediano de lo que realmente se ofrece/vende indica para qué percepción
    se construye, independiente de la demografía de la isócrona. Devuelve la clave NSE o None.
    """
    if not ft:
        return None
    precios_m = []
    for t in ft:
        a = t.get("attributes", {})
        precio = _price(a.get("F____UNIDAD"))
        pm2v = _pm2(a.get("F___M2"))
        if precio and pm2v and pm2v >= pm2_min:   # solo oferta vertical plausible
            precios_m.append(precio / 1e6)
    if not precios_m:
        return None
    mediano_m = statistics.median(precios_m)
    bands = [("A", 6.8, None), ("B", 3.05, 6.8), ("C+", 1.35, 3.05), ("C", 0.577, 1.35),
             ("D+", 0.349, 0.577), ("D", 0.2, 0.349), ("E", 0, 0.2)]
    for nse, lo, hi in bands:
        if mediano_m >= lo and (hi is None or mediano_m < hi):
            return nse
    return None


def derive_demografia(agebs: List[Dict], ft: Optional[List[Dict]] = None,
                      agebs_geo: Optional[List[Dict]] = None,
                      pin_lng: Optional[float] = None, pin_lat: Optional[float] = None) -> Dict[str, Any]:
    """Agrega los AGEBs a KPIs demográficos de zona. Ausencia => None.
    Si se pasa `ft` (oferta vertical VVV), el NSE dominante mostrado refleja la PERCEPCIÓN DE
    VALOR (el NSE para el que realmente se construye), que predomina sobre la demografía cruda
    de la isócrona cuando hay oferta vertical observada."""
    if not agebs:
        return {
            "population": None, "households": None, "tca": None,
            "personas_hogar": None, "ingreso_hogar": None, "ingreso_total_anual": None,
            "nse": {}, "nse_dominante": None, "tenencia": {}, "hog_renta": None,
            "edad_grupos": None,
        }

    pob = _sum(agebs, "Población total 2026") or _sum(agebs, "POB1")
    hog = _sum(agebs, "Hogares totales 2026") or _sum(agebs, "Hogares totales 2020")
    ing_total = _sum(agebs, "Ingresos totales 2026")
    personas_hogar = _wavg(agebs, "Personas por hogar 2026", "Hogares totales 2026")

    # TCA: tasa de crecimiento anual ponderada por población (decimal)
    tca_pct = _wavg(agebs, "Tasa de crecimiento anual", "POB1")
    tca = round(tca_pct / 100.0, 4) if tca_pct is not None else None

    # Ingreso por hogar MENSUAL ponderado (IXH ya es mensual · ver _ageb_ixh_mensual)
    ixh_vals = []
    for r in agebs:
        ixh_m = _ageb_ixh_mensual(r)
        h = _num(r.get("Hogares totales 2026")) or _num(r.get("Hogares totales 2020"))
        if ixh_m is not None and h:
            ixh_vals.append((ixh_m, h))
    if ixh_vals:
        num = sum(v * w for v, w in ixh_vals); den = sum(w for _, w in ixh_vals)
        ingreso_hogar = num / den if den else None
    else:
        ingreso_hogar = None

    # NSE por hogares, clasificado por INGRESO REAL del AGEB (universal AMAI)
    nse_hog: Dict[str, float] = {}
    nse_ing: Dict[str, List[float]] = {}
    for r in agebs:
        key = ageb_nse(r)
        h = _num(r.get("Hogares totales 2026")) or _num(r.get("Hogares totales 2020"))
        if key in NSE_MAP and h is not None:
            nse_hog[key] = nse_hog.get(key, 0.0) + h
            ixh_m = _ageb_ixh_mensual(r)
            if ixh_m is not None:
                # H10: guardar (ingreso, hogares) para promediar PONDERADO por hogares
                nse_ing.setdefault(key, []).append((ixh_m, h))

    def _wmean_pairs(pairs):
        """Promedio PONDERADO por hogares de pares (valor, hogares). H10."""
        num = sum(v * (w or 0) for v, w in pairs)
        den = sum((w or 0) for _, w in pairs)
        return (num / den) if den else (statistics.mean([v for v, _ in pairs]) if pairs else None)

    nse = {}
    total_hog = sum(nse_hog.values()) if nse_hog else 0
    # El template usa claves Cm (=C+) y Dm (=D+)
    key_map = {"C+": "Cm", "D+": "Dm"}
    for k, h in nse_hog.items():
        ingreso = round(_wmean_pairs(nse_ing[k])) if nse_ing.get(k) else None
        tkey = key_map.get(k, k)
        nse[tkey] = {"hog": round(h), "pct": round(h / total_hog * 100, 1) if total_hog else None,
                     "ingreso": ingreso}
    # Garantizar que todas las claves que el template consulta existan (hog=0 si no hay)
    for tkey in ["A", "B", "Cm", "C", "Dm", "D", "E"]:
        nse.setdefault(tkey, {"hog": 0, "pct": 0, "ingreso": None})
    # NSE dominante por masa de hogares PONDERADA por proximidad al pin (regla: el dominante es
    # el más accesible; la accesibilidad se cuantifica con distancia real). Consistente con
    # _nse_dominante_agebs. Sin geometría/pin → masa simple (factor 1, integridad).
    if nse_hog:
        _prox = _factor_proximidad_por_nse(agebs_geo, pin_lng, pin_lat)
        _masa_pond = {k: v * _prox.get(k, 1.0) for k, v in nse_hog.items()}
        nse_dom_key = max(_masa_pond, key=_masa_pond.get)
    else:
        nse_dom_key = None
    # PERCEPCIÓN DE VALOR PREDOMINA: si la oferta vertical observada indica un NSE superior al
    # demográfico, el NSE dominante mostrado se eleva a esa percepción (la zona se construye
    # para ese nivel). La isócrona define la zona; la percepción de valor define el producto.
    NSE_ORDEN_PV = ["A", "B", "C+", "C", "D+", "D", "E"]
    rank_pv = {n: i for i, n in enumerate(NSE_ORDEN_PV)}
    nse_percepcion = _nse_percepcion_valor(ft) if ft else None
    percepcion_aplicada = False
    if nse_percepcion and nse_dom_key and rank_pv.get(nse_percepcion, 99) < rank_pv.get(nse_dom_key, 99):
        nse_dom_key = nse_percepcion
        percepcion_aplicada = True
    # Etiqueta descriptiva (el front muestra el string completo y hace split(' ')[0] para la letra)
    NSE_DESC = {
        "A": "NSE alto · residencial premium", "B": "NSE medio-alto",
        "C+": "NSE medio-alto · aspiracional", "C": "NSE medio",
        "D+": "NSE medio-bajo", "D": "NSE bajo", "E": "NSE de subsistencia",
    }
    if percepcion_aplicada:
        # El % de hogares ya no aplica (es percepción de valor de oferta, no demografía)
        nse_dominante = f"{nse_dom_key} · {NSE_DESC.get(nse_dom_key, '')} · por percepción de valor"
    else:
        dom_pct = round(nse_hog[nse_dom_key] / total_hog * 100, 1) if (nse_dom_key and total_hog) else None
        nse_dominante = (f"{nse_dom_key} · {NSE_DESC.get(nse_dom_key, '')}"
                         + (f" ({dom_pct}%)" if dom_pct else "")) if nse_dom_key else None

    # INGRESO DEL HOGAR coherente con el NSE DOMINANTE (dato REAL de la base ArcGIS, no inventado).
    # El promedio ponderado de toda la isócrona diluye el ingreso del NSE alto (mezcla AGEBs A
    # con muchos D+). Cuando hay un NSE dominante (demográfico o por percepción de valor), el
    # ingreso mostrado = ingreso REAL de los AGEBs de ESE NSE (campo IXH de la base). Si ese NSE
    # no tiene AGEBs en la zona (p. ej. percepción A sin AGEB A), se cae al promedio ponderado.
    if nse_dom_key and nse_ing.get(nse_dom_key):
        ingreso_hogar = _wmean_pairs(nse_ing[nse_dom_key])   # H10: ponderado por hogares
    # else: se conserva el promedio ponderado ya calculado arriba (último recurso, dato de base)

    # Tenencia (C81-84) en %
    propia = _sum(agebs, "Propia"); alquilada = _sum(agebs, "Alquilada")
    prestada = _sum(agebs, "Prestada"); otra = _sum(agebs, "Otra situación")
    ten_total = sum(v for v in [propia, alquilada, prestada, otra] if v is not None)
    tenencia = {}
    if ten_total > 0:
        tenencia = {
            "propia": round((propia or 0) / ten_total * 100, 1),
            "alquilada": round((alquilada or 0) / ten_total * 100, 1),
            "prestada": round((prestada or 0) / ten_total * 100, 1),
            "otra": round((otra or 0) / ten_total * 100, 1),
        }
    hog_renta = round(alquilada) if alquilada is not None else None

    # Edad: agrupación a las 6 bandas del tablero (niños, adolescentes, jóvenes, jov_adultos, consolidados, maduros)
    edad_grupos = _derive_edad_grupos(agebs)

    # UBICACIÓN GEOGRÁFICA real de la zona (de la base, no inventada). El municipio/estado
    # dominante = el de mayor masa de hogares en la zona de influencia. La zona puede abarcar
    # varios municipios; se reporta el dominante y la lista completa para transparencia.
    ubic = _derive_ubicacion(agebs)

    return {
        "population": round(pob) if pob else None,
        "households": round(hog) if hog else None,
        "tca": tca,
        "personas_hogar": round(personas_hogar, 2) if personas_hogar else None,
        "ingreso_hogar": round(ingreso_hogar) if ingreso_hogar else None,
        "ingreso_total_anual": round(ing_total) if ing_total else None,
        "nse": nse, "nse_dominante": nse_dominante,
        "tenencia": tenencia, "hog_renta": hog_renta,
        "edad_grupos": edad_grupos,
        # — Ubicación geográfica real —
        "municipio": ubic.get("municipio"),
        "estado": ubic.get("estado"),
        "pais": ubic.get("pais"),
        "municipios_todos": ubic.get("municipios_todos"),
    }


def _derive_ubicacion(agebs: List[Dict]) -> Dict[str, Any]:
    """Extrae municipio/estado dominantes de la zona desde los AGEBs (campos reales 'Municipio'
    y 'Estado' de la base ArcGIS). Dominante = mayor masa de hogares. País = México (la base
    es nacional MX). Integridad: si no hay dato, devuelve None (no se inventa ubicación)."""
    if not agebs:
        return {"municipio": None, "estado": None, "pais": None, "municipios_todos": []}
    muni_masa: Dict[str, float] = {}
    edo_masa: Dict[str, float] = {}
    for r in agebs:
        h = _num(r.get("Hogares totales 2026")) or _num(r.get("Hogares totales 2020")) or 1
        muni = r.get("Municipio")
        edo = r.get("Estado")
        if muni and str(muni).strip():
            m = str(muni).strip()
            muni_masa[m] = muni_masa.get(m, 0) + h
        if edo and str(edo).strip():
            e = str(edo).strip()
            edo_masa[e] = edo_masa.get(e, 0) + h
    muni_dom = max(muni_masa, key=muni_masa.get) if muni_masa else None
    edo_dom = max(edo_masa, key=edo_masa.get) if edo_masa else None
    municipios = sorted(muni_masa, key=muni_masa.get, reverse=True)
    return {
        "municipio": muni_dom,
        "estado": edo_dom,
        "pais": "México" if (muni_dom or edo_dom) else None,
        "municipios_todos": municipios,
    }


def _derive_edad_grupos(agebs: List[Dict]) -> Optional[List[int]]:
    """6 bandas: niños(0-9), adolescentes(10-15), jóvenes(16-19), jov_adultos(20-29), consolidados(30-54), maduros(55+)."""
    bands = {
        "ninos": ["0 a 4", "5 a 9"],
        "adolescentes": ["10 a 14"],
        "jovenes": ["15 a 19"],
        "jov_adultos": ["20 a 24", "25 a 29"],
        "consolidados": ["30 a 34", "35 a 39", "40 a 44", "45 a 49", "50 a 54"],
        "maduros": ["55 a 59", "60 a 64", "65 a 69", "70 a 74", "75 y Más"],
    }
    out = []
    any_data = False
    for band, cols in bands.items():
        s = 0.0
        present = False
        for c in cols:
            v = _sum(agebs, c)
            if v is not None:
                s += v; present = True
        if present:
            any_data = True
        out.append(round(s))
    return out if any_data else None


# ──────────────────────── NSE_DIM (para DIM_DATA) ────────────────────────
def derive_nse_dim(agebs: List[Dict]) -> List[Dict[str, Any]]:
    """Construye nse_dim[] con rangos de ingreso/vivienda y población por grupo de edad."""
    if not agebs:
        return []
    # Rangos de ingreso por NSE (estándar AMAI/DIGO, preservados del template)
    nse_ranges = {
        "A":  {"ing_min": 200000, "ing_max": None, "viv_min": 6800000, "viv_max": None, "tca": NSE_TCA["A"]},
        "B":  {"ing_min": 90000,  "ing_max": 199999, "viv_min": 3050000, "viv_max": 6799999, "tca": NSE_TCA["B"]},
        "C+": {"ing_min": 40000,  "ing_max": 89999,  "viv_min": 1350000, "viv_max": 3049999, "tca": NSE_TCA["C+"]},
        "C":  {"ing_min": 17000,  "ing_max": 39999,  "viv_min": 577000,  "viv_max": 1349999, "tca": NSE_TCA["C"]},
        "D+": {"ing_min": 10000,  "ing_max": 16999,  "viv_min": 349000,  "viv_max": 576999,  "tca": NSE_TCA["D+"]},
        "D":  {"ing_min": 4000,   "ing_max": 9999,   "viv_min": 200000,  "viv_max": 348999,  "tca": NSE_TCA["D"]},
        "E":  {"ing_min": 0,      "ing_max": 3999,   "viv_min": 0,       "viv_max": 199999,  "tca": NSE_TCA["E"]},
    }
    agg: Dict[str, Dict] = {}
    for r in agebs:
        k = ageb_nse(r)   # clasificación por ingreso real (universal)
        if k not in NSE_MAP:
            continue
        d = agg.setdefault(k, {"poblacion": 0.0, "hogares": 0.0,
                               "ninos": 0.0, "adolescentes": 0.0, "jovenes": 0.0,
                               "jov_adultos": 0.0, "consolidados": 0.0, "empty_nest": 0.0,
                               "ixh_w": 0.0, "viv_w": 0.0})
        h = _num(r.get("Hogares totales 2026")) or _num(r.get("Hogares totales 2020")) or 0
        d["poblacion"] += _num(r.get("Población total 2026")) or _num(r.get("POB1")) or 0
        d["hogares"] += h
        d["ninos"] += _num(r.get("Niños")) or 0
        d["adolescentes"] += _num(r.get("Adolescentes")) or 0
        d["jovenes"] += _num(r.get("Jovenes")) or 0
        d["jov_adultos"] += _num(r.get("Jovenes_adultos")) or 0
        d["consolidados"] += _num(r.get("Consolidados")) or 0
        d["empty_nest"] += _num(r.get("Nesters")) or 0
        # IXH mensual ponderado por hogares
        ixh_m = _ageb_ixh_mensual(r)
        if ixh_m is not None and h:
            d["ixh_w"] += ixh_m * h
        # Valor de vivienda (V CASAS) ponderado por hogares
        vcasa = _ageb_vcasa(r)
        if vcasa is not None and h:
            d["viv_w"] += vcasa * h

    out = []
    order = ["A", "B", "C+", "C", "D+", "D", "E"]
    for k in order:
        rng = nse_ranges[k]
        d = agg.get(k, {"poblacion": 0, "hogares": 0, "ninos": 0, "adolescentes": 0,
                        "jovenes": 0, "jov_adultos": 0, "consolidados": 0, "empty_nest": 0,
                        "ixh_w": 0, "viv_w": 0})
        hog = d["hogares"]
        # ixh_nse: ingreso por hogar MENSUAL real del NSE; si no hay dato, punto medio del rango
        if hog and d["ixh_w"] > 0:
            ixh = round(d["ixh_w"] / hog)
        else:
            ixh = round((rng["ing_min"] + (rng["ing_max"] or rng["ing_min"] * 1.3)) / 2)
        # viv_nse: valor de vivienda real ponderado; si no hay dato, punto medio del rango
        if hog and d["viv_w"] > 0:
            viv = round(d["viv_w"] / hog)
        else:
            viv = round((rng["viv_min"] + (rng["viv_max"] or rng["viv_min"] * 1.3)) / 2)
        out.append({
            "NSE": k,
            "ing_min": rng["ing_min"], "ing_max": rng["ing_max"] or 0,
            "viv_min": rng["viv_min"], "viv_max": rng["viv_max"] or 0,
            "tca": rng["tca"],
            "ixh_nse": ixh, "viv_nse": viv,
            "poblacion": round(d["poblacion"]),
            "hogares": round(d["hogares"]),
            "ninos": round(d["ninos"], 1), "adolescentes": round(d["adolescentes"], 1),
            "jovenes": round(d["jovenes"], 1), "jov_adultos": round(d["jov_adultos"], 1),
            "consolidados": round(d["consolidados"], 1), "empty_nest": round(d["empty_nest"], 1),
        })
    return out


# ──────────────────────── Segmentos de demanda (Checkpoint A) ────────────────────────
# NOTA: la tabla canónica de buckets de precio vive DENTRO de derive_segments (BUCKETS).
# Aquí existía una segunda tabla PRICE_BUCKETS muerta (nunca usada) que se eliminó para
# evitar ediciones en la tabla equivocada (regla: una sola fuente por constante).


def _nse_dominante_agebs(agebs: List[Dict], agebs_geo: Optional[List[Dict]] = None,
                         pin_lng: Optional[float] = None, pin_lat: Optional[float] = None):
    """
    BARRERA DE NSE (regla de zona de influencia · universal):
    Dentro de la isócrona puede haber mezcla de submercados de distinto NSE. El análisis
    de demanda NO se hace sobre el promedio de la zona, sino sobre el NSE DOMINANTE:
    el de mayor representatividad (masa de hogares) dentro de la isócrona, que por la
    propia accesibilidad de la isócrona es también el más accesible desde el predio.

    PONDERACIÓN POR PROXIMIDAD (regla confirmada): "el dominante es el más accesible" se
    cuantifica con la DISTANCIA REAL al pin. La masa de hogares de cada NSE se multiplica por
    un factor de cercanía al predio: masa_ponderada[nse] = masa[nse] × proximidad[nse], donde
    proximidad[nse] = promedio de 1/(1+dist_km) de las AGEB georreferenciadas de ese NSE. Así,
    si lo cercano tiene mayor capacidad de pago, el dominante (y el precio que se ancla a él)
    sube; si lo cercano paga menos, baja. En zonas amplias (horizontal, isócronas grandes) esto
    evita que AGEB lejanas de otro NSE dominen el análisis. Solo cambia el VALOR del dominante;
    NO altera la regla (sigue siendo masa + accesibilidad). Sin geometría → masa simple (integridad).

    Devuelve (agebs_dominante, nse_dom, nse_superior, share_dom):
      • agebs_dominante: solo las AGEB del NSE dominante (sobre las que se calcula demanda).
      • nse_dom: clave del NSE dominante.
      • nse_superior: el NSE inmediatamente superior presente en la zona (define el TECHO
        de inducción de demanda hacia arriba — siempre por debajo del piso de esa zona).
      • share_dom: fracción de hogares que representa el dominante (para reportar la barrera).
    """
    if not agebs:
        return agebs, None, None, 1.0
    # Orden de NSE de mayor a menor poder adquisitivo
    NSE_ORDEN = ["A", "B", "C+", "C", "D+", "D", "E"]
    rank = {n: i for i, n in enumerate(NSE_ORDEN)}
    # Masa de hogares por NSE (representatividad real en la isócrona)
    masa = {}
    for r in agebs:
        nse = ageb_nse(r)
        if not nse:
            continue
        hog = (_num(r.get("Hogares totales 2026")) or _num(r.get("Hogares totales 2020"))
               or _num(r.get("Total de hogares")) or 1)
        masa[nse] = masa.get(nse, 0) + hog
    if not masa:
        return agebs, None, None, 1.0
    total_hog = sum(masa.values())
    # Factor de proximidad por NSE (si hay geometría georreferenciada y pin). Pondera la masa
    # por cercanía al predio; el más próximo pesa más. Sin datos → factor 1 (masa simple).
    proximidad = _factor_proximidad_por_nse(agebs_geo, pin_lng, pin_lat)
    masa_pond = {n: masa[n] * proximidad.get(n, 1.0) for n in masa}
    # El NSE dominante se elige sobre la masa PONDERADA por proximidad (accesibilidad real).
    nse_dom = max(masa_pond, key=lambda n: masa_pond[n])
    # share_dom se reporta sobre la masa real (hogares), no la ponderada (transparencia del dato).
    share_dom = masa[nse_dom] / total_hog if total_hog else 1.0
    # NSE inmediatamente superior PRESENTE en la zona (techo de inducción)
    nse_superior = None
    for n in NSE_ORDEN:
        if n in masa and rank.get(n, 99) < rank.get(nse_dom, 0):
            nse_superior = n  # el último (más cercano) por encima del dominante
    # AGEB del submercado dominante
    agebs_dom = [r for r in agebs if ageb_nse(r) == nse_dom]
    return agebs_dom, nse_dom, nse_superior, share_dom


def _factor_proximidad_por_nse(agebs_geo: Optional[List[Dict]], pin_lng: Optional[float],
                               pin_lat: Optional[float]) -> Dict[str, float]:
    """Factor de cercanía al pin por NSE, a partir de las AGEB georreferenciadas (agebs_geo trae
    nse_txt + centroide lng/lat). Para cada NSE: promedio de 1/(1+dist_km) de sus AGEB. NSE con
    AGEB cercanas → factor alto (cerca de 1); con AGEB lejanas → factor bajo. Devuelve {} si no
    hay geometría o pin (el llamador usa entonces masa simple, sin inventar posiciones)."""
    if not agebs_geo or pin_lng is None or pin_lat is None:
        return {}
    acc: Dict[str, List[float]] = {}
    for g in agebs_geo:
        nse = g.get("nse_txt")
        la = _num(g.get("lat")); ln = _num(g.get("lng"))
        if not nse or la is None or ln is None:
            continue
        # Distancia aproximada en km (equirectangular; suficiente a escala de isócrona urbana).
        dlat = (la - pin_lat) * 111.0
        dlng = (ln - pin_lng) * 111.0 * math.cos(math.radians(pin_lat))
        dist_km = math.sqrt(dlat * dlat + dlng * dlng)
        acc.setdefault(nse, []).append(1.0 / (1.0 + dist_km))
    return {n: (sum(v) / len(v)) for n, v in acc.items() if v}


def _nse_barrier_info(agebs: List[Dict], agebs_geo: Optional[List[Dict]] = None) -> Dict[str, Any]:
    """Reporte legible de la barrera de NSE aplicada a la zona de influencia.

    Combina dos señales:
      • COMPOSICIÓN (del XLSX): qué NSEs hay y en qué proporción de hogares.
      • SEPARACIÓN ESPACIAL (del KMZ georreferenciado, agebs_geo): si los AGEBs de NSE
        distinto están geográficamente separados (frontera real) o entremezclados.
    La barrera de NSE es real solo si hay mezcla relevante Y separación espacial.
    """
    agebs_geo = agebs_geo or []
    agebs_dom, nse_dom, nse_superior, share_dom = _nse_dominante_agebs(agebs)
    # Composición NSE completa de la isócrona (informativa)
    masa = {}
    for r in agebs:
        nse = ageb_nse(r)
        if not nse:
            continue
        hog = (_num(r.get("Hogares totales 2026")) or _num(r.get("Hogares totales 2020"))
               or _num(r.get("Total de hogares")) or 1)
        masa[nse] = masa.get(nse, 0) + hog
    total = sum(masa.values()) or 1
    composicion = {k: round(v / total * 100, 1) for k, v in sorted(masa.items(), key=lambda kv: -kv[1])}
    mezcla_relevante = len(masa) > 1 and share_dom < 0.85  # submercados de NSE distintos

    # SEPARACIÓN ESPACIAL (requiere geometría)
    espacial = _nse_separacion_espacial(agebs_geo)
    barrera = mezcla_relevante and espacial.get("separados", False)

    return {
        "nse_dominante": nse_dom,
        "nse_superior": nse_superior,
        "share_dominante_pct": round(share_dom * 100, 1),
        "agebs_dominante": len(agebs_dom),
        "agebs_total": len(agebs),
        "agebs_geo_total": len(agebs_geo),
        "composicion_nse": composicion,
        "mezcla_relevante": mezcla_relevante,
        "separacion_espacial": espacial,
        "barrera": barrera,
        "metodo": ("nse_espacial" if barrera else
                   "nse_mezclado_sin_frontera" if mezcla_relevante else
                   "isocrona_uniforme"),
        "nota": (
            f"Barrera de NSE: la zona combina niveles {', '.join(composicion.keys())} "
            f"espacialmente separados. El mercado del predio corresponde a su zona de NSE, "
            f"no a toda la isócrona."
            if barrera else
            f"Mezcla de NSE ({', '.join(composicion.keys())}) entremezclada sin frontera "
            f"geográfica: se mantiene la isócrona como zona."
            if mezcla_relevante else
            f"Zona homogénea en NSE {nse_dom}" if nse_dom else "Sin datos de NSE"),
    }


def _zona_por_barrera_nse(agebs_geo: List[Dict], pin_lng: float, pin_lat: float,
                          base_ring: List[List[float]]) -> Dict[str, Any]:
    """Determina si la DEMOGRAFÍA define una barrera de NSE que recorta la zona.

    Cuando la oferta no detecta barrera (o no hay oferta), la demografía puede definirla:
    si dentro de la isócrona hay una FRONTERA ESPACIAL de NSE (bloques de NSE distinto
    geográficamente separados), la zona de influencia real del predio es el BLOQUE DE NSE
    donde cae el pin, no toda la isócrona. Esto resuelve el caso de zonas sin oferta
    vertical donde el cambio drástico de NSE/IXH sí define un submercado.

    Devuelve {barrera, zona_poligono, n_bloque, nse_bloque, motivo}. Integridad: si no hay
    frontera o no se puede formar el hull, barrera=False (no se inventa recorte).
    """
    res = {"barrera": False, "zona_poligono": None, "motivo": None}
    sep = _nse_separacion_espacial(agebs_geo)
    if not sep.get("separados"):
        res["motivo"] = sep.get("motivo")
        return res

    pts = [g for g in agebs_geo if g.get("nse_rank") is not None and g.get("lng") and g.get("lat")]
    rank_med = sep.get("rank_corte")
    if rank_med is None:
        return res
    # Bloques: alto valor (rank ≤ mediana) vs bajo valor (rank > mediana)
    alto = [g for g in pts if g["nse_rank"] <= rank_med]
    bajo = [g for g in pts if g["nse_rank"] > rank_med]
    if len(alto) < 2 or len(bajo) < 2:
        alto = [g for g in pts if g["nse_rank"] < rank_med]
        bajo = [g for g in pts if g["nse_rank"] >= rank_med]
    if len(alto) < 2 or len(bajo) < 2:
        return res

    # ¿A qué bloque pertenece el PIN? Al del centroide más cercano.
    def centro(grp):
        return (sum(g["lng"] for g in grp) / len(grp), sum(g["lat"] for g in grp) / len(grp))
    c_alto, c_bajo = centro(alto), centro(bajo)
    d_alto = math.hypot(pin_lng - c_alto[0], pin_lat - c_alto[1])
    d_bajo = math.hypot(pin_lng - c_bajo[0], pin_lat - c_bajo[1])
    bloque = alto if d_alto <= d_bajo else bajo
    nse_bloque = "alto_valor" if d_alto <= d_bajo else "bajo_valor"

    # Hull del bloque de NSE del pin (zona de influencia socioeconómica).
    hull = _convex_hull([[g["lng"], g["lat"]] for g in bloque])
    if not hull:
        res["motivo"] = "bloque de NSE sin hull"
        return res

    res.update({
        "barrera": True,
        "zona_poligono": [hull],
        "n_bloque": len(bloque),
        "nse_bloque": nse_bloque,
        "ratio": sep.get("ratio"),
        "motivo": (f"Frontera espacial de NSE (ratio {sep.get('ratio')}). "
                   f"Zona = bloque {nse_bloque} del pin ({len(bloque)} AGEBs)."),
    })
    return res


def _nse_separacion_espacial(agebs_geo: List[Dict]) -> Dict[str, Any]:
    """Detecta si los AGEBs de NSE distinto están geográficamente separados.

    Agrupa los AGEBs en dos bloques por nse_rank (alto vs bajo respecto a la mediana),
    calcula el centroide de cada bloque y compara la separación ENTRE bloques contra la
    dispersión interna media. Si la distancia entre centroides supera la dispersión
    interna, hay frontera espacial real. Con pocos AGEBs o NSE uniforme → separados=False
    (no se inventa barrera sin evidencia espacial).
    """
    pts = [g for g in agebs_geo if g.get("nse_rank") is not None and g.get("lng") and g.get("lat")]
    if len(pts) < 6:
        return {"separados": False, "motivo": "muestra insuficiente", "n": len(pts)}
    ranks = sorted(set(g["nse_rank"] for g in pts))
    if len(ranks) < 2:
        return {"separados": False, "motivo": "NSE uniforme", "n": len(pts)}

    rank_med = statistics.median(g["nse_rank"] for g in pts)
    alto = [g for g in pts if g["nse_rank"] <= rank_med]
    bajo = [g for g in pts if g["nse_rank"] > rank_med]
    if len(alto) < 2 or len(bajo) < 2:
        alto = [g for g in pts if g["nse_rank"] < rank_med]
        bajo = [g for g in pts if g["nse_rank"] >= rank_med]
    if len(alto) < 2 or len(bajo) < 2:
        return {"separados": False, "motivo": "sin dos bloques de NSE", "n": len(pts)}

    def centro(grp):
        return (sum(g["lng"] for g in grp) / len(grp), sum(g["lat"] for g in grp) / len(grp))
    def disp(grp, c):
        return (sum(math.hypot(g["lng"] - c[0], g["lat"] - c[1]) for g in grp) / len(grp)) or 1e-9
    c_alto, c_bajo = centro(alto), centro(bajo)
    d_entre = math.hypot(c_alto[0] - c_bajo[0], c_alto[1] - c_bajo[1])
    disp_media = (disp(alto, c_alto) + disp(bajo, c_bajo)) / 2
    ratio = d_entre / disp_media if disp_media else 0.0
    SEP_MIN = 1.0
    return {
        "separados": ratio >= SEP_MIN,
        "ratio": round(ratio, 2),
        "n_alto": len(alto), "n_bajo": len(bajo),
        "rank_corte": rank_med,
        "motivo": ("frontera espacial de NSE" if ratio >= SEP_MIN
                   else "NSE entremezclado (sin frontera)"),
    }


def derive_segments(agebs: List[Dict], ft: List[Dict], tipo_vivienda: str = "vertical",
                    agebs_geo: Optional[List[Dict]] = None,
                    pin_lng: Optional[float] = None, pin_lat: Optional[float] = None,
                    ft_renta: Optional[List[Dict]] = None) -> List[Dict[str, Any]]:
    """
    Metodología DPO de demanda (universal para todo México):

    0. BARRERA DE NSE: el análisis se hace sobre el NSE DOMINANTE de la zona de influencia
       (mayor masa de hogares en la isócrona), no sobre el promedio. Se puede inducir demanda
       a precios mayores, pero por debajo del piso del NSE superior (mejor precio a cambio de
       sacrificar entorno urbano, con producto nuclear casi tan bueno).
    1. CAPACIDAD DE COMPRA: cada hogar puede pagar vivienda ≈ ingreso_anual × MÚLTIPLO
       (4.5 años). La demanda base se asigna al bucket de precio que su ingreso permite.
    2. ANCLA DE PERCEPCIÓN DE VALOR: el nivel de precio "natural" de la zona se valida
       contra el producto que realmente se vende (VVV).
    3. EXPANSIÓN CON EVIDENCIA: la demanda puede subir a buckets superiores solo si hay
       EVIDENCIA de unidades vendidas ahí (VVV) coherente con ingresos altos en la zona.
    4. TOPE: no se permiten más de 2 buckets por encima del nivel de valor percibido de la
       zona si NO hay evidencia de venta de producto en esos niveles.
    """
    if not agebs:
        return []

    # ── 0 · BARRERA DE NSE (identifica el submercado dominante; NO recorta la demanda) ──
    # La demanda se calcula sobre TODO el espectro de la zona de influencia. La barrera de
    # NSE se usa para: (a) anclar el SWEET SPOT al bucket del NSE dominante (el más accesible),
    # (b) fijar el TECHO DE INDUCCIÓN (piso del NSE superior). Todos los buckets con demanda
    # u oferta se muestran como productos (recomendados o no).
    agebs_dom, nse_dom, nse_superior, share_dom = _nse_dominante_agebs(agebs, agebs_geo, pin_lng, pin_lat)
    # H9 · NSE realmente PRESENTES en la zona (para exigir presencia al recalcular el superior)
    nse_presentes = {ageb_nse(r) for r in agebs}
    nse_presentes.discard(None)

    # ── Buckets de precio canónicos (M MXN) ──
    BUCKETS = [
        (0, 1.5, "< $1.5M"), (1.5, 2.5, "$1.5M-$2.5M"), (2.5, 3.5, "$2.5M-$3.5M"),
        (3.5, 5.0, "$3.5M-$5.0M"), (5.0, 7.0, "$5.0M-$7.0M"), (7.0, 10.0, "$7.0M-$10.0M"),
        (10.0, 15.0, "$10.0M-$15.0M"), (15.0, 25.0, "$15.0M-$25.0M"), (25.0, 999.0, "> $25.0M"),
    ]
    def bucket_idx(precio_m):
        for i, (lo, hi, _) in enumerate(BUCKETS):
            if lo <= precio_m < hi:
                return i
        return None

    # ── 1 · DEMANDA POR CAPACIDAD DE COMPRA ──
    # INEGI ya estima la capacidad de compra real por hogar en "Rangos demanda vivienda"
    # (captura patrimonio, apalancamiento, etc., no solo múltiplo de ingreso corriente).
    # Tomamos esa demanda por rango y la mapeamos a nuestros buckets canónicos.
    demanda_bucket = {i: {"nuevas_fam": 0.0, "hogares": 0.0, "mercado_venta": 0.0, "ixh_w": 0.0, "ixh_hog": 0.0,
                          "renta_w": 0.0, "propia_w": 0.0, "ten_hog": 0.0} for i in range(len(BUCKETS))}
    # H7 · TENENCIA REAL: la propensión venta/renta de cada bucket sale de la tenencia REAL
    # (Propia/Alquilada/Prestada/Otra) de sus AGEBs — dato de la base, no proporción inventada.
    # Fallback: tenencia agregada de la ZONA; sin dato en toda la zona → None (N/D legítimo).
    _zona_ten = {"alq": 0.0, "prop": 0.0, "tot": 0.0}

    def _tenencia_shares(r):
        """(share_renta, share_propia) del AGEB por tenencia real; (None, None) sin dato."""
        pr = _num(r.get("Propia")); al = _num(r.get("Alquilada"))
        pe = _num(r.get("Prestada")); ot = _num(r.get("Otra situación"))
        tot = sum(v for v in (pr, al, pe, ot) if v is not None)
        if not tot:
            return None, None
        return ((al or 0) / tot, (pr or 0) / tot)

    for r in agebs:
        rango = r.get("Rangos demanda vivienda")
        dem = _num(r.get("Demanda anual vivienda"))
        hog = _num(r.get("Hogares totales 2026")) or 0
        _sr, _sp = _tenencia_shares(r)
        if _sr is not None and hog:
            _zona_ten["alq"] += hog * _sr
            _zona_ten["prop"] += hog * _sp
            _zona_ten["tot"] += hog
        # "Mercado en venta" = pool de compradores potenciales activos en venta del AGEB (demanda
        # total). Regula la velocidad en segmentos sobreofertados. Distinto de "Demanda anual
        # vivienda" (flujo de nuevas familias/año).
        mkt_venta = _num(r.get("Mercado en venta")) or 0
        if not rango or dem is None:
            continue
        low = _extract_low(rango)   # valor en MXN
        if low is None:
            continue
        bi = bucket_idx(low / 1e6)
        if bi is None:
            continue
        # PONDERACIÓN POR TIPO DE VIVIENDA: en modo horizontal, la demanda de vivienda del
        # AGEB se pondera por la proporción de viviendas tipo CASA (la demanda de casa real
        # del AGEB). En zonas de departamentos la demanda horizontal baja; en zonas de casas
        # se mantiene. El modo vertical usa la demanda genérica (comportamiento validado).
        if tipo_vivienda == "horizontal":
            peso_tipo = _ageb_prop_casa(r)
            dem = dem * peso_tipo
            hog = hog * peso_tipo
            mkt_venta = mkt_venta * peso_tipo
        demanda_bucket[bi]["nuevas_fam"] += dem
        demanda_bucket[bi]["hogares"] += hog
        demanda_bucket[bi]["mercado_venta"] += mkt_venta
        # H7: tenencia real acumulada por bucket (share del AGEB × hogares del AGEB en el bucket)
        if _sr is not None and hog:
            demanda_bucket[bi]["renta_w"] += hog * _sr
            demanda_bucket[bi]["propia_w"] += hog * _sp
            demanda_bucket[bi]["ten_hog"] += hog
        # Ingreso real (IXH mensual) ponderado por hogares → NSE del bucket por INGRESO real,
        # no por el precio de la vivienda (evita etiquetar NSE C un bucket comprado por NSE D).
        ixh_m = _ageb_ixh_mensual(r)
        if ixh_m is not None and hog:
            demanda_bucket[bi]["ixh_w"] += ixh_m * hog
            demanda_bucket[bi]["ixh_hog"] += hog

    # ── 2 · Evidencia de venta real por bucket (VVV) ──
    PM2_VERTICAL_MIN = 20000   # piso de plausibilidad: pm2 vertical real
    evidencia = {i: {"vendidas": 0, "disp": 0, "n": 0, "pm2_vals": []} for i in range(len(BUCKETS))}
    for t in ft:
        a = t.get("attributes", {})
        precio = _price(a.get("F____UNIDAD"))
        if not precio:
            continue
        bi = bucket_idx(precio / 1e6)
        if bi is None:
            continue
        evidencia[bi]["vendidas"] += int(_num(a.get("UNIDADES_VENDIDAS")) or 0)
        evidencia[bi]["disp"] += int(_num(a.get("UNIDADES_DISPONIBLES")) or 0)
        evidencia[bi]["n"] += 1
        pm2v = _pm2(a.get("F___M2"))
        if pm2v and pm2v >= PM2_VERTICAL_MIN:
            evidencia[bi]["pm2_vals"].append(pm2v)
    # ¿El bucket tiene oferta VERTICAL plausible? (mediana pm2 ≥ piso vertical)
    for i in evidencia:
        pv = evidencia[i]["pm2_vals"]
        evidencia[i]["es_vertical"] = bool(pv) and statistics.median(pv) >= PM2_VERTICAL_MIN

    # ── REGLA DE PERCEPCIÓN DE VALOR (universal) ──
    # La ISÓCRONA determina la ZONA (demografía → demanda). La PERCEPCIÓN DE VALOR del
    # PRODUCTO la determina la OFERTA observada: el valor de vivienda vertical que realmente
    # se vende en la zona. La percepción de valor DOMINANTE predomina y PUEDE AUMENTAR si hay
    # evidencia de oferta en valores superiores. Así, una zona con demografía diluida (p.ej.
    # Valle Poniente sale D+ por accesibilidad) pero con oferta vertical premium ($142k/m²)
    # reconoce su percepción de valor real (A premium) para anclar el producto.
    NSE_ORDEN_PV = ["A", "B", "C+", "C", "D+", "D", "E"]
    rank_pv = {n: k for k, n in enumerate(NSE_ORDEN_PV)}
    def _nse_de_precio_m(precio_m):
        bands = [("A", 6.8, None), ("B", 3.05, 6.8), ("C+", 1.35, 3.05), ("C", 0.577, 1.35),
                 ("D+", 0.349, 0.577), ("D", 0.2, 0.349), ("E", 0, 0.2)]
        for nse, lo, hi in bands:
            if precio_m >= lo and (hi is None or precio_m < hi):
                return nse
        return None
    # Percepción de valor dominante = NSE del PRECIO MEDIANO de la oferta vertical real.
    # El precio mediano de lo que se vende/ofrece indica para qué percepción se construye en
    # la zona (la masa de oferta vertical), independiente de la demografía de la isócrona.
    precios_oferta_m = []
    for t in ft:
        a = t.get("attributes", {})
        precio = _price(a.get("F____UNIDAD"))
        pm2v = _pm2(a.get("F___M2"))
        # solo oferta vertical plausible (pm2 ≥ piso vertical)
        if precio and pm2v and pm2v >= PM2_VERTICAL_MIN:
            precios_oferta_m.append(precio / 1e6)
    nse_percepcion = None
    if precios_oferta_m:
        precio_mediano_m = statistics.median(precios_oferta_m)
        nse_percepcion = _nse_de_precio_m(precio_mediano_m)
    # La percepción de valor PREDOMINA: si la oferta observada indica un NSE superior al
    # demográfico, el ancla del producto sube a esa percepción (puede aumentar con evidencia).
    nse_dom_demografico = nse_dom
    salto_percepcion = 0   # cuántos niveles NSE sube la percepción sobre la demografía
    if nse_percepcion and rank_pv.get(nse_percepcion, 99) < rank_pv.get(nse_dom, 99):
        salto_percepcion = rank_pv.get(nse_dom, 0) - rank_pv.get(nse_percepcion, 0)
        nse_dom = nse_percepcion
        # Recalcular el NSE superior respecto a la nueva percepción dominante.
        # H9: la regla exige que el NSE superior esté PRESENTE en la zona (masa real de
        # hogares); antes se tomaba de la escala completa aunque no existiera en la zona.
        nse_superior = None
        for n in NSE_ORDEN_PV:
            if n in nse_presentes and rank_pv.get(n, 99) < rank_pv.get(nse_dom, 0):
                nse_superior = n  # el más cercano por encima PRESENTE

    # ── 3 · Nivel de valor percibido = bucket con mayor demanda por capacidad de compra ──
    nivel_percibido = max(demanda_bucket, key=lambda i: demanda_bucket[i]["nuevas_fam"]) \
        if any(demanda_bucket[i]["nuevas_fam"] for i in demanda_bucket) else 0

    # ── 4 · TOPE: no exceder +2 buckets sobre el nivel percibido SIN evidencia de venta ──
    EVIDENCIA_MIN = 3  # unidades vendidas para validar producto en ese nivel
    bucket_max_permitido = nivel_percibido + 2
    for i in range(nivel_percibido + 1, len(BUCKETS)):
        if evidencia[i]["vendidas"] >= EVIDENCIA_MIN:
            bucket_max_permitido = max(bucket_max_permitido, i)  # la evidencia extiende el tope

    # ── TECHO DE INDUCCIÓN POR BARRERA DE NSE ──
    # Se puede inducir demanda a precios mayores que el NSE dominante (mejor precio a cambio
    # de sacrificar entorno urbano, con producto nuclear casi tan bueno), PERO siempre por
    # DEBAJO del piso de valor de vivienda del NSE superior presente en la zona.
    NSE_VIV_PISO_M = {"A": 6.8, "B": 3.05, "C+": 1.35, "C": 0.577, "D+": 0.349, "D": 0.2, "E": 0.0}
    if nse_superior:
        piso_superior_m = NSE_VIV_PISO_M.get(nse_superior, 999)
        # Último bucket cuyo techo no alcanza el piso del NSE superior
        techo_induccion = 0
        for i, (lo, hi, _) in enumerate(BUCKETS):
            if lo < piso_superior_m:
                techo_induccion = i
        # El tope no puede exceder el techo de inducción (salvo evidencia dura ya contemplada)
        bucket_max_permitido = min(bucket_max_permitido, max(techo_induccion, nivel_percibido))

    # ── DEMANDA GRANULAR REAL (sin redistribución artificial) ──
    # Cada bucket conserva la demanda REAL que la base asigna a ese rango de precio (campo
    # "Demanda anual vivienda" por AGEB, ya mapeado arriba). NO se redistribuye ni se promedia:
    # se usa la información granular tal cual la genera la base ArcGIS. El "techo de inducción"
    # y el "tope" se conservan SOLO como metadato informativo (no alteran la demanda mostrada).
    techo_induccion_idx = bucket_max_permitido  # metadato: hasta dónde es inducible la demanda
    # (intencionalmente NO se redistribuye demanda entre buckets · dato granular intacto)

    # ── Construir segmentos ──
    def nse_by_viv(valor_m):
        bands = [("A", 6.8, None), ("B", 3.05, 6.8), ("C+", 1.35, 3.05), ("C", 0.577, 1.35),
                 ("D+", 0.349, 0.577), ("D", 0.2, 0.349), ("E", 0, 0.2)]
        for nse, lo, hi in bands:
            if valor_m >= lo and (hi is None or valor_m < hi):
                return nse
        return "—"
    nse_ing_band = {"A": (200000, 350000), "B": (90000, 199999), "C+": (40000, 89999),
                    "C": (17000, 39999), "D+": (10000, 16999), "D": (4000, 9999), "E": (0, 3999)}

    def nse_by_ingreso(ixh_mensual):
        """Clasifica el NSE por el INGRESO mensual real del hogar (AMAI). El NSE de un segmento
        debe reflejar QUIÉN compra (ingreso real), no el precio de la vivienda. Devuelve None si
        no hay ingreso (entonces se cae a la clasificación por valor de vivienda)."""
        if ixh_mensual is None:
            return None
        for nse, (lo, hi) in nse_ing_band.items():
            if lo <= ixh_mensual <= hi:
                return nse
        if ixh_mensual > 200000:
            return "A"
        return "E"

    # ── H8 · TASA DE RENTA OBSERVADA DE LA ZONA ──
    # renta_pct_zona = mediana($/m²/mes de la oferta de renta real vv_renta) ÷ mediana($/m² de
    # la oferta de venta real). Dato observado (oferta como laboratorio). Con menos de 3
    # observaciones en cualquiera de las dos capas → fallback a la regla base DIGO (0.4%/mes),
    # que queda DOCUMENTADA como fallback, no como regla principal.
    RENTA_PCT_BASE = 0.004
    _pm2_renta_vals = []
    for _t in (ft_renta or []):
        _a = _t.get("attributes", {})
        # $/m²/MES de renta: rango plausible 20-5,000 (NO usar _pm2, cuyo piso >1,000 es de
        # VENTA y descartaría rentas reales de $100-600/m²/mes).
        _rm = _num(_a.get("F___M2"))
        if _rm is None or not (20 <= _rm <= 5000):
            _ru = _num(_a.get("F____UNIDAD")); _ap = _num(_a.get("ÁREA_PRIVATIVA"))
            _rm = (_ru / _ap) if (_ru and _ru > 1000 and _ap and M2_MIN <= _ap <= M2_MAX) else None
        if _rm and 20 <= _rm <= 5000:
            _pm2_renta_vals.append(_rm)
    _pm2_venta_vals = []
    for _t in ft:
        _a = _t.get("attributes", {})
        _pv = _pm2(_a.get("F___M2"))
        if _pv and _pv >= PM2_VERTICAL_MIN:
            _pm2_venta_vals.append(_pv)
    if len(_pm2_renta_vals) >= 3 and len(_pm2_venta_vals) >= 3:
        renta_pct_zona = statistics.median(_pm2_renta_vals) / statistics.median(_pm2_venta_vals)
        renta_pct_fuente = "observada"
    else:
        renta_pct_zona = RENTA_PCT_BASE
        renta_pct_fuente = "base_digo"

    segments = []
    total_nf = sum(demanda_bucket[i]["nuevas_fam"] for i in demanda_bucket)

    def nse_by_viv2(valor_m):
        return nse_by_viv(valor_m)
    def _nse_de_bucket(i):
        lo, hi, _ = BUCKETS[i]
        mid = (lo + (hi if hi < 999 else lo * 1.3)) / 2
        return nse_by_viv(mid)
    NSE_ORDEN = ["A", "B", "C+", "C", "D+", "D", "E"]
    rank = {n: k for k, n in enumerate(NSE_ORDEN)}
    rank_dom = rank.get(nse_dom, 99)
    rank_sup = rank.get(nse_superior, -1) if nse_superior else -1

    # ════════ SELECCIÓN UNIFICADA DEL PRODUCTO ANCLA (sweet / featured) ════════
    # Regla: la PERCEPCIÓN DE VALOR (bucket del precio mediano de la oferta vertical) define
    # el nivel del producto ancla. Sobre ese nivel:
    #   • si el bucket tiene demanda DIM local → es SWEET SPOT (demand-driven).
    #   • si NO tiene demanda local pero sí oferta vertical → es ancla de PERCEPCIÓN DE VALOR
    #     (mercado_dual): el producto se vende a foráneos/inversionistas (nota explícita).
    # El bucket de percepción de valor: el del precio mediano de la oferta vertical observada.
    pv_idx = None
    if precios_oferta_m:
        pv_idx = bucket_idx(statistics.median(precios_oferta_m))
    # Si ese bucket no tiene oferta vertical plausible, bajar al bucket vertical más cercano
    if pv_idx is None or not evidencia.get(pv_idx, {}).get("es_vertical"):
        verticales = [i for i in range(len(BUCKETS)) if evidencia[i].get("es_vertical")
                      and evidencia[i]["vendidas"] > 0]
        if verticales and pv_idx is not None:
            pv_idx = min(verticales, key=lambda i: abs(i - pv_idx))
        elif verticales:
            pv_idx = max(verticales, key=lambda i: evidencia[i]["vendidas"])
        else:
            pv_idx = None

    sweet_idx = None
    mercado_dual = False
    dual_idx = None
    if pv_idx is not None:
        if demanda_bucket[pv_idx]["nuevas_fam"] > 0:
            # Convergencia demanda×oferta en el nivel de percepción de valor → sweet spot
            sweet_idx = pv_idx
        else:
            # Oferta vertical en el nivel de percepción de valor SIN demanda DIM local:
            # mercado anclado a percepción de valor / foráneo (dual)
            mercado_dual = True
            dual_idx = pv_idx
    else:
        # No hay oferta vertical en absoluto: mercado incipiente demand-driven puro.
        # El ancla es el bucket de mayor demanda dentro del TECHO DE INDUCCIÓN (regla
        # confirmada: se puede inducir demanda hacia arriba, pero por debajo del piso del
        # NSE superior presente). El comentario ya lo enunciaba; ahora el filtro lo aplica.
        cand = {i: demanda_bucket[i]["nuevas_fam"] for i in range(len(BUCKETS))
                if demanda_bucket[i]["nuevas_fam"] > 0
                and rank.get(_nse_de_bucket(i), 99) <= rank_dom
                and i <= bucket_max_permitido}
        if not cand:
            cand = {i: demanda_bucket[i]["nuevas_fam"] for i in range(len(BUCKETS))
                    if demanda_bucket[i]["nuevas_fam"] > 0}
        sweet_idx = max(cand, key=cand.get) if cand else None


    # RANGO DE VALOR APLICABLE DE LA ZONA (para marcar buckets fuera de rango como N/A).
    # Un bucket es APLICABLE si está dentro del rango de valor donde existe oferta vertical
    # real (VVV) en la zona. Fuera de ese rango, el bucket se MUESTRA pero se marca como
    # no-aplicable (no se inventa: la marca se basa en la oferta observada, dato de la base).
    buckets_con_oferta = [i for i in range(len(BUCKETS))
                          if (evidencia[i]["vendidas"] > 0 or evidencia[i]["disp"] > 0
                              or evidencia[i]["n"] > 0)]
    if buckets_con_oferta:
        rango_oferta_lo, rango_oferta_hi = min(buckets_con_oferta), max(buckets_con_oferta)
    else:
        rango_oferta_lo, rango_oferta_hi = None, None

    for i, (lo, hi, label) in enumerate(BUCKETS):
        db = demanda_bucket[i]
        vend = evidencia[i]["vendidas"]; disp = evidencia[i]["disp"]
        tiene_demanda = db["nuevas_fam"] > 0 or db["hogares"] > 0
        tiene_oferta = vend > 0 or disp > 0 or evidencia[i]["n"] > 0
        # OPCIÓN 3: incluir el bucket si tiene demanda O si tiene oferta (supply-driven)
        if not tiene_demanda and not tiene_oferta:
            continue

        nuevas = round(db["nuevas_fam"], 1)
        mkt_total = round(db["hogares"])
        mid_m = (lo + (hi if hi < 999 else lo * 1.3)) / 2
        # NSE del segmento = por INGRESO real de los hogares del bucket (quién compra), con
        # fallback a clasificación por valor de vivienda si no hay ingreso (integridad).
        ixh_bucket = (db["ixh_w"] / db["ixh_hog"]) if db.get("ixh_hog") else None
        nse_cls = nse_by_ingreso(ixh_bucket) or nse_by_viv(mid_m)

        # ORIGEN: demand_driven si hay demanda DIM local; supply_driven si solo hay oferta
        origen = "demand_driven" if tiene_demanda else "supply_driven"

        # Status: combina demanda (nuevas familias) con oferta disponible + evidencia de venta
        if i == sweet_idx:
            status = "sweet_spot"
        elif not tiene_demanda and tiene_oferta:
            # Supply-driven: el mercado lo atiende la oferta sin demanda DIM local medible.
            # Si se vende (evidencia), es atendido/oceano_rojo; si no, bajo_crecimiento.
            if vend >= 3 and disp < vend:
                status = "atendido"
            elif disp > 0 and vend == 0:
                status = "bajo_crecimiento"
            else:
                status = "oceano_rojo"
        elif nuevas == 0:
            status = "bajo_crecimiento"
        elif disp == 0 and vend == 0:
            status = "desatendido"
        elif disp < nuevas * 12:
            status = "oportunidad"
        elif disp < nuevas * 36:
            status = "atendido"
        else:
            status = "oceano_rojo"

        ing_lo, ing_hi = nse_ing_band.get(nse_cls, (0, 0))
        # H8 · Renta mensual = valor × TASA DE RENTA OBSERVADA de la zona (vv_renta real);
        # sin observaciones suficientes → 0.4% (regla base DIGO como fallback documentado).
        # Para el bucket inferior (lo=0), piso representativo para no devolver renta 0/N/D.
        hi_eff = (hi if hi < 999 else lo * 1.3)
        lo_eff = lo if lo > 0 else hi_eff * 0.5   # piso del bucket inferior
        rent_min = round(lo_eff * 1e6 * renta_pct_zona)
        rent_max = round(hi_eff * 1e6 * renta_pct_zona)
        # H7 · Propensión venta/renta/propia del bucket por TENENCIA REAL (AGEBs del bucket;
        # fallback: tenencia de la zona; sin dato en toda la zona → None = N/D legítimo).
        if db.get("ten_hog"):
            share_renta = db["renta_w"] / db["ten_hog"]
            share_propia = db["propia_w"] / db["ten_hog"]
        elif _zona_ten["tot"]:
            share_renta = _zona_ten["alq"] / _zona_ten["tot"]
            share_propia = _zona_ten["prop"] / _zona_ten["tot"]
        else:
            share_renta = None
            share_propia = None
        es_dual_featured = (mercado_dual and i == dual_idx)
        # APLICABILIDAD: el bucket está dentro del rango de valor donde la zona tiene oferta
        # vertical real. Fuera de ese rango se muestra pero se marca no-aplicable.
        if rango_oferta_lo is not None:
            aplicable = (rango_oferta_lo <= i <= rango_oferta_hi)
        else:
            aplicable = True  # sin oferta observada, no se descarta nada (integridad)
        seg = {
            "NSE": nse_cls, "bucket": label,
            "val_min": lo * 1e6, "val_max": (hi if hi < 999 else lo * 1.3) * 1e6,
            "mkt_total": mkt_total, "nuevas_fam": nuevas,
            # TCA (tasa de crecimiento anual de hogares) canónica por NSE — fuente única backend.
            # El front la LEE de aquí; nunca la recalcula (antes tenía una tabla hardcodeada distinta).
            "tca": NSE_TCA.get(nse_cls, 0),
            "demanda_total": round(db["mercado_venta"]),   # "Mercado en venta": pool compradores activos
            # H7: venta/renta/propia derivadas de TENENCIA REAL (antes 0.83/0.17/0.65 fijos)
            "mkt_venta": (round(mkt_total * (1 - share_renta)) if (mkt_total and share_renta is not None) else (0 if not mkt_total else None)),
            "mkt_renta": (round(mkt_total * share_renta) if (mkt_total and share_renta is not None) else (0 if not mkt_total else None)),
            "share_renta": round(share_renta, 4) if share_renta is not None else None,
            "share_propia": round(share_propia, 4) if share_propia is not None else None,
            "rent_min": rent_min, "rent_max": rent_max,
            "renta_pct_zona": round(renta_pct_zona, 5),
            "renta_pct_fuente": renta_pct_fuente,
            "ing_min": ing_lo, "ing_max": ing_hi,
            "hog_propios": (round(mkt_total * share_propia) if (mkt_total and share_propia is not None) else (0 if not mkt_total else None)),
            "evidencia_vendidas": vend, "evidencia_disp": disp,
            "status": status, "origen": origen, "segments_in_bucket": 1,
            "aplicable": aplicable,
        }
        if es_dual_featured:
            seg["dual_featured"] = True
            if salto_percepcion >= 3:
                # Salto grande (p.ej. D+→A): zona popular con oferta vertical de lujo para foráneos
                seg["nota_mercado"] = "Mercado foráneo/inversionista · oferta vertical premium sin demanda local DIM"
            else:
                # Salto moderado: la percepción de valor de la zona supera su demografía;
                # el producto vertical atiende a la percepción de valor dominante observada.
                seg["nota_mercado"] = "Producto anclado a la percepción de valor de la zona (oferta vertical observada)"
        segments.append(seg)

    return segments


# ════════════ DEM-1 · SEGMENTOS POR PERFIL (diseño aprobado · docs/DISENO_DEM1.md) ════════════
# Capacidad de pago por MENSUALIDAD (U3 corregido por Héctor: banca MX 30-35% pago-ingreso;
# el 28% era convención de EUA). Tasa/plazo/enganche configurables por entorno.
PTI_REF = float(os.environ.get("DATARIA_PTI_REF", "0.30"))          # central prudente (CONDUSEF)
PTI_MAX = float(os.environ.get("DATARIA_PTI_MAX", "0.35"))          # techo banca MX
TASA_HIPOTECARIA_REF = float(os.environ.get("DATARIA_TASA_HIP", "0.091"))   # 9.1% · CONFIRMADA por Héctor (7 jul 2026) · full backend, el front jamás la ajusta
PLAZO_HIP_MESES = int(os.environ.get("DATARIA_PLAZO_HIP", "240"))
ENGANCHE_REF = float(os.environ.get("DATARIA_ENGANCHE", "0.10"))
# U1 · umbrales INICIALES a calibrar por sensibilidad en zonas ancla (no dogma)
UMBRAL_PERFIL_PCT = float(os.environ.get("DATARIA_UMBRAL_PERFIL_PCT", "0.05"))
UMBRAL_PERFIL_HOG = int(os.environ.get("DATARIA_UMBRAL_PERFIL_HOG", "300"))

# Bandas de tamaño por ticket (M MXN) · FUENTE ÚNICA a nivel módulo (antes vivía duplicada
# dentro de derive_productos_venta; mismos valores, cero cambio de comportamiento).
_BANDA_TAMANO_TICKET = [
    (1.5, 45, 58), (2.5, 52, 68), (3.5, 62, 80), (5.0, 70, 90), (8.0, 80, 110),
    (12.0, 105, 140), (18.0, 135, 180), (25.0, 180, 260), (None, 260, 360),
]


def _banda_tamano_por_ticket(ticket_m) -> tuple:
    """(m² min, m² max) del programa habitable para un ticket (M MXN)."""
    if ticket_m is None:
        return (70, 90)
    for hi, lo_b, hi_b in _BANDA_TAMANO_TICKET:
        if hi is None or ticket_m <= hi:
            return (lo_b, hi_b)
    return (260, 360)


def _capacidad_pago_banda_M(ixh_mensual) -> tuple:
    """(precio_min_M, precio_max_M) que el ingreso soporta por MENSUALIDAD:
    pago = IXH × PTI; crédito = pago / factor(tasa, plazo); precio = crédito / (1−enganche).
    Banda = [PTI 30%, PTI 35%]. Sin ingreso → (None, None) — jamás inventar."""
    if not ixh_mensual or ixh_mensual <= 0:
        return (None, None)
    r = TASA_HIPOTECARIA_REF / 12.0
    f = r / (1.0 - (1.0 + r) ** (-PLAZO_HIP_MESES))
    def _precio(pti):
        return (ixh_mensual * pti / f) / (1.0 - ENGANCHE_REF)
    return (round(_precio(PTI_REF) / 1e6, 3), round(_precio(PTI_MAX) / 1e6, 3))


def _gmm2_bic(vals: List[float]) -> Dict[str, Any]:
    """Mezcla de 2 gaussianas 1-D (EM) vs 1 gaussiana, comparadas por BIC.
    Detecta SUBMERCADOS de ingreso dentro de un NSE (zonas en transición).
    SALVAGUARDA de integridad: n<30 → k=1 (con muestras chicas el GMM sobreajusta)."""
    v = sorted(x for x in vals if x is not None and x > 0)
    if len(v) < 30:
        return {"k": 1, "medias": [round(statistics.median(v))] if v else [],
                "corte": None, "motivo": "muestra_insuficiente"}
    n = len(v)
    mu = statistics.mean(v)
    var = statistics.pvariance(v) or 1.0
    ll1 = sum(-0.5 * math.log(2 * math.pi * var) - (x - mu) ** 2 / (2 * var) for x in v)
    bic1 = -2 * ll1 + 2 * math.log(n)
    m1, m2 = _percentil(v, 0.25), _percentil(v, 0.75)
    s1 = s2 = math.sqrt(var) or 1.0
    w = 0.5
    for _ in range(80):
        resp = []
        for x in v:
            p1 = w * math.exp(-(x - m1) ** 2 / (2 * s1 * s1)) / (s1 or 1e-9)
            p2 = (1 - w) * math.exp(-(x - m2) ** 2 / (2 * s2 * s2)) / (s2 or 1e-9)
            resp.append(p1 / (p1 + p2) if (p1 + p2) > 0 else 0.5)
        W1 = sum(resp)
        W2 = n - W1
        if W1 < 1e-6 or W2 < 1e-6:
            break
        m1n = sum(r * x for r, x in zip(resp, v)) / W1
        m2n = sum((1 - r) * x for r, x in zip(resp, v)) / W2
        s1 = math.sqrt(max(sum(r * (x - m1n) ** 2 for r, x in zip(resp, v)) / W1, 1e-6))
        s2 = math.sqrt(max(sum((1 - r) * (x - m2n) ** 2 for r, x in zip(resp, v)) / W2, 1e-6))
        w = W1 / n
        conv = abs(m1n - m1) + abs(m2n - m2)
        m1, m2 = m1n, m2n
        if conv < 1e-4:
            break
    ll2 = 0.0
    for x in v:
        p1 = w / (s1 * math.sqrt(2 * math.pi)) * math.exp(-(x - m1) ** 2 / (2 * s1 * s1))
        p2 = (1 - w) / (s2 * math.sqrt(2 * math.pi)) * math.exp(-(x - m2) ** 2 / (2 * s2 * s2))
        ll2 += math.log(max(p1 + p2, 1e-300))
    bic2 = -2 * ll2 + 5 * math.log(n)
    sep_rel = abs(m2 - m1) / (((m1 + m2) / 2) or 1e-9)
    if bic2 < bic1 and sep_rel > 0.25:
        return {"k": 2, "medias": sorted([round(m1), round(m2)]),
                "corte": round((m1 + m2) / 2), "motivo": "bic_mejora"}
    return {"k": 1, "medias": [round(mu)], "corte": None, "motivo": "bic_no_mejora"}


# Programa base por cohorte de etapa de vida (U2 aprobado; la NORMATIVA manda en cajones)
_COHORTES_DEM1 = {
    "C1": {"label": "Joven solo",              "rec": 1, "cajones": 1.0},
    "C2": {"label": "Pareja joven sin hijos",  "rec": 2, "cajones": 1.5},
    "C3": {"label": "Familia en formación",    "rec": 2, "cajones": 1.5},
    "C4": {"label": "Familia consolidada",     "rec": 3, "cajones": 2.0},
    "C5": {"label": "Corresidentes",           "rec": 2, "cajones": 2.0},
    "C6": {"label": "Adulto solo",             "rec": 1, "cajones": 1.0},
    "C7": {"label": "Nido vacío",              "rec": 2, "cajones": 1.5},
}


def derive_segmentos_dem1(agebs: List[Dict], segments: List[Dict],
                          personas_hogar: Optional[float] = None) -> Dict[str, Any]:
    """DEM-1 · Matriz de PERFILES (cohorte × NSE × ingreso) con CONSERVACIÓN de la
    demanda por bucket ya validada (los buckets se REPARTEN, no se alteran).
    Ver docs/DISENO_DEM1.md. Todo dato real; asignaciones documentadas en fuente_masas."""
    meta = {"metodo_masas": "asignacion_v1_piramide_tipologia",
            "umbral": {"pct": UMBRAL_PERFIL_PCT, "hogares": UMBRAL_PERFIL_HOG},
            "capacidad": {"pti_ref": PTI_REF, "pti_max": PTI_MAX,
                          "tasa": TASA_HIPOTECARIA_REF, "plazo_m": PLAZO_HIP_MESES,
                          "enganche": ENGANCHE_REF},
            "gmm": {}, "conservacion": None,
            "nota_captable": ("Mercado captable (zonas de origen + extranjeros) pendiente "
                              "de P1/P2; solo se reporta demanda NATURAL. No se inventa."),
            "version_modelo": "dem1_v1"}
    if not agebs:
        return {"segmentos": [], "meta": meta}

    # ── 1 · Grupos por NSE (masas y pirámide reales por AGEB, agregadas por NSE) ──
    grupos: Dict[str, Dict[str, Any]] = {}
    for r in agebs:
        nse = ageb_nse(r)
        if not nse:
            continue
        g = grupos.setdefault(nse, {"hog": 0.0, "ixh": [], "unip": 0.0, "corres": 0.0,
                                    "fam": 0.0, "ninos": 0.0, "adol": 0.0, "jovad": 0.0,
                                    "consol": 0.0, "nest": 0.0, "n_agebs": 0})
        h = _num(r.get("Hogares totales 2026")) or _num(r.get("Hogares totales 2020")) or 0
        g["hog"] += h
        g["n_agebs"] += 1
        ixh = _ageb_ixh_mensual(r)
        if ixh and h:
            g["ixh"].append((ixh, h))
        g["unip"] += _num(r.get("Hogares unipersonales")) or 0
        g["corres"] += _num(r.get("Hogares corresidentes")) or 0
        g["fam"] += _num(r.get("Hogares familiares totales")) or 0
        g["ninos"] += _num(r.get("Niños")) or 0
        g["adol"] += _num(r.get("Adolescentes")) or 0
        g["jovad"] += _num(r.get("Jovenes_adultos")) or 0
        g["consol"] += _num(r.get("Consolidados")) or 0
        g["nest"] += _num(r.get("Nesters")) or 0
    total_hog_zona = sum(g["hog"] for g in grupos.values()) or 1.0
    umbral = min(UMBRAL_PERFIL_PCT * total_hog_zona, float(UMBRAL_PERFIL_HOG))

    # ── 2 · Perfiles por NSE (con GMM de ingreso para submercados/transición) ──
    perfiles: List[Dict[str, Any]] = []
    for nse, g in grupos.items():
        if g["hog"] <= 0:
            continue
        gmm = _gmm2_bic([x for x, _ in g["ixh"]])
        meta["gmm"][nse] = gmm
        if gmm["k"] == 2 and gmm.get("corte"):
            bandas = [("bajo", [(x, h) for x, h in g["ixh"] if x <= gmm["corte"]]),
                      ("alto", [(x, h) for x, h in g["ixh"] if x > gmm["corte"]])]
        else:
            bandas = [(None, g["ixh"])]
        # Asignación de cohortes (v1 · documentada): unipersonales/corresidentes directos
        # de tipología; familiares repartidos ∝ pirámide real. Sin tipología → degradación
        # explícita (caso base familiar) con fuente marcada.
        adultos = (g["jovad"] + g["consol"] + g["nest"]) or 1.0
        unip, corres, fam = g["unip"], g["corres"], g["fam"]
        fuente = "asignacion_v1_piramide_tipologia"
        if (unip + corres + fam) <= 0:
            fam = g["hog"]
            fuente = "asignacion_v1_solo_piramide"
        masas = {"C1": unip * (g["jovad"] / adultos),
                 "C6": unip * (g["consol"] / adultos),
                 "C5": corres}
        c7u = unip * (g["nest"] / adultos)
        fam_w = {"C2": g["jovad"] * 0.5, "C3": g["ninos"], "C4": g["adol"], "C7": g["nest"] * 0.5}
        fw_tot = sum(fam_w.values()) or 1.0
        for c, wv in fam_w.items():
            masas[c] = masas.get(c, 0.0) + fam * (wv / fw_tot)
        masas["C7"] = masas.get("C7", 0.0) + c7u
        hog_con_masa = sum(masas.values()) or 1.0
        escala = g["hog"] / hog_con_masa   # normalizar a hogares reales del grupo
        for banda_tag, banda in bandas:
            hb = sum(h for _, h in banda)
            peso_banda = (hb / g["hog"]) if (g["hog"] and hb) else (1.0 if banda_tag is None else 0.0)
            if peso_banda <= 0:
                continue
            ixh_med = statistics.median([x for x, _ in banda]) if banda else None
            cap_lo, cap_hi = _capacidad_pago_banda_M(ixh_med)
            for ck, cfg in _COHORTES_DEM1.items():
                stock = masas.get(ck, 0.0) * escala * peso_banda
                if stock <= 0:
                    continue
                n_rec = cfg["rec"]
                if ck == "C5" and personas_hogar:
                    n_rec = max(1, min(4, int(round(personas_hogar))))
                perfiles.append({
                    "perfil_id": f"{ck}-{nse}" + (f"-{banda_tag}" if banda_tag else ""),
                    "cohorte": ck, "cohorte_label": cfg["label"],
                    "nse": nse, "banda_ingreso_tag": banda_tag,
                    "ixh_mediana": round(ixh_med) if ixh_med else None,
                    "capacidad_pago_banda_M": [cap_lo, cap_hi],
                    "hogares_stock": round(stock),
                    "nuevas_fam_year": 0.0, "pool_activo": 0.0,
                    "crecimiento_pct": NSE_TCA.get(nse, 0),
                    "programa": {"rec": n_rec, "cajones": cfg["cajones"] + (0.5 if nse in ("A", "B") else 0.0)},
                    "fuente_masas": fuente,
                    "confianza": {"n_agebs": g["n_agebs"], "hogares_grupo": round(g["hog"])},
                })

    # ── 3 · Umbral U1 (calibrable): perfiles chicos se agregan al mayor de su NSE ──
    grandes = [p for p in perfiles if p["hogares_stock"] >= umbral]
    chicos = [p for p in perfiles if p["hogares_stock"] < umbral]
    for p in chicos:
        destino = max((q for q in grandes if q["nse"] == p["nse"]),
                      key=lambda q: q["hogares_stock"], default=None)
        if destino:
            destino["hogares_stock"] += p["hogares_stock"]
        else:
            grandes.append(p)   # NSE sin perfil grande: se conserva (no perder masa)
    perfiles = grandes

    # ── 4 · Reparto de la demanda por bucket (CONSERVACIÓN de masa validada) ──
    nf_buckets = 0.0
    for s in (segments or []):
        nf = s.get("nuevas_fam") or 0
        pool = s.get("demanda_total") or 0
        if nf <= 0 and pool <= 0:
            continue
        nf_buckets += nf
        lo_m = (s.get("val_min") or 0) / 1e6
        hi_m = (s.get("val_max") or 0) / 1e6
        cand = [p for p in perfiles
                if p["capacidad_pago_banda_M"][0] is not None
                and p["capacidad_pago_banda_M"][1] >= lo_m
                and p["capacidad_pago_banda_M"][0] < hi_m]
        if not cand:
            cand = [p for p in perfiles if p["nse"] == s.get("NSE")] or perfiles
        masa_c = sum(p["hogares_stock"] for p in cand) or 1.0
        for p in cand:
            frac = p["hogares_stock"] / masa_c
            p["nuevas_fam_year"] = round(p["nuevas_fam_year"] + nf * frac, 2)
            p["pool_activo"] = round(p["pool_activo"] + pool * PCT_POOL_ACTIVO * frac, 2)
            p.setdefault("buckets", {})
            p["buckets"][s.get("bucket")] = round(p["buckets"].get(s.get("bucket"), 0) + nf * frac, 2)

    # ── 5 · Producto por perfil: ticket de SU capacidad, m² del programa, $/m² derivado ──
    for p in perfiles:
        cap_lo, cap_hi = p["capacidad_pago_banda_M"]
        p["bucket_principal"] = max(p.get("buckets", {}), key=p.get("buckets", {}).get) \
            if p.get("buckets") else None
        p["ticket_banda_M"] = [cap_lo, cap_hi]
        if cap_lo and cap_hi:
            n_rec = p["programa"]["rec"]
            lo_b, hi_b = _banda_tamano_por_ticket((cap_lo + cap_hi) / 2)
            area_min_fisica = (n_rec * _min_m2_recamara(p["nse"])) / _pct_area_recamaras(p["nse"])
            m2_lo = max(lo_b, math.ceil(area_min_fisica))
            m2_hi = max(hi_b, m2_lo + 5)
            p["m2_banda"] = [m2_lo, m2_hi]
            # Monotonía correcta: a mismo ticket, MÁS m² ⇒ MENOS $/m² (y viceversa)
            p["pm2_derivado_banda"] = [round(cap_lo * 1e6 / m2_hi), round(cap_hi * 1e6 / m2_lo)]
        else:
            p["m2_banda"] = None
            p["pm2_derivado_banda"] = None
        p.pop("buckets", None)
    perfiles.sort(key=lambda p: -p["hogares_stock"])
    nf_perfiles = round(sum(p["nuevas_fam_year"] for p in perfiles), 1)
    meta["conservacion"] = {"nf_buckets": round(nf_buckets, 1), "nf_perfiles": nf_perfiles,
                            "ok": abs(nf_buckets - nf_perfiles) < max(1.0, nf_buckets * 0.01)}
    return {"segmentos": perfiles, "meta": meta}


def _extract_low(rango: str) -> Optional[float]:
    """De '$2,800,000-$3,399,999' o '8,333,333 - 10,000,000' → 2800000.0 (float)."""
    if not rango:
        return None
    import re
    # Captura grupos numéricos con comas/puntos (sin el signo $)
    nums = re.findall(r"[\d][\d,\.]*", str(rango).replace("$", ""))
    if nums:
        try:
            return float(nums[0].replace(",", ""))
        except ValueError:
            return None
    return None


def _extract_high(rango: str) -> Optional[float]:
    if not rango:
        return None
    import re
    nums = re.findall(r"[\d][\d,\.]*", str(rango).replace("$", ""))
    if len(nums) >= 2:
        try:
            return float(nums[1].replace(",", ""))
        except ValueError:
            return None
    return None


# ──────────────────────── Productos venta (Checkpoint B) ────────────────────────
# Tolerancias para definir "comparable DIRECTO" producto-por-producto (regla de negocio).
# Un comparable directo compite en ticket, $/m², programa (recámaras) y tamaño. La calidad/
# equipamiento no es numérica en el dataset; se aproxima por la combinación ticket+pm² (a igual
# programa, mayor ticket/pm² ⇒ mayor calidad). Las tolerancias acotan "mismo producto".
_DIRECTO_TOL_TICKET = 0.15   # ±15% en ticket por unidad
_DIRECTO_TOL_PM2    = 0.15   # ±15% en precio por m²
_DIRECTO_TOL_M2     = 0.20   # ±20% en superficie
VENTAJA_DISENO      = 1.20   # diseñamos para vender ≥20% más rápido que la competencia directa


def _rec_to_int(rec):
    """Convierte un valor de recámaras a entero. Acepta int, float o strings tipo '2 Rec',
    '3-4 Rec', '4 Rec +'. Devuelve el primer número entero hallado, o None si no hay."""
    if rec is None:
        return None
    if isinstance(rec, (int, float)):
        return int(rec)
    m = re.search(r"\d+", str(rec))
    return int(m.group()) if m else None


def _abs_comparables_directos(ticket_M, pm2, rec, m2, ft: List[Dict]) -> Dict[str, Any]:
    """Identifica las tipologías de oferta COMPARABLES DIRECTAS del producto recomendado: mismo
    programa de recámaras y precio/m²/ticket/tamaño dentro de tolerancia (producto-por-producto,
    no por zona). Devuelve:
      • abs_vals       → absorciones reales (Abs_Demanda) de los directos → su mediana es el DATO
                         de validación que se muestra a la derecha (no entra al cálculo).
      • abs_compet     → Σ Abs_Demanda (un/mes) de los directos que AÚN tienen inventario
                         disponible > 0. Es el FLUJO competidor que seguirá consumiendo demanda;
                         se RESTA de la demanda mensual (flujo contra flujo). Los proyectos
                         agotados (disp=0) NO se cuentan: ya no competirán a futuro.
      • disp_total     → Σ unidades DISPONIBLES de los directos (referencia / tope de mercado).
      • n              → cantidad de tipologías/productos competidores directos.
    Todo proviene de `ft` traído EN VIVO del API (vvv/query) en cada análisis; nunca cacheado.
    """
    if not ft:
        return {"abs_vals": [], "abs_compet": 0.0, "disp_total": 0, "n": 0, "pm2_vals": []}
    rec_i = _rec_to_int(rec)
    abs_vals = []
    pm2_vals = []
    abs_compet = 0.0
    disp_total = 0
    n = 0
    for f in ft:
        a = f.get("attributes", f)
        c_ticket = _price(a.get("F____UNIDAD"))      # ticket por unidad (M se compara abajo)
        c_pm2 = _pm2v(a.get("F___M2"))
        c_m2 = _num(a.get("ÁREA_TOTAL")) or _num(a.get("ÁREA_PRIVATIVA"))
        c_rec = _rec_to_int(a.get("CANTIDAD_DE_RECAMARAS"))
        # Programa: mismo número de recámaras (criterio duro del comparable directo)
        if rec_i is not None and c_rec is not None and c_rec != rec_i:
            continue
        # Ticket (en M): dentro de tolerancia
        if ticket_M and c_ticket:
            if abs(c_ticket / 1e6 - ticket_M) / ticket_M > _DIRECTO_TOL_TICKET:
                continue
        # Precio por m²: dentro de tolerancia
        if pm2 and c_pm2:
            if abs(c_pm2 - pm2) / pm2 > _DIRECTO_TOL_PM2:
                continue
        # Tamaño: dentro de tolerancia
        if m2 and c_m2:
            if abs(c_m2 - m2) / m2 > _DIRECTO_TOL_M2:
                continue
        # Es un comparable directo. Cuenta para n.
        n += 1
        c_disp = _num(a.get("UNIDADES_DISPONIBLES"))
        c_abs = _num(a.get("Abs_Demanda"))
        if c_pm2 and c_pm2 > 0:
            pm2_vals.append(c_pm2)   # $/m² real del directo → mediana = ancla del precio recomendado
        if c_disp is not None and c_disp > 0:
            disp_total += c_disp
            # Solo los que AÚN tienen inventario seguirán compitiendo → su absorción consume demanda.
            if c_abs is not None and c_abs > 0:
                abs_compet += c_abs
        # La mediana de absorción (dato de validación) usa todos los directos con Abs_Demanda>0.
        if c_abs is not None and c_abs > 0:
            abs_vals.append(c_abs)
    return {"abs_vals": abs_vals, "abs_compet": round(abs_compet, 4),
            "disp_total": disp_total, "n": n, "pm2_vals": pm2_vals}


# Tolerancias del PRECIO RECOMENDADO (regla de negocio del Monitor, tomada del tablero estático
# findCompetitorsForMixItem). Son DISTINTAS y más amplias que las del comparable directo de
# absorción: aquí NO se filtra por $/m² (el $/m² es justo lo que vamos a recomendar; filtrarlo
# sería circular). Un competidor directo de precio comparte ticket, tamaño y programa aproximado.
_PRECIO_TOL_TICKET = 0.15   # ±15% en ticket por unidad
_PRECIO_TOL_AREA   = 0.30   # ±30% en superficie
_PRECIO_TOL_REC    = 1      # ±1 recámara


def _precio_recomendado_directos(ticket_M, rec, m2, ft: List[Dict]) -> Dict[str, Any]:
    """PRECIO RECOMENDADO del producto a partir de la competencia directa (regla de Monitor,
    replicada del tablero estático findCompetitorsForMixItem). Identifica las tipologías de la
    oferta real (ft, EN VIVO) que compiten directamente con el producto:
      • Solo con inventario DISPONIBLE > 0 (las agotadas ya no marcan precio de mercado).
      • Ticket dentro de ±15%, área dentro de ±30%, recámaras dentro de ±1.
      • NO se filtra por $/m²: es lo que se va a recomendar.
    El precio recomendado $/m² = MEDIANA de los $/m² de esos competidores directos (ancla al
    mercado observado). El ticket recomendado = pm²_recomendado × m² del producto.
    Devuelve {pm2_recomendado, ticket_recomendado_M, n}. Sin directos → None (N/D, no se inventa).
    """
    if not ft or not ticket_M or ticket_M <= 0:
        return {"pm2_recomendado": None, "ticket_recomendado_M": None, "n": 0}
    rec_i = _rec_to_int(rec)
    pm2_vals = []
    for f in ft:
        a = f.get("attributes", f)
        c_disp = _num(a.get("UNIDADES_DISPONIBLES"))
        if c_disp is None or c_disp <= 0:
            continue   # solo inventario que aún compite marca el precio de mercado
        c_ticket = _price(a.get("F____UNIDAD"))
        c_pm2 = _pm2v(a.get("F___M2"))
        if not c_ticket or not c_pm2 or c_ticket <= 100000:
            continue   # precio inválido / N/D
        # Ticket dentro de ±15%
        if abs(c_ticket / 1e6 - ticket_M) / ticket_M > _PRECIO_TOL_TICKET:
            continue
        # Área dentro de ±30% (si ambos tienen dato)
        c_m2 = _num(a.get("ÁREA_TOTAL")) or _num(a.get("ÁREA_PRIVATIVA"))
        if c_m2 and m2 and m2 > 0:
            if abs(c_m2 - m2) / m2 > _PRECIO_TOL_AREA:
                continue
        # Recámaras dentro de ±1 (si ambos tienen dato)
        c_rec = _rec_to_int(a.get("CANTIDAD_DE_RECAMARAS"))
        if rec_i is not None and c_rec is not None and abs(c_rec - rec_i) > _PRECIO_TOL_REC:
            continue
        pm2_vals.append(c_pm2)
    if not pm2_vals:
        return {"pm2_recomendado": None, "ticket_recomendado_M": None, "n": 0}
    pm2_rec = round(statistics.median(pm2_vals))
    ticket_rec_M = round(pm2_rec * m2 / 1e6, 2) if (m2 and pm2_rec) else None
    return {"pm2_recomendado": pm2_rec, "ticket_recomendado_M": ticket_rec_M, "n": len(pm2_vals)}


# Umbral del veredicto caro/barato/en línea del precio ingresado vs el recomendado.
PRECIO_TOL_VEREDICTO = 0.05   # ±5%


PCT_POOL_ACTIVO = 0.05   # fracción del pool de mercado en venta del segmento que se anualiza


# ════════════════════════ MONITOR · AMENAZA COMPETITIVA (regla de negocio) ════════════════════════
# Toda la lógica del Monitor vive aquí (antes estaba duplicada en el front: findCompetitorsForMixItem,
# findMatchingSegment, computeCompetitiveThreat). El front solo arma el mix interactivo y muestra.
# Mismas tolerancias que el precio recomendado (regla del tablero estático findCompetitorsForMixItem).
# H11 · CONSOLIDACIÓN: los _MON_TOL_* ahora REFERENCIAN los _PRECIO_TOL_* (misma regla de
# comparable directo de precio). Ya no pueden divergir; se conservan los nombres porque el
# Monitor los referencia (catálogo: no se renombra lo que otros consumen).
_MON_TOL_TICKET = _PRECIO_TOL_TICKET   # ±15% en ticket por unidad
_MON_TOL_AREA   = _PRECIO_TOL_AREA     # ±30% en superficie
_MON_TOL_REC    = _PRECIO_TOL_REC      # ±1 recámara


def _competidores_mix_item(item: Dict[str, Any], typologies: Dict[str, List[Dict]]) -> List[Dict[str, Any]]:
    """Competidores DIRECTOS de un producto del mix del Monitor, sobre las tipologías reales de la
    zona (typologies que el backend ya entregó en el payload). Solo inventario disponible, precio
    válido, ticket ±15%, área ±30%, recámaras ±1. Misma regla que el precio recomendado."""
    ticket_M = _num(item.get("ticket_M")) or 0
    if ticket_M <= 0 or not typologies:
        return []
    item_area = _num(item.get("m2")) or 60
    item_rec = _rec_to_int(item.get("rec"))
    comp = []
    for proj, typs in (typologies or {}).items():
        for t in (typs or []):
            disp = _num(t.get("unid_disp"))
            if disp is None or disp <= 0:
                continue
            pud = _num(t.get("precio_ud"))
            if pud is None or pud <= 100000:
                continue
            if abs(pud / 1e6 - ticket_M) / ticket_M > _MON_TOL_TICKET:
                continue
            t_area = _num(t.get("area_priv")) or 0
            if t_area > 0 and item_area > 0 and abs(t_area - item_area) / item_area > _MON_TOL_AREA:
                continue
            t_rec = _rec_to_int(t.get("rec"))
            if item_rec is not None and t_rec is not None and abs(t_rec - item_rec) > _MON_TOL_REC:
                continue
            comp.append({
                "project": proj, "tipo": t.get("tipo"), "precio_ud": pud,
                "precio_m2": _num(t.get("precio_m2")), "area_priv": t_area,
                "rec": t_rec, "unid_disp": disp, "abs": _num(t.get("abs")) or 0,
            })
    return comp


def _segmento_para_ticket(ticket_M: float, segments: List[Dict]) -> Optional[Dict]:
    """Segmento de demanda cuyo rango de valor contiene el ticket del producto (M)."""
    if not segments or not ticket_M or ticket_M <= 0:
        return None
    for s in segments:
        lo = (s.get("val_min") or 0) / 1e6
        hi = (s.get("val_max") or 0) / 1e6
        if lo <= ticket_M <= hi:
            return s
    return None


def _precio_recomendado_mix_item(item: Dict[str, Any], typologies: Dict[str, List[Dict]]) -> Dict[str, Any]:
    """PRECIO RECOMENDADO de un producto del mix del Monitor, sobre las tipologías reales de la zona
    (typologies, mismo formato que usa _competidores_mix_item). Misma regla que el precio recomendado
    de la sección de producto: directos = inventario disponible>0, ticket ±15%, área ±30%, rec ±1, sin
    filtrar por $/m²; pm²_recomendado = MEDIANA de $/m² de esos directos; ticket_rec = pm²_rec × m².
    Consistente para vertical y horizontal (opera sobre la misma estructura de oferta). Sin directos
    → None (N/D, no se inventa)."""
    ticket_M = _num(item.get("ticket_M")) or 0
    m2 = _num(item.get("m2")) or 0
    rec_i = _rec_to_int(item.get("rec"))
    if not typologies or ticket_M <= 0:
        return {"pm2_recomendado": None, "ticket_recomendado_M": None, "n": 0}
    pm2_vals = []
    for proj, typs in (typologies or {}).items():
        for t in (typs or []):
            disp = _num(t.get("unid_disp"))
            if disp is None or disp <= 0:
                continue
            pud = _num(t.get("precio_ud"))
            pm2 = _num(t.get("precio_m2"))
            if not pud or not pm2 or pud <= 100000:
                continue
            if abs(pud / 1e6 - ticket_M) / ticket_M > _PRECIO_TOL_TICKET:
                continue
            t_area = _num(t.get("area_priv")) or 0
            if t_area > 0 and m2 and m2 > 0 and abs(t_area - m2) / m2 > _PRECIO_TOL_AREA:
                continue
            t_rec = _rec_to_int(t.get("rec"))
            if rec_i is not None and t_rec is not None and abs(t_rec - rec_i) > _PRECIO_TOL_REC:
                continue
            pm2_vals.append(pm2)
    if not pm2_vals:
        return {"pm2_recomendado": None, "ticket_recomendado_M": None, "n": 0}
    pm2_rec = round(statistics.median(pm2_vals))
    ticket_rec_M = round(pm2_rec * m2 / 1e6, 2) if (m2 and pm2_rec) else None
    return {"pm2_recomendado": pm2_rec, "ticket_recomendado_M": ticket_rec_M, "n": len(pm2_vals)}


def amenaza_competitiva(item: Dict[str, Any], period: int, capture: float,
                        typologies: Dict[str, List[Dict]], segments: List[Dict]) -> Dict[str, Any]:
    """REGLA DE NEGOCIO del Monitor: dado un producto del mix (item), el horizonte (period, meses) y
    la tasa de captación (capture), calcula la amenaza competitiva contra la oferta directa real:
      • threat_ratio = inventario competidor disponible / demanda del horizonte.
      • net_demand   = demanda potencial del producto − absorción competidora en el horizonte.
      • strategy/action/color: expansión / acelerar / monitorear / aguantar / reposicionar.
    El front NO recalcula nada de esto; solo envía el mix interactivo y muestra el resultado.
    """
    comp = _competidores_mix_item(item, typologies)
    seg = _segmento_para_ticket(_num(item.get("ticket_M")) or 0, segments)
    comp_units = sum(c["unid_disp"] for c in comp)
    comp_abs_month = sum(c["abs"] for c in comp)
    comp_projects = len({c["project"] for c in comp})
    months_to_deplete = (comp_units / comp_abs_month) if comp_abs_month > 0 else None
    cr = capture if isinstance(capture, (int, float)) else 1.0
    new_fam_period = (seg["nuevas_fam"] * (period / 12)) if seg else 0
    item_potential = new_fam_period * cr
    threat_ratio = (comp_units / new_fam_period) if new_fam_period > 0 else (99 if comp_units > 0 else 0)
    comp_absorbed = min(comp_units, comp_abs_month * period)
    net_demand = max(0, item_potential - comp_absorbed)
    item_units = _num(item.get("units")) or 0
    item_abs_month = (seg["nuevas_fam"] / 12 * cr) if seg else 0
    item_months_to_sell = (item_units / item_abs_month) if item_abs_month > 0 else None
    mtd = round(months_to_deplete) if months_to_deplete is not None else None
    if comp_units == 0:
        strategy = "Mercado libre · sin competidores directos · expansión sin obstáculos"
        color, action = "green", "expansion"
    elif threat_ratio > 2.0:
        strategy = f"Ajustar precio o reposicionar · {comp_units:,} ud competidoras ({threat_ratio:.1f}× la demanda de {period}m) · océano rojo"
        color, action = "red", "reprice"
    elif threat_ratio > 1.0:
        strategy = f"Aguantar precio · {mtd}m para agotar competencia · esperar absorción del mercado"
        color, action = "amber", "hold"
    elif threat_ratio > 0.5:
        strategy = f"Mercado equilibrado · monitorear · {mtd}m absorción competidora"
        color, action = "blue", "monitor"
    else:
        strategy = f"Acelerar comercialización · publicidad e incentivos al equipo · ventaja clara ({round(threat_ratio*100)}% del horizonte)"
        color, action = "green", "accelerate"
    # PRECIO RECOMENDADO del producto (consistente vertical/horizontal): mediana $/m² de los
    # directos reales. Mismo cálculo que la sección de producto; el front lo muestra como veredicto.
    precio_rec = _precio_recomendado_mix_item(item, typologies)
    pm2_usuario = None
    _m2 = _num(item.get("m2")) or 0
    if _m2 > 0 and (_num(item.get("ticket_M")) or 0) > 0:
        pm2_usuario = round((_num(item.get("ticket_M")) or 0) * 1e6 / _m2)
    veredicto_precio = None
    if precio_rec.get("pm2_recomendado") and pm2_usuario:
        d = (pm2_usuario - precio_rec["pm2_recomendado"]) / precio_rec["pm2_recomendado"]
        if d > PRECIO_TOL_VEREDICTO:
            veredicto_precio = "caro"
        elif d < -PRECIO_TOL_VEREDICTO:
            veredicto_precio = "barato"
        else:
            veredicto_precio = "en_linea"
    return {
        "competitors": comp, "matching_seg": seg,
        "comp_units": comp_units, "comp_abs_month": round(comp_abs_month, 4),
        "comp_projects": comp_projects,
        "months_to_deplete": (round(months_to_deplete, 1) if months_to_deplete is not None else None),
        "threat_ratio": round(threat_ratio, 4), "new_families_period": round(new_fam_period, 2),
        "item_potential_demand": round(item_potential, 2),
        "comp_absorbed_in_period": round(comp_absorbed, 2), "net_demand": round(net_demand, 2),
        "item_units": item_units, "item_abs_month": round(item_abs_month, 4),
        "item_months_to_sell": (round(item_months_to_sell, 1) if item_months_to_sell is not None else None),
        "strategy": strategy, "color": color, "action": action,
        # Precio recomendado (mismo cálculo que la sección de producto)
        "pm2_recomendado": precio_rec.get("pm2_recomendado"),
        "ticket_recomendado_M": precio_rec.get("ticket_recomendado_M"),
        "precio_n_directos": precio_rec.get("n", 0),
        "pm2_usuario": pm2_usuario,
        "precio_tol_veredicto": PRECIO_TOL_VEREDICTO,
        "veredicto_precio": veredicto_precio,
    }



# Curva de maduración del pronóstico de venta (factores del ritmo base por tramo de meses).
RAMP_ARRANQUE   = 0.60   # meses 1–6   (preventa / lanzamiento)
RAMP_CONSOLIDA  = 0.70   # meses 7–18  (consolidación)
RAMP_COLA       = 0.20   # meses 19–24 (cola / agotamiento)


def _factor_curva(mes: int) -> float:
    """Factor de maduración del ritmo de venta según el mes del proyecto (tramos fijos)."""
    if mes <= 6:
        return RAMP_ARRANQUE
    if mes <= 18:
        return RAMP_CONSOLIDA
    return RAMP_COLA


def pronostico_ventas(abs_mensual: float, horizonte_meses: int,
                      unidades_proyecto: Optional[int] = None) -> Dict[str, Any]:
    """Pronóstico acumulado de venta del proyecto nuevo a N meses (12/18/24 del selector).

        ventas(N) = min( Σ mes=1..N [ abs_mensual × factor_curva(mes) ] , unidades_proyecto )

    • abs_mensual        → ritmo base captable por el proyecto (absorción flujo-contra-flujo).
    • factor_curva       → 60% (m1-6), 70% (m7-18), 20% (m19-24).
    • unidades_proyecto  → tope de inventario propio (campo del front). Si None, sin tope.
    Devuelve acumulado, ritmo mensual efectivo y mes de agotamiento (si aplica el tope).
    """
    if not abs_mensual or abs_mensual <= 0:
        return {"horizonte": horizonte_meses, "acumulado": 0.0, "mensual": [],
                "mes_agotamiento": None, "topado": False}
    acumulado = 0.0
    mensual = []
    mes_agotamiento = None
    topado = False
    for mes in range(1, horizonte_meses + 1):
        ritmo = abs_mensual * _factor_curva(mes)
        if unidades_proyecto is not None and acumulado + ritmo >= unidades_proyecto:
            ritmo = max(0.0, unidades_proyecto - acumulado)
            acumulado += ritmo
            mensual.append(round(ritmo, 2))
            if mes_agotamiento is None and acumulado >= unidades_proyecto:
                mes_agotamiento = mes
                topado = True
            # Una vez agotado el inventario propio, los meses siguientes venden 0.
            for _ in range(mes + 1, horizonte_meses + 1):
                mensual.append(0.0)
            break
        acumulado += ritmo
        mensual.append(round(ritmo, 2))
    return {"horizonte": horizonte_meses, "acumulado": round(acumulado, 1),
            "mensual": mensual, "mes_agotamiento": mes_agotamiento, "topado": topado}


def _absorcion_producto(nuevas_fam_year: float, demanda_total: float,
                        directos: Dict[str, Any],
                        unidades_proyecto: Optional[int] = None) -> Dict[str, Any]:
    """Absorción estimada de UN producto (un/mes), regla de negocio definitiva (flujo-contra-flujo):

        numerador_anual = nuevas_familias_anual + (demanda_total × 0.05)
            (ambos del MISMO bucket de precio comparable, no del acumulado de la zona)
        demanda_mensual = numerador_anual / 12

        • SIN comparables directos → absorción = demanda_mensual
        • CON comparables directos → absorción = demanda_mensual − Σ Abs_Demanda de los
          competidores directos QUE AÚN tienen inventario disponible (flujo competidor que
          seguirá consumiendo demanda). El proyecto nuevo capta lo que la competencia deja libre.

        · numerador_anual ≤ 0  → N/D, no se recomienda (sin demanda en el bucket).
        · absorción ≤ 0        → sobreofertado (la competencia activa consume toda la demanda).

    Además devuelve el pronóstico acumulado a 12/18/24 meses (curva de maduración + tope de
    inventario del proyecto). La mediana de absorción de los directos NO entra al cálculo: es
    dato de validación (producto ganador vs mediocre).
    """
    abs_vals = directos.get("abs_vals", [])
    abs_compet = directos.get("abs_compet", 0.0) or 0.0
    disp_total = directos.get("disp_total", 0) or 0
    n_directos = directos.get("n", 0)
    mediana = round(statistics.median(abs_vals), 2) if abs_vals else None

    numerador = (nuevas_fam_year or 0) + (demanda_total or 0) * PCT_POOL_ACTIVO
    if numerador <= 0:
        # Sin demanda DIM en el bucket de precio. PERO si hay comparables DIRECTOS con absorción
        # real (competidores vendiendo activamente), el mercado SÍ existe aunque INEGI no reporte
        # demanda en ese rango: la evidencia de venta de la oferta lo demuestra. Regla de negocio:
        # absorción derivada de la OFERTA = mediana de absorción de los directos × nº de directos
        # (el potencial del rango lo marca lo que la competencia comparable ya está absorbiendo).
        if mediana is not None and n_directos > 0:
            abs_oferta = round(mediana * n_directos, 2)
            pron = {h: pronostico_ventas(abs_oferta, h, unidades_proyecto) for h in (12, 18, 24)}
            return {"abs": abs_oferta, "origen": "absorcion_oferta_directa", "n_directos": n_directos,
                    "mediana_directos": mediana, "abs_competidores": round(abs_compet, 2),
                    "inventario_competidor": disp_total, "pronostico": pron}
        # Sin demanda y sin directos con absorción → no hay base; el producto no se recomienda.
        return {"abs": None, "origen": "sin_demanda", "n_directos": n_directos,
                "mediana_directos": mediana, "abs_competidores": round(abs_compet, 2),
                "inventario_competidor": disp_total, "pronostico": {}}

    demanda_mensual = numerador / 12.0
    if n_directos > 0:
        abs_proyecto = demanda_mensual - abs_compet   # flujo contra flujo
        origen = "comparables_directos"
    else:
        abs_proyecto = demanda_mensual
        origen = "demanda_sin_competencia"

    if abs_proyecto <= 0:
        # La demanda DIM la consume toda la competencia activa (flujo-contra-flujo ≤ 0). PERO si hay
        # comparables DIRECTOS con absorción real, el rango SÍ tiene mercado (lo prueban los
        # competidores vendiendo). Regla de negocio: absorción derivada de la OFERTA = mediana de
        # absorción de los directos × nº de directos. Se marca sobreofertado (el motivo se muestra),
        # pero se entrega la absorción de referencia en vez de un 0/N/D que oculta el mercado real.
        if mediana is not None and n_directos > 0:
            abs_oferta = round(mediana * n_directos, 2)
            pron = {h: pronostico_ventas(abs_oferta, h, unidades_proyecto) for h in (12, 18, 24)}
            return {"abs": abs_oferta, "origen": "sobreofertado", "n_directos": n_directos,
                    "mediana_directos": mediana, "abs_competidores": round(abs_compet, 2),
                    "inventario_competidor": disp_total, "pronostico": pron}
        # Sobreofertado y sin directos con absorción medible → 0 explícito (no N/D).
        return {"abs": 0.0, "origen": "sobreofertado", "n_directos": n_directos,
                "mediana_directos": mediana, "abs_competidores": round(abs_compet, 2),
                "inventario_competidor": disp_total, "pronostico": {}}

    abs_proyecto = round(abs_proyecto, 2)
    pron = {h: pronostico_ventas(abs_proyecto, h, unidades_proyecto) for h in (12, 18, 24)}
    return {
        "abs": abs_proyecto,
        "origen": origen,
        "n_directos": n_directos,
        "mediana_directos": mediana,           # dato de validación (no entra al cálculo)
        "abs_competidores": round(abs_compet, 2),
        "inventario_competidor": disp_total,
        "pronostico": pron,                    # {12:{...},18:{...},24:{...}}
    }


def _recomendable_por_absorcion(abs_origen: str) -> Dict[str, Any]:
    """Define si un producto es recomendable según el origen de su absorción (la nueva fórmula
    flujo-contra-flujo manda sobre el status demográfico). Devuelve {recomendable, motivo}.

      • comparables_directos / demanda_sin_competencia → recomendable (hay absorción positiva).
      • sin_demanda    → NO recomendable: no hay demanda en el bucket de precio (nadie compra ahí).
      • sobreofertado  → NO recomendable: la competencia activa ya consume toda la demanda.
    El motivo se muestra en el front en vez de un 'N/D' a secas.
    """
    if abs_origen == "sin_demanda":
        return {"recomendable": False, "motivo": "Sin demanda en el rango de precio"}
    if abs_origen == "sobreofertado":
        return {"recomendable": False, "motivo": "Sobreofertado · competencia satura la demanda"}
    # comparables_directos, demanda_sin_competencia, absorcion_oferta_directa u otros con absorción positiva
    return {"recomendable": True, "motivo": None}


def derive_productos_horizontal(ft: List[Dict], segments: List[Dict], unidades_proyecto: Optional[int] = None) -> List[Dict[str, Any]]:
    """
    PRODUCTO DE VIVIENDA HORIZONTAL (casa unifamiliar en venta).

    Replica EXACTAMENTE la metodología DPO de vivienda vertical (misma absorción canónica
    nuevas_fam/12 × captación, mismo anclaje del programa al ticket, recomendado⊆aplicable),
    cambiando solo lo propio de la CASA:
      • DOS áreas: ÁREA_CONSTRUCCIÓN (la que paga el cliente, tamaño principal) y
        ÁREA_TERRENO (dato adicional del lote). En vertical era ÁREA_PRIVATIVA.
      • Bandas de tamaño de CASA (no depa): una casa es más grande que un depa al mismo
        nivel. Anclado a datos reales vh_venta: económica 2 rec ~58-110 m² constr (Valle de
        Lincoln 58/53 NSE D+ $533k); premium 3-4 rec 150-280 m² (Los Cenizos 212/150 NSE A
        $10M; Privada Encinos 270 NSE A $11.5M).
      • m² de CONSTRUCCIÓN observado real de vh_venta manda si existe.
    """
    color_by_status = {
        "sweet_spot": "green", "desatendido": "blue", "oportunidad": "purple",
        "atendido": "teal", "oceano_rojo": "red", "bajo_crecimiento": "amber",
    }
    # Precio/m² de construcción de referencia por NSE para CASA (MXN). Menor que vertical
    # porque la casa reparte valor entre construcción y terreno (no todo es construcción).
    PM2_POR_NSE = {"A": 95000, "B": 70000, "C+": 45000, "C": 32000,
                   "D+": 24000, "D": 18000, "E": 14000}

    def recamaras_por_m2(m2):
        if m2 is None: return "N/D"
        if m2 < 70:  return "2 Rec"
        if m2 < 110: return "2-3 Rec"
        if m2 < 160: return "3 Rec"
        if m2 < 230: return "3-4 Rec"
        return "4 Rec"

    # PROGRAMA DE RECÁMARAS de CASA anclado al TICKET (monotónico). Una casa parte de 2 rec
    # (no hay studios/1 rec en unifamiliar); crece a 3, 3-4 y 4+ con el valor.
    def recamaras_por_ticket(ticket_m, m2=None):
        if ticket_m is None:
            return recamaras_por_m2(m2)
        if ticket_m <= 1.0:   return "2 Rec"      # económica (Valle de Lincoln tipo)
        if ticket_m <= 2.0:   return "2 Rec"
        if ticket_m <= 3.5:   return "3 Rec"
        if ticket_m <= 6.0:   return "3 Rec"
        if ticket_m <= 10.0:  return "3-4 Rec"
        if ticket_m <= 16.0:  return "4 Rec"
        return "4 Rec +"

    # Bandas de tamaño de CONSTRUCCIÓN de casa por ticket (MXN), de datos reales vh_venta:
    #   ≤ 1.0M  → 2 Rec · 55-75 m²    (económica: Valle de Lincoln 58 m²)
    #   ≤ 2.0M  → 2 Rec · 70-95 m²
    #   ≤ 3.5M  → 3 Rec · 90-130 m²
    #   ≤ 6.0M  → 3 Rec · 120-170 m²
    #   ≤ 10.0M → 3-4 Rec · 160-220 m² (Los Cenizos 212 m²)
    #   ≤ 16.0M → 4 Rec · 210-290 m²   (Privada Encinos 270 m²)
    #   > 16.0M → 4 Rec+ · 280-400 m²
    def _banda_constr(ticket_m):
        if ticket_m is None:   return (90, 130)
        if ticket_m <= 1.0:    return (55, 75)
        if ticket_m <= 2.0:    return (70, 95)
        if ticket_m <= 3.5:    return (90, 130)
        if ticket_m <= 6.0:    return (120, 170)
        if ticket_m <= 10.0:   return (160, 220)
        if ticket_m <= 16.0:   return (210, 290)
        return (280, 400)

    # Relación terreno/construcción típica por NSE: en económica el terreno es ~igual o menor
    # (casa de 2 plantas en lote chico); en premium el terreno crece (más jardín/frente).
    def _ratio_terreno(nse):
        return {"A": 0.95, "B": 0.85, "C+": 0.80, "C": 0.85, "D+": 0.90, "D": 0.95, "E": 1.0}.get(nse, 0.85)

    # Competencia directa por bucket (tipologías de vh_venta en el rango de precio)
    competencia = {}
    for t in ft:
        a = t.get("attributes", {})
        precio = _price(a.get("F____UNIDAD"))
        if not precio:
            continue
        for s in segments:
            if s.get("val_min") and s.get("val_max") and s["val_min"] <= precio < s["val_max"]:
                competencia[s["bucket"]] = competencia.get(s["bucket"], 0) + 1
                break

    productos = []
    for s in segments:
        vmin = s.get("val_min"); vmax = s.get("val_max")
        val_mid = None
        if vmin is not None and vmax is not None and (vmin or vmax):
            if vmin and vmax:
                val_mid = (float(vmin) + float(vmax)) / 2
            elif vmax:
                # Bucket inferior (val_min=0): NO usar el techo del bucket como precio. El ticket
                # debe reflejar la CAPACIDAD DE PAGO REAL de la demanda (ingreso del segmento ×
                # 12 × múltiplo hipotecario), no el tope del rango. Evita productos por encima de
                # lo que la población puede pagar (p. ej. casa $1.12M en zona con capacidad $0.66M).
                MULTIPLO_HIPOTECARIO = 4.5
                ing_rep = s.get("ing_min")   # ingreso mensual representativo del segmento (base real)
                cap_compra = (ing_rep * 12 * MULTIPLO_HIPOTECARIO) if ing_rep else None
                if cap_compra:
                    val_mid = min(cap_compra, float(vmax) * 0.75)  # capado al techo del bucket
                else:
                    val_mid = float(vmax) * 0.5   # sin ingreso: punto medio-bajo conservador
            elif vmin:
                val_mid = float(vmin) * 1.1
        ticket_M = round(val_mid / 1e6, 2) if val_mid else None
        if ticket_M and vmin:
            ticket_M = max(ticket_M, round(float(vmin) / 1e6, 2))

        # CONSTRUCCIÓN observada real de vh_venta en el rango (preferida)
        PM2_CASA_MIN = 8000
        constr_obs, terreno_obs, pm2_obs = [], [], []
        for t in ft:
            a = t.get("attributes", {})
            precio = _price(a.get("F____UNIDAD"))
            ac = _num(a.get("ÁREA_CONSTRUCCIÓN"))
            at = _num(a.get("ÁREA_TERRENO"))
            pm2v = _pm2(a.get("F___M2"))
            if precio and val_mid and abs(precio - val_mid) / val_mid < 0.20:
                if ac and M2_MIN <= ac <= M2_MAX: constr_obs.append(ac)
                if at and M2_MIN <= at <= M2_MAX: terreno_obs.append(at)
                if pm2v and pm2v >= PM2_CASA_MIN: pm2_obs.append(pm2v)
        constr_zona = round(statistics.median(constr_obs)) if constr_obs else None
        terreno_zona = round(statistics.median(terreno_obs)) if terreno_obs else None
        pm2_zona = round(statistics.median(pm2_obs)) if pm2_obs else None

        # TAMAÑO DE CONSTRUCCIÓN: 1º observado real; 2º banda del programa por NSE.
        lo_b, hi_b = _banda_constr(ticket_M)
        if constr_zona is not None:
            m2 = max(M2_MIN, min(M2_MAX, constr_zona))
        else:
            pos = {"A": 0.85, "B": 0.65, "C+": 0.5, "C": 0.4, "D+": 0.3, "D": 0.25, "E": 0.2}.get(s["NSE"], 0.5)
            m2 = max(M2_MIN, min(M2_MAX, round(lo_b + (hi_b - lo_b) * pos)))

        # TERRENO: observado real si existe; si no, derivado de la construcción por ratio NSE.
        if terreno_zona is not None:
            m2_terreno = max(M2_MIN, min(M2_MAX, terreno_zona))
        else:
            m2_terreno = round(m2 * _ratio_terreno(s["NSE"]))

        # pm² de construcción derivado (ticket/m²); ancla al observado si existe.
        if ticket_M and m2:
            pm2 = round(ticket_M * 1e6 / m2)
        elif pm2_zona:
            pm2 = pm2_zona
        else:
            pm2 = PM2_POR_NSE.get(s["NSE"], 35000)

        # ABSORCIÓN (regla de negocio): si hay comparables DIRECTOS (mismo programa/ticket/pm²/
        # tamaño), absorción = mediana real de los directos × 1.20 (vender 20% más rápido). Si NO
        # hay directos, el producto captura el 100% de la demanda mensual del segmento (sin
        # competencia → se diseña para ser dominante). Sin techo artificial.
        n_comp = competencia.get(s["bucket"], 0)
        abs_dir = _abs_comparables_directos(ticket_M, pm2, recamaras_por_ticket(ticket_M, m2), m2, ft)
        precio_rec = _precio_recomendado_directos(ticket_M, recamaras_por_ticket(ticket_M, m2), m2, ft)
        abs_info = _absorcion_producto(s.get("nuevas_fam", 0) or 0, s.get("demanda_total", 0) or 0, abs_dir, unidades_proyecto)
        abs_rate = abs_info["abs"]

        mix_pct = None
        tot_nf = sum(x.get("nuevas_fam", 0) for x in segments) or 0
        if tot_nf and s.get("nuevas_fam"):
            mix_pct = round(s["nuevas_fam"] / tot_nf * 100)
        nse_tca = NSE_TCA.get(s["NSE"], 0)   # TCA canónica (fuente única); antes tabla hardcodeada errónea

        productos.append({
            "tipo": f"{recamaras_por_ticket(ticket_M, m2)} · {s['bucket']}" if m2 else f"{s['NSE']} · {s['bucket']}",
            "badge": f"{mix_pct}%" if mix_pct is not None else "—",
            "color": color_by_status.get(s["status"], "teal"),
            "rec": recamaras_por_ticket(ticket_M, m2),
            "m2": f"{m2} m²" if m2 else "N/D",                     # construcción (tamaño principal)
            "m2_construccion": f"{m2} m²" if m2 else "N/D",
            "m2_terreno": f"{m2_terreno} m²" if m2_terreno else "N/D",
            "pm2": f"${pm2:,}" if pm2 else "N/D",
            "ticket": f"${ticket_M:.2f}M" if ticket_M else "N/D",
            "abs": f"{abs_rate:.1f} un/mes" if abs_rate else "N/D",
            "abs_origen": abs_info["origen"],
            "abs_n_directos": abs_info["n_directos"],
            "abs_mediana_directos": abs_info["mediana_directos"],
            "abs_inv_competidor": abs_info.get("inventario_competidor", 0),
            "abs_competidores": abs_info.get("abs_competidores", 0),
            "abs_pronostico": abs_info.get("pronostico", {}),
            "pm2_recomendado": precio_rec.get("pm2_recomendado"),
            "ticket_recomendado_M": precio_rec.get("ticket_recomendado_M"),
            "precio_tol_veredicto": PRECIO_TOL_VEREDICTO,
            # Consistencia con vertical (catálogo universal): mismos nombres en ambos modos.
            "tca": nse_tca,
            "competidores": n_comp,
            "mercado": f"NSE {s['NSE']} · {s['bucket']} · {n_comp} competidores directos",
            "status": s["status"],
            "recomendado": ((s.get("aplicable", True)
                            and (s["status"] in ("sweet_spot", "desatendido", "oportunidad", "atendido")))
                           or s.get("dual_featured", False))
                           and _recomendable_por_absorcion(abs_info["origen"])["recomendable"],
            "no_recomendable_motivo": _recomendable_por_absorcion(abs_info["origen"])["motivo"],
            "aplicable": s.get("aplicable", True),
            "featured": (s["status"] == "sweet_spot" or s.get("dual_featured", False))
                        and _recomendable_por_absorcion(abs_info["origen"])["recomendable"],
            "seg_dim": f"{s['NSE']} · {s['bucket']}",
            "mkt_segmento": s["mkt_total"], "nuevas_fam": s["nuevas_fam"],
            "categoria": _status_label(s["status"]),
            "perfiles": _perfiles_por_segmento(s["NSE"], m2, ticket_M),
            "nota_mercado": s.get("nota_mercado"),
            "nuevas_fam_year": s.get("nuevas_fam", 0),
            "depth": s.get("mkt_total", 0),
            "abs_num": round(abs_rate, 2) if abs_rate else 0.0,
            "pm2_num": pm2,
            "m2_num": m2,
            "m2_terreno_num": m2_terreno,
            "ticket_num": round(ticket_M * 1e6) if ticket_M else None,
            "mix_num": mix_pct if mix_pct is not None else 0,
        })
    # Un solo featured (idéntico a vertical)
    feats = [p for p in productos if p.get("featured")]
    if len(feats) > 1:
        best = max(feats, key=lambda p: p.get("nuevas_fam", 0))
        for p in feats:
            if p is not best:
                p["featured"] = False
    elif not feats:
        def _abs_num(p):
            try:
                return float(str(p.get("abs", "0")).replace(" un/mes", "")) if p.get("abs") not in (None, "N/D") else 0.0
            except ValueError:
                return 0.0
        pool = [p for p in productos if p.get("recomendado")] or productos
        if pool:
            max(pool, key=_abs_num)["featured"] = True
    return productos


# ════════════════════ MULTI-PROGRAMA POR BUCKET (regla de negocio Héctor) ════════════════════
# A un MISMO precio, variar el metraje cambia cuántas recámaras caben: menos m² → más recámaras
# chicas; más m² → menos recámaras amplias. Así un bucket de precio atiende a 2-3 segmentos de
# demanda distintos. Límites físicos por NSE (lujo A/B vs económico C-).
def _min_m2_recamara(nse: str) -> float:
    """m² mínimos por recámara: lujo (A/B) 13 m²; económico (C y abajo) 9 m²."""
    return 13.0 if nse in ("A", "B") else 9.0


def _pct_area_recamaras(nse: str) -> float:
    """Fracción del área total que ocupan las recámaras: lujo (A/B) 50%; económico (C-) 60%."""
    return 0.50 if nse in ("A", "B") else 0.60


def _recamaras_que_caben(area_total: float, nse: str) -> int:
    """Máximo nº de recámaras que físicamente caben en un área dada, según la regla:
        n_rec × m²_min_recámara ≤ %_recámaras × área_total
    Topa el resultado entre 1 y 4 (PH). Si el área no alcanza ni para 1, devuelve 1 (estudio)."""
    if not area_total or area_total <= 0:
        return 1
    presupuesto = _pct_area_recamaras(nse) * area_total
    min_m2 = _min_m2_recamara(nse)
    n = int(presupuesto // min_m2)
    return max(1, min(4, n))


def _demanda_recamaras_bucket(lo_m: float, hi_m: float, ft: List[Dict],
                              personas_hogar: Optional[float],
                              hogares_comp: Optional[Dict] = None) -> set:
    """Recámaras con DEMANDA REAL en el rango de precio [lo_m, hi_m] (en M de MXN).
    Combina dos evidencias:
      1) OFERTA VENDIDA: nº de recámaras de las tipologías de ft cuyo ticket cae en el rango y que
         registran ventas (UNIDADES_VENDIDAS>0) — el mercado ya absorbe esas configuraciones.
      2) HOGARES (composición): se infiere la recámara natural según la composición del hogar.
         REGLA DE RECÁMARA COMPARTIDA: salvo hogares unipersonales y de corresidentes, se asume que
         al menos 2 adultos (papá y mamá) comparten la recámara principal. Por tanto, para hogares
         FAMILIARES las recámaras necesarias = personas − 1 (los hijos no comparten entre sí en el
         caso base). Unipersonal → 1 recámara. Corresidentes → 1 recámara por residente (no pareja).
         Se incluyen las recámaras adyacentes ±1 para reflejar configuraciones cercanas que el mismo
         tipo de hogar considera.
    Devuelve un set de enteros de recámaras (1..4). Si no hay evidencia, set vacío."""
    recs: set = set()
    # 1) Oferta vendida en el rango
    for t in ft:
        a = t.get("attributes", {}) if isinstance(t, dict) else {}
        precio = _price(a.get("F____UNIDAD"))
        vend = _num(a.get("UNIDADES_VENDIDAS")) or 0
        rec = _rec_to_int(a.get("CANTIDAD_DE_RECAMARAS"))
        if precio is None or rec is None:
            continue
        pM = precio / 1e6
        if lo_m <= pM <= hi_m and vend > 0:
            recs.add(max(1, min(4, rec)))
    # 2) Composición de hogar (demanda demográfica) con regla de recámara compartida.
    if personas_hogar and personas_hogar > 0:
        # Mezcla de tipos de hogar de la zona (pct). Se infiere la recámara base de cada tipo
        # presente y se agregan, así un bucket refleja la diversidad real de hogares de la zona.
        pesos = _pesos_tipologia_hogar(hogares_comp)
        # Recámara base familiar: 2 adultos comparten la principal → recámaras = personas − 1.
        fam_base = max(1, int(round(personas_hogar)) - 1)
        if pesos.get("familiar", 0) > 0:
            for r in (fam_base - 1, fam_base, fam_base + 1):
                if 1 <= r <= 4:
                    recs.add(r)
        # Unipersonal → 1 recámara (vive solo).
        if pesos.get("unipersonal", 0) > 0:
            recs.add(1)
        # Corresidentes → cada residente su propia recámara (no son pareja): personas recámaras.
        if pesos.get("corresidente", 0) > 0:
            cores = max(1, min(4, int(round(personas_hogar))))
            recs.add(cores)
        # Si no hay desglose de tipología, aplicar la regla familiar (caso dominante) por defecto.
        if not pesos:
            for r in (fam_base - 1, fam_base, fam_base + 1):
                if 1 <= r <= 4:
                    recs.add(r)
    return recs


def _pesos_tipologia_hogar(hogares_comp: Optional[Dict]) -> Dict[str, float]:
    """Extrae los pesos (pct) de los tipos de hogar relevantes para la regla de recámaras:
    familiar (papá+mamá comparten principal), unipersonal (vive solo), corresidente (no pareja).
    Devuelve {} si no hay datos (entonces el llamador aplica la regla familiar por defecto)."""
    if not hogares_comp:
        return {}
    tip = hogares_comp.get("tipologia_hogar") if isinstance(hogares_comp, dict) else None
    if not tip:
        return {}
    out: Dict[str, float] = {}
    for t in tip:
        label = str(t.get("label", "")).lower()
        pct = _num(t.get("pct")) or 0
        if pct <= 0:
            continue
        if "familiar" in label and t.get("parent"):
            out["familiar"] = pct
        elif "unipersonal" in label:
            out["unipersonal"] = pct
        elif "corresidente" in label:
            out["corresidente"] = pct
    return out


def _variantes_multiprograma(s: Dict, ticket_M: float, m2_base: float, pm2_base: float,
                             ft: List[Dict], personas_hogar: Optional[float],
                             hogares_comp: Optional[Dict] = None) -> List[Dict[str, Any]]:
    """Genera las VARIANTES de un bucket de precio: por cada nº de recámaras con demanda real, el
    metraje que resulta de ese mismo ticket a un $/m² coherente, validando que las recámaras quepan
    (regla física). Devuelve lista de dicts {rec, m2, pm2, fits} ordenada por recámaras. La variante
    cuyo m² ≈ m2_base se marca como _base (es la que el cálculo principal ya generó)."""
    nse = s.get("NSE", "C")
    if not ticket_M or ticket_M <= 0 or not m2_base or m2_base <= 0:
        return []
    lo_m = (s.get("val_min") or 0) / 1e6
    hi_m = (s.get("val_max") or 0) / 1e6
    demanda_recs = _demanda_recamaras_bucket(lo_m, hi_m, ft, personas_hogar, hogares_comp)
    if not demanda_recs:
        return []
    min_m2 = _min_m2_recamara(nse)
    pct = _pct_area_recamaras(nse)
    # m² mínimo de vivienda para alojar n recámaras (área de recámaras = pct del total):
    #   n × min_m2 ≤ pct × area  →  area ≥ n × min_m2 / pct
    variantes = []
    for n_rec in sorted(demanda_recs):
        area_min_fisica = (n_rec * min_m2) / pct
        # El metraje de la variante: parte del m2_base y se ajusta para que quepan n_rec recámaras.
        # Menos recámaras que la base → vivienda más amplia por recámara (sube algo el m²);
        # más recámaras → vivienda más compacta por recámara, pero NUNCA por debajo del mínimo físico.
        area_var = max(area_min_fisica, round(m2_base * (0.78 + 0.16 * n_rec), 0))
        area_var = max(M2_MIN, min(M2_MAX, area_var))
        # ¿Caben físicamente n_rec en area_var? (regla de tope)
        caben = _recamaras_que_caben(area_var, nse)
        fits = caben >= n_rec
        if not fits:
            # Topar: el área no alcanza para n_rec; esta variante no es viable a este precio.
            continue
        # $/m² de la variante: mismo ticket repartido en el nuevo metraje (precio constante).
        pm2_var = round(ticket_M * 1e6 / area_var) if area_var else pm2_base
        variantes.append({
            "rec": n_rec,
            "m2": int(round(area_var)),
            "pm2": pm2_var,
            "area_recamaras_max": round(pct * area_var),
            "min_m2_recamara": min_m2,
            "_base": abs(area_var - m2_base) <= max(3.0, m2_base * 0.05),
        })
    return variantes


def derive_productos_venta(ft: List[Dict], segments: List[Dict], unidades_proyecto: Optional[int] = None,
                           personas_hogar: Optional[float] = None,
                           hogares_comp: Optional[Dict] = None) -> List[Dict[str, Any]]:
    """
    Un producto por bucket de demanda. Metodología DPO vigente (ver docs/METODOLOGIA_DIGO.md):
      • RECÁMARAS ancladas al TICKET (programa monotónico por valor).
      • TAMAÑO anclado al PROGRAMA del ticket (banda habitable); el m² observado real de la
        oferta (laboratorio) manda si existe. pm² = ticket / m² (derivado, no al revés).
      • ABSORCIÓN flujo-contra-flujo: (Demanda anual del bucket + Mercado en venta × 5%) / 12
        menos Σ Abs_Demanda de comparables directos con inventario disponible. Pronóstico
        12/18/24 con curva de maduración. La mediana de directos es dato de validación.
      • MULTI-PROGRAMA: variantes de recámaras con demanda real en el bucket.
    """
    color_by_status = {
        "sweet_spot": "green", "desatendido": "blue", "oportunidad": "purple",
        "atendido": "teal", "oceano_rojo": "red", "bajo_crecimiento": "amber",
    }
    # Precio/m² de referencia por NSE (de los tableros estáticos validados · MXN)
    PM2_POR_NSE = {"A": 140000, "B": 95000, "C+": 55000, "C": 40000,
                   "D+": 30000, "D": 22000, "E": 16000}
    # Recámaras según área interior (m²) — usado solo como respaldo cuando no hay ticket.
    def recamaras_por_m2(m2):
        if m2 is None: return "N/D"
        if m2 < 45:  return "Studio"
        if m2 < 65:  return "1 Rec"
        if m2 < 95:  return "2 Rec"
        if m2 < 140: return "3 Rec"
        return "3-4 Rec"

    # PROGRAMA DE RECÁMARAS anclado al TICKET (regla DPO): crece monotónicamente con el
    # valor del segmento, igual que en las tablas de los reportes. El m² afina el tamaño
    # dentro del programa, pero no determina el nº de recámaras (evita no-monotonías como
    # un bucket de mayor ticket con menos recámaras que uno inferior).
    def recamaras_por_ticket(ticket_m, m2=None):
        if ticket_m is None:
            return recamaras_por_m2(m2)
        if ticket_m <= 1.5:   return "1 Rec"      # económico: 1 recámara habitable
        if ticket_m <= 2.5:   return "1 Rec"
        if ticket_m <= 3.5:   return "2 Rec"
        if ticket_m <= 5.0:   return "2 Rec"
        if ticket_m <= 8.0:   return "2 Rec"      # 2 Rec Core / Sweet Spot
        if ticket_m <= 12.0:  return "3 Rec"
        if ticket_m <= 18.0:  return "3 Rec"
        if ticket_m <= 25.0:  return "3-4 Rec"
        return "4 Rec PH"

    # Competencia directa por bucket: nº de tipologías de la oferta (ft) en cada rango de precio
    competencia = {}
    for t in ft:
        a = t.get("attributes", {})
        precio = _price(a.get("F____UNIDAD"))
        if not precio:
            continue
        for s in segments:
            if s.get("val_min") and s.get("val_max") and s["val_min"] <= precio < s["val_max"]:
                competencia[s["bucket"]] = competencia.get(s["bucket"], 0) + 1
                break

    # Familias-mes máximas de toda la zona (para normalizar elasticidad)
    max_nf_month = max((s.get("nuevas_fam", 0) / 12 for s in segments), default=0)

    productos = []
    for s in segments:
        vmin = s.get("val_min")
        vmax = s.get("val_max")
        val_mid = None
        if vmin is not None and vmax is not None and (vmin or vmax):
            # Si val_min es 0 (bucket inferior), usar val_max como referencia superior
            if vmin and vmax:
                val_mid = (float(vmin) + float(vmax)) / 2
            elif vmax:
                # Bucket inferior (val_min=0): anclar a la CAPACIDAD DE PAGO REAL del segmento
                # (ingreso × 12 × múltiplo hipotecario), capado al techo del bucket. Evita
                # productos por encima de lo que la población puede pagar.
                MULTIPLO_HIPOTECARIO = 4.5
                ing_rep = s.get("ing_min")
                cap_compra = (ing_rep * 12 * MULTIPLO_HIPOTECARIO) if ing_rep else None
                if cap_compra:
                    val_mid = min(cap_compra, float(vmax) * 0.75)
                else:
                    val_mid = float(vmax) * 0.5
            elif vmin:
                val_mid = float(vmin) * 1.1

        # DPO anchor: ticket no por debajo del val_min del segmento
        ticket_M = round(val_mid / 1e6, 2) if val_mid else None
        if ticket_M and vmin:
            ticket_M = max(ticket_M, round(float(vmin) / 1e6, 2))

        # ════════ TAMAÑO ANCLADO AL PROGRAMA · REGLA DPO DOCUMENTADA ════════
        # Metodología (reportes Valle Poniente V2 / Calzada): el TAMAÑO NO se deriva de
        # ticket/pm². Se ancla al PROGRAMA DE RECÁMARAS que corresponde al ticket del
        # segmento (un programa habitable real), y el pm² se deriva después como ticket/m².
        # Esto evita metrajes absurdos (p.ej. 28 m² para vivienda económica) cuando el pm²
        # de referencia es alto. El reporte lo enuncia: "a este ticket exigen más de 140 m²"
        # y "al cambiar el tamaño, el mercado ajusta el precio por m²".
        #
        # Bandas de tamaño por ticket (MXN), de las tablas de producto reales de los reportes:
        #   ≤ 1.5M  → 1 Rec · 45-58 m²   (económico: programa mínimo habitable, no studio)
        #   ≤ 2.5M  → 1-2 Rec · 52-68 m²
        #   ≤ 3.5M  → 2 Rec · 62-80 m²
        #   ≤ 5.0M  → 2 Rec · 70-90 m²   (2 Rec Core VP: $7.4M→80 m²; entry: ~70)
        #   ≤ 8.0M  → 2-3 Rec · 80-110 m² (Sweet Spot Calzada: 2 Rec 105 m²)
        #   ≤ 12.0M → 3 Rec · 105-140 m² (Smart Luxury 1Rec+Flex 75; mid 3R ~120)
        #   ≤ 18.0M → 3 Rec · 135-180 m² (Premium Calzada: 3 Rec 160 m²)
        #   ≤ 25.0M → 3-4 Rec · 180-260 m² (Sky-Residence VP: $20M→240 m²)
        #   > 25.0M → 4 Rec PH · 260-360 m² (Grand PH VP: $25M+→320 m²)
        # FUENTE ÚNICA: la tabla vive a nivel módulo (_BANDA_TAMANO_TICKET · DEM-1 la
        # comparte). Mismos valores que la tabla local anterior — sin cambio de conducta.
        def _banda_tamano(ticket_m):
            return _banda_tamano_por_ticket(ticket_m)

        # pm² OBSERVADO real de la oferta en el rango (si existe) — para anclar a la zona.
        PM2_VERTICAL_MIN = 20000
        pm2_obs = []
        for t in ft:
            a = t.get("attributes", {})
            precio = _price(a.get("F____UNIDAD"))
            pm2v = _pm2(a.get("F___M2"))
            if (precio and val_mid and abs(precio - val_mid) / val_mid < 0.20
                    and pm2v and pm2v >= PM2_VERTICAL_MIN):
                pm2_obs.append(pm2v)
        pm2_zona = round(statistics.median(pm2_obs)) if pm2_obs else None

        # m² OBSERVADO real de la oferta en el rango (si existe) — fuente preferida.
        m2_obs = []
        for t in ft:
            a = t.get("attributes", {})
            precio = _price(a.get("F____UNIDAD"))
            area = _num(a.get("ÁREA_PRIVATIVA"))
            if (precio and val_mid and abs(precio - val_mid) / val_mid < 0.20
                    and area and M2_MIN <= area <= M2_MAX):
                m2_obs.append(area)
        m2_zona = round(statistics.median(m2_obs)) if m2_obs else None

        # TAMAÑO: 1º el observado real de la zona (si hay oferta); 2º la banda de programa
        # del ticket, posicionada según el NSE (NSE alto → extremo superior de la banda).
        lo_b, hi_b = _banda_tamano(ticket_M)
        if m2_zona is not None:
            m2 = max(M2_MIN, min(M2_MAX, m2_zona))   # dato real de la base manda
        else:
            # posición dentro de la banda según NSE (A=alto, C/D=bajo)
            pos = {"A": 0.85, "B": 0.65, "C+": 0.5, "C": 0.4, "D+": 0.3, "D": 0.25, "E": 0.2}.get(s["NSE"], 0.5)
            m2 = round(lo_b + (hi_b - lo_b) * pos)
            m2 = max(M2_MIN, min(M2_MAX, m2))

        # pm² DERIVADO del ticket y el tamaño anclado (regla: pm² = ticket / m²).
        # Si hay pm² observado real de la zona se reporta como ancla de mercado, pero el
        # tamaño ya respeta el programa, así que el pm² derivado es coherente.
        if ticket_M and m2:
            pm2 = round(ticket_M * 1e6 / m2)
        elif pm2_zona:
            pm2 = pm2_zona
        else:
            pm2 = PM2_POR_NSE.get(s["NSE"], 50000)

        # ════════ ABSORCIÓN (regla de negocio) ════════
        # Con comparables DIRECTOS (mismo programa/ticket/pm²/tamaño): mediana real × 1.20.
        # Sin directos: captura 100% de la demanda mensual del segmento (sin competencia). Sin techo.
        n_comp = competencia.get(s["bucket"], 0)
        abs_dir = _abs_comparables_directos(ticket_M, pm2, recamaras_por_ticket(ticket_M, m2), m2, ft)
        precio_rec = _precio_recomendado_directos(ticket_M, recamaras_por_ticket(ticket_M, m2), m2, ft)
        abs_info = _absorcion_producto(s.get("nuevas_fam", 0) or 0, s.get("demanda_total", 0) or 0, abs_dir, unidades_proyecto)
        abs_rate = abs_info["abs"]

        mix_pct = None
        tot_nf = sum(x.get("nuevas_fam", 0) for x in segments) or 0
        if tot_nf and s.get("nuevas_fam"):
            mix_pct = round(s["nuevas_fam"] / tot_nf * 100)
        nse_tca = NSE_TCA.get(s["NSE"], 0)   # TCA canónica (fuente única); antes tabla hardcodeada errónea

        # MULTI-PROGRAMA: variantes de recámara que el mismo ticket puede atender ajustando el m²
        # (solo las recámaras con demanda real en el rango; cada una validada físicamente).
        variantes_mp = _variantes_multiprograma(s, ticket_M, m2, pm2, ft, personas_hogar, hogares_comp)

        productos.append({
            "tipo": f"{recamaras_por_ticket(ticket_M, m2)} · {s['bucket']}" if m2 else f"{s['NSE']} · {s['bucket']}",
            "badge": f"{mix_pct}%" if mix_pct is not None else "—",
            "color": color_by_status.get(s["status"], "teal"),
            "rec": recamaras_por_ticket(ticket_M, m2),
            "variantes": variantes_mp,   # [{rec, m2, pm2, area_recamaras_max, min_m2_recamara, _base}]
            "m2": f"{m2} m²" if m2 else "N/D",
            "pm2": f"${pm2:,}" if pm2 else "N/D",
            "ticket": f"${ticket_M:.2f}M" if ticket_M else "N/D",
            "abs": f"{abs_rate:.1f} un/mes" if abs_rate else "N/D",
            "abs_origen": abs_info["origen"],
            "abs_n_directos": abs_info["n_directos"],
            "abs_mediana_directos": abs_info["mediana_directos"],
            "abs_inv_competidor": abs_info.get("inventario_competidor", 0),
            "abs_competidores": abs_info.get("abs_competidores", 0),
            "abs_pronostico": abs_info.get("pronostico", {}),
            "pm2_recomendado": precio_rec.get("pm2_recomendado"),
            "ticket_recomendado_M": precio_rec.get("ticket_recomendado_M"),
            "precio_tol_veredicto": PRECIO_TOL_VEREDICTO,
            "tca": nse_tca,
            "competidores": n_comp,
            "mercado": f"NSE {s['NSE']} · {s['bucket']} · {n_comp} competidores directos",
            "status": s["status"],
            # Recomendado (entra a la mezcla) SOLO si está en el rango de oferta vertical real
            # de la zona (aplicable). Un bucket con demanda demográfica alta pero SIN oferta
            # vertical (p.ej. vivienda económica de la zona ampliada) se muestra pero NO se
            # recomienda para producto vertical: no es captable con este tipo de producto.
            "recomendado": ((s.get("aplicable", True)
                            and (s["status"] in ("sweet_spot", "desatendido", "oportunidad", "atendido")))
                           or s.get("dual_featured", False))
                           and _recomendable_por_absorcion(abs_info["origen"])["recomendable"],
            "no_recomendable_motivo": _recomendable_por_absorcion(abs_info["origen"])["motivo"],
            "aplicable": s.get("aplicable", True),
            "featured": (s["status"] == "sweet_spot" or s.get("dual_featured", False))
                        and _recomendable_por_absorcion(abs_info["origen"])["recomendable"],
            "seg_dim": f"{s['NSE']} · {s['bucket']}",
            "mkt_segmento": s["mkt_total"], "nuevas_fam": s["nuevas_fam"],
            "categoria": _status_label(s["status"]),
            "perfiles": _perfiles_por_segmento(s["NSE"], m2, ticket_M),
            "nota_mercado": s.get("nota_mercado"),
            # — Campos numéricos para el front (slider de captación y sync con Demanda) —
            "nuevas_fam_year": s.get("nuevas_fam", 0),     # RITMO (flow): nuevas familias/año
            "depth": s.get("mkt_total", 0),                # PROFUNDIDAD (depth): hogares del segmento
            "abs_num": round(abs_rate, 2) if abs_rate else 0.0,   # absorción demand-driven (captación 100%)
            "pm2_num": pm2,
            "m2_num": m2,
            "ticket_num": round(ticket_M * 1e6) if ticket_M else None,
            "mix_num": mix_pct if mix_pct is not None else 0,
        })
    # Solo un featured
    feats = [p for p in productos if p.get("featured")]
    if len(feats) > 1:
        best = max(feats, key=lambda p: p.get("nuevas_fam", 0))
        for p in feats:
            if p is not best:
                p["featured"] = False
    elif not feats:
        # Mercado puramente supply-driven (sin sweet spot demand-driven): destacar el producto
        # que el mercado valida con mayor absorción/venta real (el más relevante para invertir).
        def _abs_num(p):
            try:
                return float(str(p.get("abs", "0")).replace(" un/mes", "")) if p.get("abs") not in (None, "N/D") else 0.0
            except ValueError:
                return 0.0
        recomendables = [p for p in productos if p.get("recomendado")]
        pool = recomendables or productos
        if pool:
            best = max(pool, key=_abs_num)
            best["featured"] = True
    return productos


def _status_label(status: str) -> str:
    return {
        "desatendido": "Gap · Desatendido", "sweet_spot": "Sweet Spot Estructural",
        "atendido": "Mercado Core · Atendido", "oportunidad": "Sub-oferta · Oportunidad",
        "oceano_rojo": "Océano Rojo · Sobreoferta", "bajo_crecimiento": "Supply-Driven · Sin Demanda",
    }.get(status, status)


# ──────────────────────── Productos renta (Checkpoint F) ────────────────────────
def derive_productos_renta(ft_renta: List[Dict], segments: List[Dict]) -> List[Dict[str, Any]]:
    """
    Productos de renta derivados de los segments DPO. Para cada bucket con mercado de renta:
      • RENTA mensual = punto medio de rent_min/rent_max (~0.4% del valor de vivienda).
      • TAMAÑO (m²) = mismo criterio que venta (ticket de compra / precio_m²).
      • RENTA/m² = renta_mensual / m².
      • ABSORCIÓN de contratos/mes = nuevas familias que RENTAN/mes del segmento.
    Si hay oferta de renta real (ft_renta), se usa para validar renta/m² de la zona.
    """
    if not segments:
        return []
    # H8 ESTRICTO (regresión reportada 8 jul 2026): sin oferta de RENTA OBSERVADA en la zona
    # no se publican productos de renta modelados (integridad: la escalera base DIGO daba
    # $/m²/mes sin sentido en zonas sin mercado de renta). Dato ausente = N/D.
    _renta_obs = [r for r in (ft_renta or [])
                  if _num((r.get("attributes") or {}).get("F___M2"))]
    if len(_renta_obs) < 3:
        return []
    PM2_POR_NSE = {"A": 140000, "B": 95000, "C+": 55000, "C": 40000,
                   "D+": 30000, "D": 22000, "E": 16000}

    def recamaras_por_m2(m2):
        if m2 is None: return "N/D"
        if m2 < 45:  return "Studio"
        if m2 < 65:  return "1 Rec"
        if m2 < 95:  return "2 Rec"
        if m2 < 140: return "3 Rec"
        return "3-4 Rec"

    # Programa y banda de tamaño anclados al ticket (regla DPO, igual que venta).
    def _banda_tamano_r(ticket_m):
        if ticket_m is None:   return (70, 90)
        if ticket_m <= 1.5:    return (45, 58)
        if ticket_m <= 2.5:    return (52, 68)
        if ticket_m <= 3.5:    return (62, 80)
        if ticket_m <= 5.0:    return (70, 90)
        if ticket_m <= 8.0:    return (80, 110)
        if ticket_m <= 12.0:   return (105, 140)
        if ticket_m <= 18.0:   return (135, 180)
        if ticket_m <= 25.0:   return (180, 260)
        return (260, 360)
    def recamaras_por_ticket_r(ticket_m, m2=None):
        if ticket_m is None:   return recamaras_por_m2(m2)
        if ticket_m <= 2.5:    return "1 Rec"
        if ticket_m <= 8.0:    return "2 Rec"
        if ticket_m <= 18.0:   return "3 Rec"
        if ticket_m <= 25.0:   return "3-4 Rec"
        return "4 Rec PH"

    # renta/m² observada de la oferta real (para validar/anclar)
    pm2_renta_obs = []
    for t in ft_renta:
        a = t.get("attributes", {})
        ap = _num(a.get("ÁREA_PRIVATIVA")); pr = _num(a.get("F____UNIDAD"))
        if ap and 25 <= ap <= 500 and pr and pr > 1000:
            pm2_renta_obs.append(pr / ap)
    pm2_renta_zona = round(statistics.median(pm2_renta_obs)) if pm2_renta_obs else None

    # propensión a rentar de la zona (mkt_renta / mkt_total) para distribuir el sweet spot
    best = max(segments, key=lambda s: s.get("mkt_renta", 0) or 0, default=None)

    productos = []
    for s in segments:
        if s.get("status") == "bajo_crecimiento":
            continue
        rent_min = s.get("rent_min"); rent_max = s.get("rent_max")
        renta_mid = ((rent_min or 0) + (rent_max or 0)) / 2 if (rent_min is not None and rent_max) else None
        # ticket de compra del segmento (para anclar el programa de tamaño)
        vmin = s.get("val_min"); vmax = s.get("val_max")
        val_mid = None
        if vmin and vmax:
            val_mid = (vmin + vmax) / 2
        elif vmax:
            val_mid = vmax * 0.75
        ticket_M_r = (val_mid / 1e6) if val_mid else None
        # TAMAÑO anclado al programa por ticket, posición por NSE (NO val_mid/pm2)
        lo_b, hi_b = _banda_tamano_r(ticket_M_r)
        pos = {"A": 0.85, "B": 0.65, "C+": 0.5, "C": 0.4, "D+": 0.3, "D": 0.25, "E": 0.2}.get(s["NSE"], 0.5)
        m2 = max(M2_MIN, min(M2_MAX, round(lo_b + (hi_b - lo_b) * pos))) if ticket_M_r else None
        pm2_renta = round(renta_mid / m2) if (renta_mid and m2) else None
        # ABSORCIÓN de renta = nuevas familias que rentan/mes (mkt_renta es stock; usamos nuevas_fam × propensión a rentar)
        nf = s.get("nuevas_fam", 0) or 0
        mkt_total = s.get("mkt_total", 0) or 1
        prop_renta = (s.get("mkt_renta", 0) or 0) / mkt_total if mkt_total else 0
        abs_renta = round(nf / 12 * prop_renta, 2) if nf and prop_renta else None

        productos.append({
            "tipo": f"{recamaras_por_ticket_r(ticket_M_r, m2)} Renta · {s['bucket']}" if m2 else f"Renta {s['NSE']}",
            "rec": recamaras_por_ticket_r(ticket_M_r, m2),
            "m2": f"{m2} m²" if m2 else "N/D",
            "pm2_renta": f"${pm2_renta:,}/m²/mes" if pm2_renta else "N/D",
            "renta_ud": f"${round(renta_mid):,}/mes" if renta_mid else "N/D",
            "abs_renta": f"{abs_renta} contratos/mes" if abs_renta else "N/D",
            "ocupacion_target": "92%",
            "status": s.get("status", "atendido"),
            "recomendado": (s.get("aplicable", True)
                            and s.get("status") in ("sweet_spot", "desatendido", "oportunidad", "atendido")),
            "aplicable": s.get("aplicable", True),
            "featured": (best is not None and s is best),
            "seg_renta": f"{s['NSE']} · {s['bucket']}",
            "mkt_segmento": round(s.get("mkt_renta") or 0),
            "nuevas_fam_year": nf,
            "mercado": f"NSE {s['NSE']} · renta {s['bucket']}",
        })
    return productos


# ──────────────────────── Comercio (potencial retail) ────────────────────────
def derive_comercio(agebs: List[Dict], nse_dom_key: Optional[str] = None,
                    ft: Optional[List[Dict]] = None) -> Dict[str, Any]:
    """Agrega gasto retail por categoría y deriva GLA potencial vía ventas-por-m². Renta $/m²,
    inquilinos y giros son DINÁMICOS por zona (no fijos):
      • Renta $/m²/mes  → derivada del gasto real captable por m² de cada giro en la zona
        (a mayor gasto captado por m² de GLA, mayor renta que el local soporta).
      • Inquilinos      → catálogo por NSE dominante de la zona (premium ≠ popular).
      • Giros mostrados → solo los que alcanzan GLA captable ≥ GLA_MIN_VIABLE m².
    """
    if not agebs:
        return {}
    # Gasto anual por categoría (montos MXN · campos C193-201)
    gasto_cats = {
        "Retail": "Gasto Retail", "Supermercado": "Gasto Supermercado",
        "Servicios": "Gasto Servicios", "Educacion": "Gasto Educación",
        "Restaurantes": "Gasto Restaurantes", "Cuidado": "Gasto Cuidado personal",
        "Entret": "Gasto Entretenimiento", "MV": "Gasto Mantenimiento de la vivienda",
        "Mueb": "Gasto Mueblería", "Salud": "Gasto Cuidado de la salud",
    }
    demanda = {}
    for k, col in gasto_cats.items():
        v = _sum(agebs, col)
        demanda[k] = round(v) if v is not None else None

    ing_total = _sum(agebs, "Ingresos totales 2026")

    # Ventas por m²/año típicas del giro (MXN) → GLA = gasto capturado / ventas_m2
    CAPTURA = 0.25
    ventas_m2_anual = {
        "Supermercado": 95000, "Retail": 65000, "Restaurantes": 55000,
        "Educacion": 40000, "Servicios": 60000, "Salud": 70000,
        "Entret": 45000, "Cuidado": 58000, "Mueb": 38000, "MV": 35000,
    }
    # Categoría (giro) del rubro
    categoria_giro = {
        "Supermercado": "Supermercado", "Retail": "Retail / Departamental",
        "Restaurantes": "Restaurantes / F&B", "Educacion": "Educación",
        "Servicios": "Servicios financieros", "Salud": "Salud / Farmacia",
        "Entret": "Entretenimiento / Gym", "Cuidado": "Cuidado personal",
        "Mueb": "Mueblería / Hogar", "MV": "Mejoras del hogar",
    }
    # Inquilino potencial sugerido · catálogo por NSE de la zona (premium ≠ popular).
    # Se elige el catálogo según el NSE dominante; los conceptos cambian con el poder adquisitivo.
    TENANTS_POR_NSE = {
        "premium": {  # A / B
            "Supermercado": "City Market · La Comer · Fresko", "Retail": "Liverpool · Palacio · boutiques",
            "Restaurantes": "Sonora Grill · fine dining · cafés de autor", "Educacion": "Colegios privados · idiomas premium",
            "Servicios": "Banca patrimonial · notaría · seguros", "Salud": "Hospital privado · especialistas · óptica",
            "Entret": "Cinépolis VIP · boutique gym · spa", "Cuidado": "Spa · clínica estética · barber premium",
            "Mueb": "Design house · galería · decoración", "MV": "Boutique ferretera · home improvement",
        },
        "medio": {  # C+ / C
            "Supermercado": "HEB · Soriana · Chedraui Selecto", "Retail": "Suburbia · Sears · Coppel",
            "Restaurantes": "Toks · Italianni's · casual dining", "Educacion": "Kumon · academias · colegios",
            "Servicios": "BBVA · Banorte · servicios", "Salud": "Farmacias · clínica · laboratorio",
            "Entret": "Smart Fit · Cinépolis · gym", "Cuidado": "Estética · barbería · spa",
            "Mueb": "Showroom hogar · decoración", "MV": "Ferretería · mejoras",
        },
        "popular": {  # D+ / D / E
            "Supermercado": "Bodega Aurrerá · Soriana Mercado · Merco", "Retail": "Coppel · Elektra · Waldos",
            "Restaurantes": "Fondas · taquerías · comida corrida", "Educacion": "Academias · escuelas públicas",
            "Servicios": "Banco Azteca · corresponsalías · OXXO", "Salud": "Farmacias del Ahorro · Similares · clínica",
            "Entret": "Gimnasio de barrio · billar · cine popular", "Cuidado": "Estética · barbería de barrio",
            "Mueb": "Mueblería a crédito · bazar", "MV": "Tlapalería · ferretería de barrio",
        },
    }
    def _nivel_nse(k):
        if not k: return "medio"
        k0 = str(k).strip().upper()
        if k0 in ("A", "B"): return "premium"
        if k0 in ("D+", "D", "E"): return "popular"
        return "medio"
    # HÍBRIDO: el comercio puede aspirar al MAYOR de los dos NSE — el demográfico (quién vive en
    # la zona) y el de percepción de valor (posicionamiento premium de la oferta). Si la zona
    # tiene oferta premium, el comercio acompaña ese nivel aunque la demografía sea más baja.
    _RANK_NSE = {"A": 0, "B": 1, "C+": 2, "C": 3, "D+": 4, "D": 5, "E": 6}
    nse_percepcion = _nse_percepcion_valor(ft) if ft else None
    nse_efectivo = nse_dom_key
    if nse_percepcion and (nse_efectivo is None or
            _RANK_NSE.get(str(nse_percepcion).strip().upper(), 99)
            < _RANK_NSE.get(str(nse_efectivo).strip().upper(), 99)):
        nse_efectivo = nse_percepcion   # percepción de valor es mayor → el comercio aspira a ella
    nivel = _nivel_nse(nse_efectivo)
    tenant_ideal = TENANTS_POR_NSE[nivel]
    GLA_MIN_VIABLE = 100  # m² mínimos de GLA captable para que un giro sea viable y se muestre
    oportunidad = {}
    for k, gasto in demanda.items():
        vm2 = ventas_m2_anual.get(k)
        if gasto is not None and gasto > 0 and vm2:
            m2 = round(gasto * CAPTURA / vm2)
            if m2 > 0:
                oportunidad[k] = m2
    gla_target = round(sum(oportunidad.values())) if oportunidad else None

    # Demanda captable @ 15% (lo que un nuevo desarrollo puede capturar realísticamente)
    CAPTABLE_PCT = 0.15
    captable = {k: round(m2 * CAPTABLE_PCT) for k, m2 in oportunidad.items()}

    # RENTA $/m²/mes DINÁMICA: derivada del gasto real captable por m² de GLA de cada giro en la
    # zona. Intuición: el gasto anual captable por el local (gasto×captura×captable) repartido en
    # su GLA da el "poder de venta" por m²; la renta sostenible es una fracción de ese poder (un
    # local paga renta como % de sus ventas). renta_m2_mes = (gasto_captable_anual / m2) × OCC / 12.
    OCC_RENTA = 0.10   # la renta sostenible ≈ 10% de las ventas anuales por m² (ratio retail típico)
    renta_m2_cat = {}
    for k, m2 in oportunidad.items():
        gasto = demanda.get(k)
        if gasto and m2 and m2 > 0:
            ventas_captables_anual_m2 = (gasto * CAPTURA * CAPTABLE_PCT) / m2
            renta_m2_cat[k] = round(ventas_captables_anual_m2 * OCC_RENTA / 12)
    # Filtro de GIROS: solo los que alcanzan la GLA captable mínima viable.
    giros_viables = {k: m2 for k, m2 in oportunidad.items()
                     if captable.get(k, 0) >= GLA_MIN_VIABLE}
    if not giros_viables:   # si ninguno llega al mínimo, mostrar al menos el de mayor GLA
        if oportunidad:
            kmax = max(oportunidad, key=lambda kk: oportunidad[kk])
            giros_viables = {kmax: oportunidad[kmax]}
    gla_target = round(sum(giros_viables.values())) if giros_viables else None

    product_mix = []
    for k, m2 in sorted(giros_viables.items(), key=lambda kv: -kv[1]):
        rm2 = renta_m2_cat.get(k)
        # Rango de renta como texto: ±12% alrededor del valor derivado.
        renta_txt = (f"${int(rm2*0.88)}-{int(rm2*1.12)}" if rm2 else "N/D")
        product_mix.append({
            "giro": categoria_giro.get(k, k),
            "m2": m2,
            "pct": round(m2 / gla_target * 100, 1) if gla_target else 0,
            "tenant": tenant_ideal.get(k, k),
            "renta": renta_txt,
            "renta_m2": rm2,
            "anchor": (k == "Supermercado"),
        })
    renta_low = None
    renta_high = None
    if gla_target and product_mix:
        renta_mensual = sum((p["m2"] * (p["renta_m2"] or 0)) for p in product_mix) * 0.85
        renta_low = round(renta_mensual / 1e6, 2)
        renta_high = round(renta_mensual * 1.35 / 1e6, 2)  # escenario alto (~+35%)

    # Formato abreviado del ingreso anual (texto que el front muestra como $XX.XB)
    def _fmt_abrev(v):
        if v is None:
            return None
        if v >= 1e9:
            return f"{v/1e9:.1f}B"
        if v >= 1e6:
            return f"{v/1e6:.1f}M"
        if v >= 1e3:
            return f"{v/1e3:.1f}k"
        return str(round(v))

    return {
        "ingreso_anual": _fmt_abrev(ing_total),
        "demanda": demanda,
        "oportunidad": oportunidad,
        "captable": captable,
        "product_mix": product_mix,
        "gla_target": gla_target,
        "renta_low": renta_low,
        "renta_high": renta_high,
    }


def _derive_comercio_OLD(agebs):
    cats = {
        "Supermercado": "Demanda Supermercado", "Retail": "Demanda Retail",
        "Restaurantes": "Demanda Restaurantes", "Educacion": "Demanda Educacion",
        "Servicios": "Demanda Servicios", "Salud": "Demanda CuidadodeSalud",
        "Entret": "Demanda Entretenimiento", "Cuidado": "Demanda Cuidado_personal",
        "Mueb": "Demanda Muebleria",
    }
    demanda = {}
    for k, col in cats.items():
        v = _sum(agebs, col)
        demanda[k] = round(v) if v is not None else None
    ing_total = _sum(agebs, "Ingresos totales 2026")

    # m² potenciales de oferta comercial por categoría (campos "... m2")
    cats_m2 = {
        "Supermercado": "Demanda Supermercado m2", "Retail": "Demanda Retail m2",
        "Restaurantes": "Demanda Restaurantes m2", "Educacion": "Demanda Educacion m2",
        "Servicios": "Demanda Servicios m2", "Salud": "Demanda CuidadodeSalud m2",
        "Entret": "Demanda Entretenimiento m2", "Cuidado": "Demanda Cuidado_personal m2",
        "Mueb": "Demanda Muebleria m2",
    }
    oportunidad = {}
    for k, col in cats_m2.items():
        v = _sum(agebs, col)
        if v is not None and v > 0:
            oportunidad[k] = round(v)
    gla_target = round(sum(oportunidad.values())) if oportunidad else None

    # Product mix comercial: distribuir GLA por categoría con tenant ideal y renta/m²
    tenant_ideal = {
        "Supermercado": "Supermercado ancla", "Retail": "Tiendas departamentales",
        "Restaurantes": "Food & beverage", "Educacion": "Academias / colegios",
        "Servicios": "Bancos / servicios", "Salud": "Clínica / farmacia",
        "Entret": "Cine / gimnasio", "Cuidado": "Spa / estética", "Mueb": "Mueblería / hogar",
    }
    renta_m2_cat = {  # renta comercial $/m²/mes típica por giro (referencia de mercado)
        "Supermercado": 180, "Retail": 320, "Restaurantes": 420, "Educacion": 250,
        "Servicios": 380, "Salud": 300, "Entret": 280, "Cuidado": 400, "Mueb": 220,
    }
    product_mix = []
    for k, m2 in sorted(oportunidad.items(), key=lambda kv: -kv[1]):
        is_anchor = (k == "Supermercado")
        product_mix.append({
            "giro": k, "m2": m2,
            "pct": round(m2 / gla_target * 100, 1) if gla_target else None,
            "tenant": tenant_ideal.get(k, k),
            "renta_m2": renta_m2_cat.get(k),
            "anchor": is_anchor,
        })
    # Renta estimada del centro comercial (escenario low): GLA × renta promedio × factor ocupación
    renta_low = None
    if gla_target and product_mix:
        renta_mensual = sum((p["m2"] * (p["renta_m2"] or 0)) for p in product_mix) * 0.85
        renta_low = round(renta_mensual / 1e6, 2)  # en M MXN/mes

    return {
        "ingreso_anual": round(ing_total) if ing_total else None,
        "demanda": demanda,
        "oportunidad": oportunidad,
        "product_mix": product_mix,
        "gla_target": gla_target,
        "renta_low": renta_low,
    }


# ╔══════════════ SECCIÓN 2: ENSAMBLADOR (assembler) ══════════════╗
from typing import Optional, List, Dict, Any


# H11 · CONSOLIDACIÓN: aquí vivían copias duplicadas de _num y _price (idénticas a las del
# inicio del módulo, líneas ~22-42) — eliminadas para que exista UNA sola fuente por helper.
# _pm2v era una tercera copia de _pm2 con otro nombre; se conserva el nombre como alias
# porque varias funciones lo referencian (catálogo: no se renombra lo que otros consumen).
_pm2v = _pm2


# ════════════ PROD-PERFIL · Resta oferta−demanda POR PERFIL (DEM-1 → Producto) ════════════
# Diseño (aprobado como parte del GO "avanza con todo"): para cada perfil de DEM-1 se
# identifica la OFERTA que lo atiende (tipologías cuyo programa de recámaras coincide y
# cuyo ticket cae en su banda de capacidad de pago), y se hace la resta FLUJO CONTRA FLUJO
# (misma regla validada de absorción): demanda del perfil (nuevas_fam + pool)/12 menos
# Σ Abs_Demanda de las tipologías activas que lo atienden → demanda INSATISFECHA mensual.
# El producto SUGERIDO por perfil ancla su $/m² dentro de la percepción de valor (ZA-6):
# ajusta el m² dentro de la banda del programa hasta que ticket/m² caiga en el núcleo
# P25-P75 de la zona (funcional primero, comprable después). Universal para todos los usos.
def derive_producto_perfil(perfiles: List[Dict], ft: List[Dict],
                           percepcion_detalle: Optional[Dict] = None) -> Dict[str, Any]:
    """Enriquece cada perfil DEM-1 con: oferta que lo atiende, insatisfecha y sugerido."""
    rows = _tipologias_planas(ft)
    nucleo = ((percepcion_detalle or {}).get("banda_nucleo") or [None, None])
    p25, p75 = (nucleo + [None, None])[:2]
    resumen = {"perfiles_desatendidos": 0, "perfiles_equilibrados": 0, "perfiles_sobreofertados": 0}
    for p in (perfiles or []):
        cap = p.get("capacidad_pago_banda_M") or [None, None]
        n_rec = (p.get("programa") or {}).get("rec")
        if cap[0] is None or n_rec is None:
            p["oferta_perfil"] = None
            p["insatisfecha_mensual"] = None
            p["status_perfil"] = "sin_capacidad_medible"
            p["producto_sugerido"] = None
            continue
        # Oferta que ATIENDE al perfil: mismo programa (±0 rec; C5 acepta ±1 por diversidad
        # de corresidentes) y ticket dentro de su banda de capacidad.
        tol_rec = 1 if p.get("cohorte") == "C5" else 0
        atiende = [t for t in rows
                   if t.get("precio") and cap[0] <= t["precio"] / 1e6 <= cap[1]
                   and t.get("rec") is not None and abs(t["rec"] - n_rec) <= tol_rec]
        activas = [t for t in atiende if (t.get("disp") or 0) > 0]
        oferta_flujo = round(sum(t.get("abs") or 0 for t in activas), 2)
        inventario = int(sum(t.get("disp") or 0 for t in atiende))
        demanda_mensual = round(((p.get("nuevas_fam_year") or 0) + (p.get("pool_activo") or 0)) / 12.0, 2)
        insat = round(demanda_mensual - oferta_flujo, 2)
        p["oferta_perfil"] = {"n_tipologias": len(atiende), "n_activas": len(activas),
                              "inventario_disp": inventario, "oferta_flujo_mensual": oferta_flujo}
        p["demanda_mensual"] = demanda_mensual
        p["insatisfecha_mensual"] = insat
        if demanda_mensual <= 0 and oferta_flujo <= 0:
            p["status_perfil"] = "sin_movimiento"
        elif insat > max(0.5, demanda_mensual * 0.25):
            p["status_perfil"] = "desatendido"
            resumen["perfiles_desatendidos"] += 1
        elif insat < -max(0.5, demanda_mensual * 0.25):
            p["status_perfil"] = "sobreofertado"
            resumen["perfiles_sobreofertados"] += 1
        else:
            p["status_perfil"] = "equilibrado"
            resumen["perfiles_equilibrados"] += 1
        # PRODUCTO SUGERIDO: ticket central de SU capacidad; m² dentro de la banda del
        # programa ajustado para que el $/m² caiga en el núcleo de percepción (si existe).
        ticket_obj = round((cap[0] + cap[1]) / 2, 2)
        m2b = p.get("m2_banda") or list(_banda_tamano_por_ticket(ticket_obj))
        m2_lo, m2_hi = m2b[0], m2b[1]
        m2_obj = round((m2_lo + m2_hi) / 2)
        pm2_obj = round(ticket_obj * 1e6 / m2_obj)
        ajuste = None
        if p25 and p75:
            if pm2_obj > p75:      # caro para la zona → crecer m² dentro de banda
                m2_obj = min(m2_hi, math.ceil(ticket_obj * 1e6 / p75))
                pm2_obj = round(ticket_obj * 1e6 / m2_obj)
                ajuste = "m2_arriba_para_entrar_al_nucleo"
            elif pm2_obj < p25:    # barato → compactar m² dentro de banda
                m2_obj = max(m2_lo, math.floor(ticket_obj * 1e6 / p25))
                pm2_obj = round(ticket_obj * 1e6 / m2_obj)
                ajuste = "m2_abajo_para_entrar_al_nucleo"
        en_nucleo = bool(p25 and p75 and p25 <= pm2_obj <= p75)
        p["producto_sugerido"] = {
            "rec": n_rec, "cajones": (p.get("programa") or {}).get("cajones"),
            "m2": int(m2_obj), "ticket_M": ticket_obj, "pm2": pm2_obj,
            "en_nucleo_percepcion": en_nucleo, "ajuste": ajuste,
        }
    return {"perfiles": perfiles, "resumen": resumen,
            "nota": ("Resta oferta−demanda por PERFIL (flujo contra flujo, misma regla "
                     "validada). Producto sugerido anclado al núcleo de percepción de la "
                     "zona: funcional primero, comprable después.")}


# ════════════ F-C · RESUMEN COMERCIAL (RES-1 parcial · RES-3 · RES-5) ════════════
EVIDENCIA_MIN_ESTRELLA = 3   # vendidas mínimas para calificar como producto estrella


def _tipologias_planas(ft: List[Dict]) -> List[Dict[str, Any]]:
    """Aplana ft a filas de tipología con nombres canónicos del catálogo."""
    rows = []
    for f in ft:
        a = f.get("attributes", f)
        nombre = a.get("PROYECTO") or a.get("Nombre")
        if not nombre:
            continue
        rows.append({
            "proyecto": nombre,
            "precio": _price(a.get("F____UNIDAD")),
            "pm2": _pm2v(a.get("F___M2")),
            "m2": _num(a.get("ÁREA_TOTAL")) or _num(a.get("ÁREA_PRIVATIVA")),
            "rec": _rec_to_int(a.get("CANTIDAD_DE_RECAMARAS")),
            "vendidas": int(_num(a.get("UNIDADES_VENDIDAS")) or 0),
            "disp": int(_num(a.get("UNIDADES_DISPONIBLES")) or 0),
            "abs": _num(a.get("Abs_Demanda")),
        })
    return rows


def _estrella_de(rows: List[Dict], min_vendidas: int = EVIDENCIA_MIN_ESTRELLA) -> Optional[Dict[str, Any]]:
    """PRODUCTO ESTRELLA (RES-5 · criterio documentado): entre tipologías con evidencia
    real de venta (vendidas ≥ 3): 1º mayor % DESPLAZADO, 2º mayor ABSORCIÓN observada,
    3º más vendidas. Sin candidatos → None (no se inventa estrella)."""
    cand = []
    for t in rows:
        tot = (t.get("vendidas") or 0) + (t.get("disp") or 0)
        if (t.get("vendidas") or 0) >= min_vendidas and tot > 0:
            cand.append({**t, "desplazamiento_pct": round(t["vendidas"] / tot * 100, 1)})
    if not cand:
        return None
    b = max(cand, key=lambda t: (t["desplazamiento_pct"], t.get("abs") or 0, t["vendidas"]))
    return {"proyecto": b["proyecto"], "rec": b.get("rec"),
            "m2": round(b["m2"]) if b.get("m2") else None,
            "precio_M": round(b["precio"] / 1e6, 2) if b.get("precio") else None,
            "pm2": round(b["pm2"]) if b.get("pm2") else None,
            "vendidas": b["vendidas"], "disp": b["disp"],
            "desplazamiento_pct": b["desplazamiento_pct"],
            "abs": round(b["abs"], 2) if b.get("abs") else None}


def derive_resumen_comercial(ft: List[Dict], segments: Optional[List[Dict]] = None) -> Dict[str, Any]:
    """F-C · Clasificación comercial por proyecto + métricas ROBUSTAS duales + estrella.
      • estatus: activo (inventario disponible) · agotado (vendió todo) · sin_dato.
      • Medianas de m²/$m²/precio para TODO el inventario y para SOLO DISPONIBLE
        (decisión RES-3: disponible manda para precio vigente; total para velocidad/mezcla).
      • producto_estrella por proyecto + TOP 3 de la zona + por segmento cuando hay >1
        mercado con estrella (misma lógica para todos los usos inmobiliarios).
    NOTA P3: la ventana 'agotados en los últimos 8 meses' y las absorciones por periodo
    (mes/trimestre/histórica) se activan cuando la base exponga la serie temporal."""
    rows = _tipologias_planas(ft)
    por_proy: Dict[str, List[Dict]] = {}
    for t in rows:
        por_proy.setdefault(t["proyecto"], []).append(t)
    proyectos: Dict[str, Any] = {}
    for nombre, ts in por_proy.items():
        vend = sum(t["vendidas"] for t in ts)
        disp = sum(t["disp"] for t in ts)
        con_dato = (vend + disp) > 0
        estatus = ("activo" if disp > 0 else "agotado") if con_dato else "sin_dato"
        d_ts = [t for t in ts if t["disp"] > 0]
        tot = vend + disp
        proyectos[nombre] = {
            "estatus": estatus,
            "vendidas": vend, "disp": disp,
            "desplazamiento_pct": round(vend / tot * 100, 1) if tot else None,
            "m2_mediana": _stats_robustas([t["m2"] for t in ts])["mediana"],
            "m2_mediana_disp": _stats_robustas([t["m2"] for t in d_ts])["mediana"],
            "pm2_mediana": _stats_robustas([t["pm2"] for t in ts])["mediana"],
            "pm2_mediana_disp": _stats_robustas([t["pm2"] for t in d_ts])["mediana"],
            "precio_mediana_M": (lambda v: round(v / 1e6, 2) if v else None)(
                _stats_robustas([t["precio"] for t in ts])["mediana"]),
            "abs_mediana": _stats_robustas([t["abs"] for t in ts if t.get("abs")])["mediana"],
            "n_tipologias": len(ts),
            "estrella": _estrella_de(ts),
        }
    d_rows = [t for t in rows if t["disp"] > 0]
    oferta_stats = {
        "pm2_total": _stats_robustas([t["pm2"] for t in rows]),
        "pm2_disponible": _stats_robustas([t["pm2"] for t in d_rows]),
        "m2_total": _stats_robustas([t["m2"] for t in rows]),
        "m2_disponible": _stats_robustas([t["m2"] for t in d_rows]),
        "precio_M_total": _stats_robustas([t["precio"] / 1e6 for t in rows if t.get("precio")]),
        "precio_M_disponible": _stats_robustas([t["precio"] / 1e6 for t in d_rows if t.get("precio")]),
        "abs_total": _stats_robustas([t["abs"] for t in rows if t.get("abs")]),
        "abs_disponible": _stats_robustas([t["abs"] for t in d_rows if t.get("abs")]),
    }
    # TOP 3 de la zona (excluyendo la tipología ganadora en cada ronda)
    top_zona: List[Dict] = []
    pool = rows[:]
    for _ in range(3):
        e = _estrella_de(pool)
        if not e:
            break
        top_zona.append(e)
        for i, t in enumerate(pool):
            if (t["proyecto"] == e["proyecto"] and t.get("rec") == e.get("rec")
                    and (round(t["m2"]) if t.get("m2") else None) == e.get("m2")):
                pool.pop(i)
                break
    # Estrella por SEGMENTO de precio (solo cuando hay MÁS de un mercado con estrella)
    por_segmento: Dict[str, Any] = {}
    for s in (segments or []):
        vmin, vmax = s.get("val_min"), s.get("val_max")
        if vmin is None or vmax is None:
            continue
        e = _estrella_de([t for t in rows if t.get("precio") and vmin <= t["precio"] < vmax])
        if e:
            por_segmento[s.get("bucket")] = e
    if len(por_segmento) < 2:
        por_segmento = {}
    return {
        "proyectos": proyectos,
        "oferta_stats": oferta_stats,
        "top_estrella": {"zona": top_zona, "por_segmento": por_segmento},
        "criterio_estrella": ("1º % desplazado · 2º absorción observada · 3º vendidas · "
                              f"evidencia mínima {EVIDENCIA_MIN_ESTRELLA} vendidas"),
    }


def _build_proyectos(resumen: List[Dict]) -> List[Dict[str, Any]]:
    out = []
    for row in resumen:
        a = row.get("attributes", {})
        g = row.get("geometry", {})
        lat = _num(g.get("y")) if g.get("y") is not None else _num(a.get("Y_coor"))
        lng = _num(g.get("x")) if g.get("x") is not None else _num(a.get("X_coor"))
        total = _num(a.get("UNIDADES_TOTALES"))
        disp = _num(a.get("UNIDADES_DISPONIBLES"))
        vend = _num(a.get("UNIDADES_VENDIDAS"))
        ticket = _num(a.get("F__UD_PROM"))
        pm2 = _num(a.get("F__M2_PROM"))
        absn = _num(a.get("ABSORCIÓN_PROYECTO"))
        avance = _num(a.get("AVANCE_COMERCIAL"))
        out.append({
            "name": a.get("PROYECTO") or a.get("Nombre") or "N/D",
            "lat": lat, "lng": lng,
            "abs": round(absn, 2) if absn is not None else None,
            "avance": round(avance * 100, 1) if avance is not None else None,
            "ticket": round(ticket / 1e6, 2) if ticket else None,
            "pm2": round(pm2) if pm2 else None,
            "total": int(total) if total else None,
            "disp": int(disp) if disp is not None else None,
            "vend": int(vend) if vend is not None else None,
        })
    return out


def _build_kpis(proyectos: List[Dict], ft: List[Dict]) -> Dict[str, Any]:
    total = sum(p["total"] for p in proyectos if p.get("total"))
    vend = sum(p["vend"] for p in proyectos if p.get("vend"))
    disp = sum(p["disp"] for p in proyectos if p.get("disp"))
    abs_vals = [p["abs"] for p in proyectos if p.get("abs")]
    avg_abs = round(sum(abs_vals) / len(abs_vals), 2) if abs_vals else None

    # ticket / pm2 ponderado por unidades
    tnum = tden = pnum = pden = 0.0
    for t in ft:
        a = t.get("attributes", {})
        precio = _price(a.get("F____UNIDAD")); pm2 = _pm2v(a.get("F___M2"))
        u = _num(a.get("UNIDADES_TOTALES")) or 0
        if precio and u:
            tnum += precio * u; tden += u
        if pm2 and u:
            pnum += pm2 * u; pden += u
    avg_ticket = round(tnum / tden / 1e6, 2) if tden else None
    avg_pm2 = round(pnum / pden) if pden else None

    return {
        "proyectos": len(proyectos),
        "unidades": total or None,
        "vendidas": vend or None,
        "disponibles": disp or None,
        "avg_abs": avg_abs, "avg_pm2": avg_pm2, "avg_ticket": avg_ticket,
    }


def _build_inventario(ft: List[Dict]) -> Dict[str, List]:
    PRICE = [(0, 2.5, "< $2.5M"), (2.5, 3.5, "$2.5M - $3.5M"), (3.5, 5.0, "$3.5M - $5.0M"),
             (5.0, 7.0, "$5.0M - $7.0M"), (7.0, 10.0, "$7.0M - $10.0M"),
             (10.0, 15.0, "$10.0M - $15.0M"), (15.0, 99.0, "> $15.0M")]
    M2 = [(0, 60, "< 60 m²"), (60, 80, "60-80 m²"), (80, 110, "80-110 m²"),
          (110, 150, "110-150 m²"), (150, 9999, "> 150 m²")]

    rows = []
    for t in ft:
        a = t.get("attributes", {})
        rows.append({
            "precio": _price(a.get("F____UNIDAD")), "m2": _num(a.get("ÁREA_PRIVATIVA")),
            "total": _num(a.get("UNIDADES_TOTALES")) or 0,
            "vend": _num(a.get("UNIDADES_VENDIDAS")) or 0,
        })
    total_units = sum(r["total"] for r in rows) or 0

    def bucketize(buckets, key, div):
        res = []
        for lo, hi, label in buckets:
            inb = [r for r in rows if r[key] is not None and lo <= r[key] / div < hi]
            uds = sum(r["total"] for r in inb)
            vend = sum(r["vend"] for r in inb)
            res.append({"rango": label, "uds": int(uds),
                        "pct": round(uds / total_units * 100, 1) if total_units else None,
                        "abs": None})
        return res

    return {
        "inventario_precio": bucketize(PRICE, "precio", 1e6),
        "inventario_m2": bucketize(M2, "m2", 1.0),
    }


# ════════════════════════════════════════════════════════════════════════════
# INVENTARIO POR PROYECTO Y TIPOLOGÍA (formato TYPOLOGIES del front)
# Agrupa las features VVV (ft) por proyecto, en el formato que tplInventario consume.
# Filtra SOLO a los proyectos del universo de los 3 sets (directos/primarios/secundarios)
# para que el inventario muestre exclusivamente los comparables del análisis, no toda
# la ciudad. Integridad: dato ausente = "No disponible"/null, nunca inventado.
# ════════════════════════════════════════════════════════════════════════════
def _build_typologies(ft: List[Dict], nombres_universo: set) -> Dict[str, List[Dict]]:
    def _val(v):
        # -1 es marcador de "no disponible" en los datasets de origen
        n = _num(v)
        return n if (n is not None and n != -1) else None
    typ: Dict[str, List[Dict]] = {}
    for f in ft:
        a = f.get("attributes", f)
        nombre = a.get("PROYECTO") or a.get("Nombre")
        if not nombre:
            continue
        # Filtro: solo proyectos del universo de los 3 sets (si se proporcionó)
        if nombres_universo and nombre not in nombres_universo:
            continue
        area_priv = _val(a.get("ÁREA_PRIVATIVA"))
        area_terr = _val(a.get("ÁREA_DE_TERRAZA"))
        area_total = _val(a.get("ÁREA_TOTAL"))
        precio_ud = _price(a.get("F____UNIDAD"))     # preventa sin tracking → None (nunca 0)
        precio_m2 = _pm2v(a.get("F___M2"))
        total = _val(a.get("UNIDADES_TOTALES"))
        disp = _val(a.get("UNIDADES_DISPONIBLES"))
        vend = _val(a.get("UNIDADES_VENDIDAS"))
        absn = _num(a.get("Abs_Demanda"))
        cajones = _val(a.get("CAJONES_ASIGNADOS"))
        avance = (vend / total) if (total and vend is not None) else None
        typ.setdefault(nombre, []).append({
            "tipo": int(_val(a.get("TIPO"))) if _val(a.get("TIPO")) is not None else len(typ.get(nombre, [])) + 1,
            "area_priv": area_priv,
            "area_terr": area_terr if area_terr is not None else "No disponible",
            "area_total": area_total,
            "rec": int(_val(a.get("CANTIDAD_DE_RECAMARAS"))) if _val(a.get("CANTIDAD_DE_RECAMARAS")) is not None else None,
            "precio_ud": precio_ud,
            "precio_m2": precio_m2,
            "unid_total": int(total) if total is not None else None,
            "unid_disp": int(disp) if disp is not None else None,
            "unid_vend": int(vend) if vend is not None else None,
            "abs": round(absn, 4) if absn is not None else 0.0,
            "avance": round(avance, 4) if avance is not None else 0.0,
            "cajones": int(cajones) if cajones is not None else "No disponible",
        })
    # Ordenar tipologías de cada proyecto por precio_ud ascendente (None al final)
    for nombre in typ:
        typ[nombre].sort(key=lambda t: (t["precio_ud"] is None, t["precio_ud"] or 0))
    return typ


# ════════════════════════════════════════════════════════════════════════════
# VARIABLES NOMBRADAS DEL ANÁLISIS (ANALYSIS_VARS)
# Objeto único generado por cada análisis a partir del PIN. Es la FUENTE ÚNICA DE
# VERDAD que TODAS las secciones del front consumen (z._vars). Evita fuentes
# desincronizadas: Mapa, Inventario, Demanda, Producto, Mezclas y Monitor leen de aquí.
# ════════════════════════════════════════════════════════════════════════════
def build_analysis_vars(req, profile, isocronas, perception, demografia,
                        segments, productos, productos_renta, proyectos, kpis) -> Dict[str, Any]:
    """Consolida todas las variables clave del análisis en un solo objeto nombrado.
    Todo proviene de datos reales del análisis (sin promedios inventados). El front
    llama z._vars.<nombre> en cualquier sección para mostrar valores coherentes."""
    comp = (perception or {}).get("competidores") or {}
    merc = (perception or {}).get("mercados") or {}
    iso_keys = sorted(int(m) for m in (isocronas or {}).keys())
    # Universo de los 3 sets (azul+verde+morado) = todos los proyectos clasificados
    universo = (comp.get("directos") or []) + (comp.get("primarios") or []) + (comp.get("secundarios") or [])
    return {
        # — Pin y predio —
        "pin": {"lat": req.lat, "lng": req.lng},
        "predio_m2": req.predio_m2,
        "uso_comercial": req.uso_comercial,
        # — Isócronas (zonas) —
        "perfil_iso": profile["key"],
        "perfil_label": profile["label"],
        "isocrona_primaria_min": iso_keys[0] if iso_keys else None,      # azul
        "isocrona_secundaria_min": iso_keys[-1] if len(iso_keys) > 1 else None,  # verde
        "isocronas_min": iso_keys,
        # — Zona de influencia real (morado) —
        "zona_poligono": (perception or {}).get("zona_poligono"),
        "barrera_mercado": (perception or {}).get("barrera", False),
        "mercados_detectados": merc.get("k"),
        "gap_separabilidad": merc.get("gap"),
        "varianza_explicada": merc.get("var_explained"),
        # — 3 sets de competidores —
        "competidores_directos": comp.get("directos") or [],
        "competidores_primarios": comp.get("primarios") or [],
        "competidores_secundarios": comp.get("secundarios") or [],
        "n_directos": comp.get("n_directos", 0),
        "n_primarios": comp.get("n_primarios", 0),
        "n_secundarios": comp.get("n_secundarios", 0),
        # — Universo de oferta (azul+verde+morado) para Inventario/KPIs/Mezclas —
        "universo_proyectos": universo,
        "n_universo": len(universo),
        "proyectos": proyectos,         # detalle completo de oferta vertical
        "kpis": kpis,
        # — Demografía / demanda (datos granulares reales de la base) —
        "poblacion": demografia.get("population"),
        "hogares": demografia.get("households"),
        "ingreso_hogar": demografia.get("ingreso_hogar"),     # IXH real del NSE dominante
        "nse_dominante": demografia.get("nse_dominante"),
        "tca": demografia.get("tca"),
        # — Ubicación geográfica real (de la base ArcGIS) —
        "municipio": demografia.get("municipio"),
        "estado": demografia.get("estado"),
        "pais": demografia.get("pais"),
        "municipios_todos": demografia.get("municipios_todos"),
        "segmentos_demanda": segments,                        # estratos granulares + aplicable
        # — Producto / sweet spot —
        "productos": productos,
        "productos_renta": productos_renta,
        "sweet_spot": next((p for p in (productos or []) if p.get("featured")), None),
    }


def derive_percepcion_detalle(perception, segments, demografia, productos) -> Dict[str, Any]:
    """ZA-6 · Delimita la PERCEPCIÓN DE VALOR de la zona y describe el MERCADO META que
    genera demanda para una zona así. Todo con datos reales y métricas robustas (RES-2):
      • límites inferior/superior = P10/P90 del $/m² del mercado del pin,
      • núcleo = P25-P75 (donde vive el 50% central de la oferta comparable),
      • los outliers (Tukey) se cuentan pero NO definen la zona,
      • mercado meta = NSE/ingreso/etapas de vida del ancla de demanda derivada.
    Sin oferta comparable → banda None y nota explícita (valor por percepción/NSE)."""
    p = perception or {}
    sr = p.get("stats_robustas") or {}
    det = {
        "pm2_mediana": sr.get("mediana"),
        "pm2_mad": sr.get("mad"),
        "banda_nucleo": [sr.get("p25"), sr.get("p75")],
        "limite_inferior": sr.get("p10"),
        "limite_superior": sr.get("p90"),
        "extremos": [sr.get("min"), sr.get("max")],
        "n_comparables": sr.get("n") or 0,
        "outliers_n": sr.get("outliers_n") or 0,
        "nse_percepcion": None,
        "mercado_meta": None,
        "nota": None,
    }
    nd = (demografia or {}).get("nse_dominante") or ""
    det["nse_percepcion"] = nd.split(" ")[0] if nd else None
    # Mercado meta desde el producto ANCLA (featured) y su segmento de demanda
    feat = next((x for x in (productos or []) if x.get("featured")), None)
    seg_feat = None
    if feat:
        seg_feat = next((s for s in (segments or [])
                         if f"{s.get('NSE')} · {s.get('bucket')}" == feat.get("seg_dim")), None)
    etiquetas = ["niños", "adolescentes", "jóvenes", "jóvenes adultos", "consolidados", "maduros"]
    edad = (demografia or {}).get("edad_grupos") or []
    etapas_top = []
    if edad and len(edad) == 6:
        orden = sorted(range(6), key=lambda i: -(edad[i] or 0))
        etapas_top = [{"etapa": etiquetas[i], "poblacion": edad[i]} for i in orden[:3] if edad[i]]
    det["mercado_meta"] = {
        "nse": (seg_feat or {}).get("NSE") or det["nse_percepcion"],
        "ingreso_min": (seg_feat or {}).get("ing_min"),
        "ingreso_max": (seg_feat or {}).get("ing_max"),
        "ticket_ancla": (feat or {}).get("ticket"),
        "perfiles": (feat or {}).get("perfiles") or [],
        "etapas_top": etapas_top,
        "share_renta": (seg_feat or {}).get("share_renta"),
    }
    if det["pm2_mediana"]:
        try:
            det["nota"] = (f"El valor percibido de la zona vive entre "
                           f"${det['limite_inferior']:,.0f} y ${det['limite_superior']:,.0f}/m² "
                           f"(núcleo ${det['banda_nucleo'][0]:,.0f}–${det['banda_nucleo'][1]:,.0f}). "
                           f"{det['outliers_n']} proyecto(s) fuera de norma no definen la zona.")
        except Exception:
            det["nota"] = None
    else:
        det["nota"] = ("Sin oferta comparable: la percepción de valor se establece por NSE y "
                       "capacidad de pago de la demanda (valor indicativo).")
    return det


# ════════════ CAPTABLE · Mercado captable por MODELO GRAVITACIONAL (Huff v1) ════════════
# Regla de negocio (metodología §3/§6): el captable viene de FUERA de la zona natural.
# HONESTIDAD DEL DATO: los motores de ruteo NO entregan flujos origen-destino observados
# (eso es telemetría · ticket P2). v1 usa el ESTÁNDAR Huff que la propia metodología pide:
# atracción de cada AGEB de la CORONA (anillo captable − zona natural) ∝ masa/(1+dist)^exp,
# sobre demografía REAL del KMZ. El share queda ETIQUETADO como modelo y se actualiza a
# flujos observados cuando el OD de Predik esté disponible. Extranjero/turístico: la regla
# se ENCIENDE SOLA cuando la base publique columnas de población extranjera.
HUFF_EXP = float(os.environ.get("DATARIA_HUFF_EXP", "2.0"))
CAPTABLE_MIN_TOPE = int(os.environ.get("DATARIA_CAPTABLE_MIN_TOPE", "30"))
UMBRAL_EXTRANJERO_TURISTICO = float(os.environ.get("DATARIA_UMBRAL_EXTRANJERO", "0.03"))
_RE_EXTRANJERO = re.compile(r"extranj|otro\s*pa[ií]s", re.I)
_RE_MASA_HOG = re.compile(r"hogares", re.I)
_RE_MASA_POB = re.compile(r"poblaci", re.I)
_RE_MUNICIPIO = re.compile(r"municip", re.I)


def _attr_num(attrs: Dict, regex, excl=None) -> Optional[float]:
    """Primer valor numérico de attrs cuya CLAVE matchea regex (parsea comas).
    `excl`: regex de EXCLUSIÓN de claves (p. ej. población TOTAL sin contar la columna
    'Población nacida en otro país', que también contiene 'población')."""
    for k, v in (attrs or {}).items():
        ks = str(k)
        if not regex.search(ks):
            continue
        if excl is not None and excl.search(ks):
            continue
        try:
            return float(str(v).replace(",", "").replace("$", "").strip())
        except (ValueError, TypeError):
            continue
    return None


def derive_mercado_captable(agebs_geo_capt: List[Dict], natural_ring: List[List[float]],
                            pin_lng: float, pin_lat: float, anillo_min: int,
                            fuente_iso: Optional[str],
                            poligono_captable: Optional[List[List[float]]] = None) -> Dict[str, Any]:
    """Orígenes del mercado captable en la CORONA (fuera de la zona natural), con share
    estimado por gravedad (Huff) sobre masas reales del KMZ. Integridad: sin masa
    publicada → conteo de AGEBs con confianza declarada; sin corona → captable vacío."""
    corona = []
    for g in (agebs_geo_capt or []):
        la, ln = g.get("lat"), g.get("lng")
        if la is None or ln is None:
            continue
        if natural_ring and _point_in_ring(ln, la, natural_ring):
            continue   # dentro de la zona natural → NO es captable
        corona.append(g)
    out: Dict[str, Any] = {
        "activo": True, "metodo": "huff_gravity_v1", "exponente": HUFF_EXP,
        "anillo_min": anillo_min, "fuente_isocrona": fuente_iso,
        "n_agebs_corona": len(corona), "masa_fuente": None,
        "origenes": [], "extranjero": None,
        "poligono_captable": poligono_captable,
        "nota": ("Share estimado por MODELO GRAVITACIONAL (Huff) sobre demografía real de "
                 "la corona captable. Se actualizará a flujos origen-destino OBSERVADOS "
                 "cuando el API de movilidad (Predik OD) esté disponible."),
    }
    if not corona:
        out["activo"] = False
        out["nota"] = "Sin AGEBs en la corona captable (anillo captable ≈ zona natural)."
        return out
    # Masa por AGEB: hogares publicados en el KMZ; si no, población; si no, conteo (declarado)
    con_hog = sum(1 for g in corona if _attr_num(g.get("attrs"), _RE_MASA_HOG))
    con_pob = sum(1 for g in corona if _attr_num(g.get("attrs"), _RE_MASA_POB))
    if con_hog >= len(corona) * 0.5:
        out["masa_fuente"] = "hogares_kmz"
        masa_de = lambda g: _attr_num(g.get("attrs"), _RE_MASA_HOG) or 0.0
    elif con_pob >= len(corona) * 0.5:
        out["masa_fuente"] = "poblacion_kmz"
        masa_de = lambda g: _attr_num(g.get("attrs"), _RE_MASA_POB, excl=_RE_EXTRANJERO) or 0.0
    else:
        out["masa_fuente"] = "conteo_agebs (la base no publica masa en el KMZ · menor confianza)"
        masa_de = lambda g: 1.0
    lat0 = math.radians(pin_lat)
    acumul: Dict[tuple, Dict[str, Any]] = {}
    total_w = 0.0
    for g in corona:
        dlat = (g["lat"] - pin_lat) * 111.32
        dlng = (g["lng"] - pin_lng) * 111.32 * math.cos(lat0)
        dist = math.sqrt(dlat * dlat + dlng * dlng)
        masa = masa_de(g)
        w = masa / ((1.0 + dist) ** HUFF_EXP) if masa else 0.0
        total_w += w
        muni = None
        for k, v in (g.get("attrs") or {}).items():
            if _RE_MUNICIPIO.search(str(k)) and v:
                muni = str(v)[:40]
                break
        key = (g.get("nse_txt") or "N/D", muni)
        a = acumul.setdefault(key, {"nse": key[0], "municipio": key[1], "n_agebs": 0,
                                    "hogares_est": 0.0, "w": 0.0, "dists": []})
        a["n_agebs"] += 1
        a["hogares_est"] += masa if out["masa_fuente"] != "conteo_agebs (la base no publica masa en el KMZ · menor confianza)" else 0
        a["w"] += w
        a["dists"].append(dist)
    origenes = []
    for a in acumul.values():
        origenes.append({
            "nse": a["nse"], "municipio": a["municipio"],
            "n_agebs": a["n_agebs"],
            "hogares_est": round(a["hogares_est"]) if a["hogares_est"] else None,
            "dist_km_mediana": round(statistics.median(a["dists"]), 1),
            "share_pct": round(a["w"] / total_w * 100, 1) if total_w else None,
        })
    origenes.sort(key=lambda x: -(x["share_pct"] or 0))
    out["origenes"] = origenes[:12]
    # ── Regla EXTRANJERO/TURÍSTICO: se enciende sola cuando la base publique columnas ──
    pob_ext = 0.0
    pob_tot = 0.0
    hay_columna = False
    for g in corona:
        ext = _attr_num(g.get("attrs"), _RE_EXTRANJERO)
        pob = _attr_num(g.get("attrs"), _RE_MASA_POB, excl=_RE_EXTRANJERO)
        if ext is not None:
            hay_columna = True
            pob_ext += ext
            pob_tot += pob or 0.0
    if hay_columna and pob_tot <= 0:
        out["extranjero"] = {
            "presente": True, "poblacion_extranjera": round(pob_ext), "share_pct": None,
            "zona_turistica": None,
            "nota": ("La base publica población extranjera pero no población total en el "
                     "KMZ; el share se calculará cuando ambas columnas estén presentes."),
        }
    elif hay_columna and pob_tot > 0:
        share = pob_ext / pob_tot
        out["extranjero"] = {
            "presente": True, "poblacion_extranjera": round(pob_ext),
            "share_pct": round(share * 100, 2),
            "zona_turistica": share >= UMBRAL_EXTRANJERO_TURISTICO,
            "umbral_pct": UMBRAL_EXTRANJERO_TURISTICO * 100,
            "nota": ("ZONA CON COMPONENTE EXTRANJERO RELEVANTE: el mercado captable debe "
                     "incluir al visitante/residente extranjero (regla turística de la "
                     "metodología)." if share >= UMBRAL_EXTRANJERO_TURISTICO else
                     "Componente extranjero presente pero por debajo del umbral turístico."),
        }
    else:
        out["extranjero"] = {
            "presente": False,
            "nota": ("Próximamente · la base aún no publica columnas de población "
                     "extranjera en esta capa; la regla se activa sola al "
                     "aparecer."),
        }
    return out


# ════════════ OD-SINTÉTICO v2 · Ancla censal INEGI (diseño DISENO_OD_SINTETICO.md) ════════════
# Independiente de Predik y de cualquier servicio vivo: el ancla (flujos municipio→municipio
# de TRABAJO/ESTUDIO del Censo 2020) se carga por CADENA DE FUENTES con degradación declarada:
#   1) datos/od_censo/<estado>.csv versionado en el repo (formato canónico) — inmune a caídas.
#   2) Autofetch remoto (env DATARIA_OD_URL_TPL) al primer uso — Render tiene egreso abierto;
#      el sandbox de desarrollo no, por eso existe /api/od/status para verificar desde Safari.
#   3) Sin ancla → v1 Huff (modelo declarado). NUNCA se inventa flujo.
# Emparejamiento por NOMBRE normalizado de municipio (sin catálogos externos que inventar).
# URL REAL verificada en INEGI vía navegador (8 jul 2026): tabulado "Movilidad cotidiana"
# del Cuestionario Ampliado, nivel Estatal/Municipal. Patrón confirmado con eum, nl y jal.
OD_URL_TPL = os.environ.get(
    "DATARIA_OD_URL_TPL",
    "https://www.inegi.org.mx/contenidos/programas/ccpv/2020/tabulados/ampliado/"
    "cpv2020_a_{abrev}_10_movilidad_cotidiana.xlsx")
# Abreviaturas INEGI de entidad (verificadas: nl, jal, eum; resto = convención estándar
# INEGI de estos mismos tabulados — cualquier desviación se diagnostica con /api/od/status)
ESTADO_ABREV_INEGI = {
    "AGUASCALIENTES": "ags", "BAJA CALIFORNIA": "bc", "BAJA CALIFORNIA SUR": "bcs",
    "CAMPECHE": "camp", "COAHUILA DE ZARAGOZA": "coah", "COAHUILA": "coah",
    "COLIMA": "col", "CHIAPAS": "chis", "CHIHUAHUA": "chih",
    "CIUDAD DE MEXICO": "cdmx", "DURANGO": "dgo", "GUANAJUATO": "gto",
    "GUERRERO": "gro", "HIDALGO": "hgo", "JALISCO": "jal", "MEXICO": "mex",
    "MICHOACAN DE OCAMPO": "mich", "MICHOACAN": "mich", "MORELOS": "mor",
    "NAYARIT": "nay", "NUEVO LEON": "nl", "OAXACA": "oax", "PUEBLA": "pue",
    "QUERETARO": "qro", "QUINTANA ROO": "qroo", "SAN LUIS POTOSI": "slp",
    "SINALOA": "sin", "SONORA": "son", "TABASCO": "tab", "TAMAULIPAS": "tamps",
    "TLAXCALA": "tlax", "VERACRUZ DE IGNACIO DE LA LLAVE": "ver", "VERACRUZ": "ver",
    "YUCATAN": "yuc", "ZACATECAS": "zac",
}
OD_AUTOFETCH = os.environ.get("DATARIA_OD_AUTOFETCH", "1") == "1"
OD_DIR_LOCAL = Path(os.environ.get("DATARIA_OD_DIR", str(Path(__file__).resolve().parent / "datos" / "od_censo")))
OD_CACHE: Dict[str, Dict[str, Any]] = {}
OD_COBERTURA_MIN = float(os.environ.get("DATARIA_OD_COBERTURA_MIN", "0.6"))


def _norm_nombre(s) -> str:
    """Normaliza nombres (municipio/estado) para emparejar: mayúsculas, sin acentos."""
    if not s:
        return ""
    t = str(s).strip().upper()
    for a, b in (("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"), ("Ü", "U"), ("Ñ", "N")):
        t = t.replace(a, b)
    return re.sub(r"\s+", " ", t)


def _od_parse_canonico(texto_csv: str) -> Dict[str, Dict[tuple, float]]:
    """Parsea el formato CANÓNICO del ancla: columnas con origen, destino, motivo, viajes
    (detección de encabezados por regex; tolerante a orden). → {motivo: {(orig,dest): v}}"""
    lineas = [l for l in texto_csv.splitlines() if l.strip()]
    if len(lineas) < 2:
        raise ValueError("ancla OD vacía")
    sep = "," if lineas[0].count(",") >= lineas[0].count(";") else ";"
    head = [_norm_nombre(h) for h in lineas[0].split(sep)]
    def _col(rx):
        for i, h in enumerate(head):
            if re.search(rx, h):
                return i
        return None
    io_, id_ = _col(r"ORIGEN|RESIDEN"), _col(r"DESTINO|TRABAJ|ESTUDI|LUGAR")
    im, iv = _col(r"MOTIVO"), _col(r"VIAJES|POBLACION|PERSONAS|VALOR|TOTAL")
    if io_ is None or id_ is None or iv is None:
        raise ValueError(f"ancla OD sin columnas reconocibles: {head[:6]}")
    out: Dict[str, Dict[tuple, float]] = {"trabajo": {}, "estudio": {}}
    for l in lineas[1:]:
        p = [c.strip().strip('"').strip("'") for c in l.split(sep)]
        if len(p) <= max(io_, id_, iv):
            continue
        try:
            v = float(str(p[iv]).replace(",", "").strip() or 0)
        except ValueError:
            continue
        if v <= 0:
            continue
        mot = _norm_nombre(p[im]) if (im is not None and len(p) > im) else "TRABAJO"
        key = "estudio" if "ESTUDI" in mot else "trabajo"
        o, d = _norm_nombre(p[io_]), _norm_nombre(p[id_])
        if o and d:
            out[key][(o, d)] = out[key].get((o, d), 0.0) + v
    if not out["trabajo"] and not out["estudio"]:
        raise ValueError("ancla OD sin flujos > 0")
    return out


def _od_parse_tabulado_inegi(xlsx_bytes: bytes) -> Dict[str, Dict[tuple, float]]:
    """Parsea el TABULADO REAL de Movilidad cotidiana (Censo 2020 · Cuestionario Ampliado,
    estatal/municipal) — estructura verificada contra cpv2020_a_nl_10 (8 jul 2026):
      hoja '14' = movilidad LABORAL por municipio · hoja '02' = ESCOLAR por municipio;
      columnas: [1] Municipio ('001 Abasolo*') · [2] Sexo · [3] Estimador · [4] población ·
      [6] % en el mismo municipio · [7] % en otro municipio · [9] % en otra entidad o país.
    IMPORTANTE (distinción de Héctor): esto es movilidad COTIDIANA (viajes diarios al
    trabajo/escuela), NO migración. El tabulado da VOLÚMENES DE SALIDA por categoría
    (mismo municipio / otro municipio / otra entidad), no el municipio destino específico
    (ese vive en microdatos · v2.2). Se emiten flujos canónicos con destinos-categoría:
    '(MISMO MUNICIPIO)', '(OTRO MUNICIPIO)', '(OTRA ENTIDAD)'."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    hojas = {"trabajo": "14", "estudio": "02"}
    out: Dict[str, Dict[tuple, float]] = {"trabajo": {}, "estudio": {}}
    for motivo, hoja in hojas.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        for row in ws.iter_rows(values_only=True, min_row=10):
            if len(row) < 10:
                continue
            muni_raw, sexo, estim = row[1], row[2], row[3]
            if not muni_raw or str(sexo).strip() != "Total" or str(estim).strip() != "Valor":
                continue
            m = str(muni_raw).strip()
            if m.lower() == "total":
                continue
            nombre = _norm_nombre(re.sub(r"^\d+\s*", "", m).rstrip("*").strip())
            if not nombre:
                continue
            try:
                pob = float(row[4] or 0)
                pct_mismo = float(row[6] or 0)
                pct_otro = float(row[7] or 0)
                pct_otra_ent = float(row[9] or 0)
            except (TypeError, ValueError):
                continue
            if pob <= 0:
                continue
            out[motivo][(nombre, "(MISMO MUNICIPIO)")] = round(pob * pct_mismo / 100.0, 1)
            out[motivo][(nombre, "(OTRO MUNICIPIO)")] = round(pob * pct_otro / 100.0, 1)
            out[motivo][(nombre, "(OTRA ENTIDAD)")] = round(pob * pct_otra_ent / 100.0, 1)
    if not out["trabajo"] and not out["estudio"]:
        raise ValueError("tabulado sin hojas 14/02 reconocibles")
    return out


def _od_salientes(flujos: Dict[str, Dict[tuple, float]], muni_pin: str) -> Dict[str, float]:
    """Masa de commuters de cada municipio que puede llegar al pin (v2.1):
    munis distintos al del pin → salientes '(OTRO MUNICIPIO)' (+ '(OTRA ENTIDAD)' si el
    pin estuviera en otra entidad; se incluye para coronas interestatales);
    el municipio DEL pin → su flujo '(MISMO MUNICIPIO)' (trabajan dentro del municipio)."""
    d_pin = _norm_nombre(muni_pin)
    masa: Dict[str, float] = {}
    for motivo, m in (flujos or {}).items():
        for (o, cat), v in m.items():
            if o == d_pin and "MISMO MUNICIPIO" in cat:
                masa[o] = masa.get(o, 0.0) + v
            elif o != d_pin and ("OTRO MUNICIPIO" in cat or "OTRA ENTIDAD" in cat):
                masa[o] = masa.get(o, 0.0) + v
    return masa


async def _od_cargar_estado(estado_nombre: str) -> Dict[str, Any]:
    """Carga el ancla OD del estado por la cadena de fuentes. Cachea en memoria."""
    ent = _norm_nombre(estado_nombre)
    if not ent:
        return {"ok": False, "error": "estado no determinado"}
    if ent in OD_CACHE:
        return OD_CACHE[ent]
    res: Dict[str, Any] = {"ok": False, "estado": ent, "fuente": None, "error": None}
    # 1 · archivo local versionado
    try:
        slug = re.sub(r"[^A-Z0-9]+", "_", ent).strip("_").lower()
        f = OD_DIR_LOCAL / f"{slug}.csv"
        if f.is_file():
            res.update(ok=True, fuente=f"local:{f.name}",
                       flujos=_od_parse_canonico(f.read_text(encoding="utf-8")))
            OD_CACHE[ent] = res
            return res
    except Exception as e:
        res["error"] = f"local: {str(e)[:80]}"
    # 2 · autofetch remoto (URL plantilla verificable vía /api/od/status)
    if OD_AUTOFETCH and OD_URL_TPL:
        try:
            abrev = ESTADO_ABREV_INEGI.get(ent, ent.replace(" ", "").lower())
            url = OD_URL_TPL.format(abrev=abrev, estado=ent.replace(" ", "%20"))
            async with httpx.AsyncClient(timeout=90) as client:
                r = await client.get(url, headers={"User-Agent": "Dataria/2.0"})
                r.raise_for_status()
                contenido = r.content
            if contenido[:2] == b"PK":   # zip o xlsx
                # ¿Es el TABULADO INEGI de movilidad? (xlsx con hojas 14/02) → parser real
                try:
                    flujos = _od_parse_tabulado_inegi(contenido)
                    res.update(ok=True, fuente=f"autofetch_tabulado:{url[:80]}", flujos=flujos)
                    OD_CACHE[ent] = res
                    return res
                except Exception:
                    pass
                zf = zipfile.ZipFile(io.BytesIO(contenido))
                csvs = [n for n in zf.namelist() if n.lower().endswith(".csv")]
                if csvs:
                    texto = zf.read(csvs[0]).decode("utf-8", errors="ignore")
                else:   # xlsx genérico → volcar a canónico
                    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
                    ws = wb.active
                    texto = "\n".join(",".join("" if c is None else str(c) for c in row)
                                      for row in ws.iter_rows(values_only=True))
            else:
                texto = contenido.decode("utf-8", errors="ignore")
            res.update(ok=True, fuente=f"autofetch:{url[:80]}",
                       flujos=_od_parse_canonico(texto))
            OD_CACHE[ent] = res
            return res
        except Exception as e:
            res["error"] = (res.get("error") or "") + f" · autofetch: {str(e)[:90]}"
    if not OD_URL_TPL:
        res["error"] = (res.get("error") or "") + " · DATARIA_OD_URL_TPL sin configurar"
    OD_CACHE[ent] = res
    return res


def _od_marginales_hacia(flujos: Dict[str, Dict[tuple, float]], muni_destino: str) -> Dict[str, Dict[str, float]]:
    """{motivo: {muni_origen: viajes}} de los flujos que ENTRAN al municipio del pin."""
    d = _norm_nombre(muni_destino)
    out: Dict[str, Dict[str, float]] = {}
    for motivo, m in (flujos or {}).items():
        ac: Dict[str, float] = {}
        for (o, dd), v in m.items():
            if dd == d and o != d:
                ac[o] = ac.get(o, 0.0) + v
        if ac:
            out[motivo] = ac
    return out


async def derive_mercado_captable_v2(agebs_geo_capt, natural_ring, pin_lng, pin_lat,
                                     anillo_min, fuente_iso, poligono_captable,
                                     muni_pin: Optional[str], estado_pin: Optional[str]) -> Dict[str, Any]:
    """Captable ANCLADO al censo: el share por municipio de origen es el OBSERVADO
    (flujos trabajo+estudio hacia el municipio del pin) y dentro de cada municipio se
    desagrega por gravedad (Huff) entre sus AGEBs. Municipios de la corona sin flujo
    censal → residual por modelo, DECLARADO. Sin ancla → v1 Huff (declarado)."""
    v1 = derive_mercado_captable(agebs_geo_capt, natural_ring, pin_lng, pin_lat,
                                 anillo_min, fuente_iso, poligono_captable)
    if not v1.get("activo") or not muni_pin:
        return v1
    ancla = await _od_cargar_estado(estado_pin or "")
    if not ancla.get("ok"):
        v1["ancla_censal"] = {"ok": False, "detalle": ancla.get("error"),
                              "nota": "Sin ancla censal disponible · share por modelo Huff (v1)."}
        return v1
    marg = _od_marginales_hacia(ancla["flujos"], muni_pin)
    entr = {}
    for motivo, mm in marg.items():
        for o, v in mm.items():
            entr[o] = entr.get(o, 0.0) + v
    modo_ancla = "destinos_censo"
    if not entr:
        # v2.1 · TABULADO (volúmenes de salida observados, sin destino específico):
        # la masa commuter observada de cada municipio ancla los VOLÚMENES; el destino
        # (hacia el pin) se modela por gravedad. Declarado como ancla PARCIAL.
        entr = _od_salientes(ancla["flujos"], muni_pin)
        modo_ancla = "salientes_tabulado"
    if not entr:
        v1["ancla_censal"] = {"ok": True, "fuente": ancla.get("fuente"),
                              "nota": f"El ancla censal no registra flujos hacia {muni_pin} · share por modelo (v1)."}
        return v1
    tot_censo = sum(entr.values())
    # Peso Huff por municipio dentro de la corona (para desagregar y para el residual)
    peso_muni: Dict[str, float] = {}
    for o in v1["origenes"]:
        m = _norm_nombre(o.get("municipio"))
        peso_muni[m] = peso_muni.get(m, 0.0) + (o.get("share_pct") or 0.0)
    con_ancla = {m: entr[m] for m in peso_muni if m in entr}
    cobertura = sum(con_ancla.values()) / tot_censo if tot_censo else 0.0
    share_censo_total = sum(peso_muni[m] for m in con_ancla) or 1.0
    origenes_v2 = []
    for o in v1["origenes"]:
        m = _norm_nombre(o.get("municipio"))
        o2 = dict(o)
        if m in con_ancla and peso_muni.get(m):
            # share municipal ANCLADO × distribución Huff dentro del municipio
            share_m_obs = con_ancla[m] / tot_censo * 100.0
            o2["share_pct"] = round(share_m_obs * ((o.get("share_pct") or 0.0) / peso_muni[m]), 1)
            o2["fuente_share"] = ("censo_2020" if modo_ancla == "destinos_censo"
                                  else "tabulado_salientes_2020")
            o2["viajes_censales_muni"] = round(con_ancla[m])
        else:
            o2["fuente_share"] = "modelo_huff"
        origenes_v2.append(o2)
    # Renormalizar a 100 conservando la proporción observado/modelo
    s = sum(x["share_pct"] or 0 for x in origenes_v2) or 1.0
    for x in origenes_v2:
        x["share_pct"] = round((x["share_pct"] or 0) / s * 100.0, 1)
    origenes_v2.sort(key=lambda x: -(x["share_pct"] or 0))
    v1["origenes"] = origenes_v2
    v1["metodo"] = "od_sintetico_inegi_v1"
    v1["ancla_censal"] = {
        "ok": True, "fuente": ancla.get("fuente"),
        "modo": modo_ancla,
        "muni_destino": muni_pin,
        "viajes_totales_censo": round(tot_censo),
        "cobertura_corona_pct": round(cobertura * 100, 1),
        "confianza": (("anclado_censo" if cobertura >= OD_COBERTURA_MIN else "parcialmente_anclado")
                      if modo_ancla == "destinos_censo" else "anclado_parcial_salientes"),
        "motivos_disponibles": sorted((marg or ancla["flujos"]).keys()),
    }
    v1["nota"] = (("Share por municipio ANCLADO a flujos observados del Censo 2020 "
                   "(trabajo/estudio) y desagregado por gravedad dentro de cada municipio. "
                   "Municipios sin flujo censal → modelo declarado.")
                  if modo_ancla == "destinos_censo" else
                  ("Volúmenes de commuters por municipio ANCLADOS al Censo 2020 (movilidad "
                   "cotidiana: salen a trabajar/estudiar a otro municipio); el destino hacia "
                   "el pin se modela por gravedad (ancla PARCIAL · la matriz destino-específica "
                   "llegará de los microdatos · v2.2). NO es migración: son viajes diarios."))
    return v1


def _analisis_identidad(analisis_nombre, colonia, municipio, estado) -> Dict[str, Any]:
    """ZA-8 · Identidad universal del análisis, presente en TODOS los encabezados:
    'nombre_del_usuario · colonia · municipio · estado · vAAAAMMDDHHMM'.
    La versión usa hora del centro de México (UTC-6 fijo desde 2022, sin horario de
    verano). Componentes ausentes se OMITEN (integridad: no se inventa)."""
    now = _dt.datetime.now(_dt.timezone(_dt.timedelta(hours=-6)))
    version = now.strftime("%Y%m%d%H%M")
    partes = [str(p).strip() for p in (analisis_nombre, colonia, municipio, estado)
              if p and str(p).strip()]
    idstr = " · ".join(partes + [f"v{version}"]) if partes else f"v{version}"
    return {
        "analisis_nombre": (str(analisis_nombre).strip() if analisis_nombre else None),
        "analisis_version": version,
        "analisis_id_str": idstr,
        "analisis_fecha": now.strftime("%Y-%m-%d %H:%M"),
    }


def _zone_label(producto_modo: str, kpis: Dict) -> Optional[str]:
    """Etiqueta de la zona acorde al MODO analizado y a los KPIs REALES de oferta.
    Si no hay oferta (proyectos=0), lo indica explícitamente en vez de inventar cifras."""
    label_modo = {
        "vivienda_vertical": "Vivienda vertical",
        "vivienda_horizontal": "Vivienda horizontal",
        "lotes": "Lotes urbanizados", "industrial": "Industrial",
        "logistica": "Logística", "oficinas": "Oficinas", "hotel": "Hotel",
    }.get(producto_modo, "Análisis")
    n_proy = kpis.get("proyectos") or 0
    n_uds = kpis.get("unidades")
    if n_proy == 0:
        return f"{label_modo} · sin oferta en comercialización en la zona"
    uds_txt = f"{n_uds:,} unidades" if n_uds else "unidades N/D"
    return f"{label_modo} · {n_proy} proyectos en comercialización · {uds_txt}"


def assemble_zone_payload(req, profile, isocronas, resumen, ft, pagos, resumen_renta,
                          demografia, nse_dim, segments, productos, productos_renta,
                          comercio, perception, agebs_count,
                          agebs=None, ft_renta=None, agebs_geo=None,
                          producto_modo="vivienda_vertical") -> Dict[str, Any]:
    agebs = agebs or []
    ft_renta = ft_renta or []
    agebs_geo = agebs_geo or []
    proyectos = _build_proyectos(resumen)
    kpis = _build_kpis(proyectos, ft)
    inv = _build_inventario(ft)
    # F-C · clasificación comercial, medianas duales y producto estrella (RES-1/3/5)
    res_com = derive_resumen_comercial(ft, segments)

    # Centro = pin del predio
    center = [req.lat, req.lng]

    # IDENTIFICACIÓN DE ZONA: de la base ArcGIS (real), no del request. El front manda el pin;
    # la zona (municipio, estado) se DERIVA de los AGEBs que caen en la isócrona (fuente de verdad
    # geográfica de la base). La COLONIA no está en el XLSX de demografía: la captura el usuario en
    # el formulario (req.colonia). El país se fija a México (base nacional MX) si hay ubicación.
    municipio = demografia.get("municipio") or req.municipio or None
    estado = demografia.get("estado") or req.estado or None
    pais = demografia.get("pais") or req.pais or ("México" if (municipio or estado) else None)
    colonia = (req.colonia or "").strip() or None
    # NOMBRE de la zona = lo más específico disponible: colonia capturada; si no, municipio de la
    # base; si no, el zone_name del request. Integridad: si no hay nada, queda None.
    zone_name = colonia or municipio or req.zone_name or None
    # SUBTÍTULO = jerarquía completa con los componentes que existen (sin "N/D"). Si el nombre ya
    # es la colonia, el subtítulo empieza en municipio para no repetir la colonia.
    if colonia:
        subt_parts = [p for p in (municipio, estado, pais) if p]
    else:
        subt_parts = [p for p in (estado, pais) if p]  # nombre=municipio → subtítulo desde estado
    subtitle = " · ".join(subt_parts) if subt_parts else None

    # ZA-8 · Identidad universal del análisis (una sola vez por procesamiento)
    identidad = _analisis_identidad(getattr(req, "analisis_nombre", None), colonia, municipio, estado)

    # ZA-6 · percepción detallada (se usa en _zona_analisis Y como ancla de PROD-PERFIL)
    pd_detalle = derive_percepcion_detalle(perception, segments, demografia, productos)

    # DEM-1 · matriz de perfiles (cohorte × NSE × ingreso) con conservación por bucket
    dem1 = derive_segmentos_dem1(agebs, segments, demografia.get("personas_hogar"))
    # PROD-PERFIL · resta oferta−demanda por perfil + producto sugerido anclado a percepción
    prod_perfil = derive_producto_perfil(dem1["segmentos"], ft, pd_detalle)
    dem1["meta"]["producto_perfil"] = {"resumen": prod_perfil["resumen"],
                                       "nota": prod_perfil["nota"]}

    # Derivaciones de demanda/renta que dependen de varios insumos
    renta_segmentos = derive_renta_segmentos(segments, agebs)
    demanda_segmentos = derive_demanda_segmentos(segments, productos)
    recamaras = derive_recamaras(agebs)

    # Enlazar m² y recámaras de los productos de renta a los renta_segmentos (por bucket)
    if productos_renta and renta_segmentos:
        pr_by_bucket = {p.get("seg_renta", "").split(" · ")[-1]: p for p in productos_renta}
        for rs in renta_segmentos:
            pr = pr_by_bucket.get(rs.get("seg"))
            if pr:
                if pr.get("m2") and pr["m2"] != "N/D":
                    rs["m2"] = pr["m2"].replace(" m²", "")
                if pr.get("rec") and pr["rec"] != "N/D":
                    rs["rec"] = pr["rec"]

    # Renta baseline para el simulador: medianas reales de la oferta de renta
    rentas_oferta = []
    m2_oferta = []
    for t in ft_renta:
        a = t.get("attributes", {})
        ap = _num(a.get("ÁREA_PRIVATIVA")); pr = _num(a.get("F____UNIDAD"))
        if ap and 25 <= ap <= 500 and pr and pr > 1000:
            rentas_oferta.append(pr / ap); m2_oferta.append(ap)
    rb_m2 = round(statistics.median(m2_oferta)) if m2_oferta else None
    rb_pm2 = round(statistics.median(rentas_oferta)) if rentas_oferta else None
    # Fallback: si no hay oferta de renta levantada, derivar del producto de renta featured (sweet spot)
    if (rb_m2 is None or rb_pm2 is None) and productos_renta:
        sweet = next((p for p in productos_renta if p.get("featured")), productos_renta[0])
        import re as _re
        if rb_m2 is None and sweet.get("m2") and sweet["m2"] != "N/D":
            nums = _re.findall(r"\d+", sweet["m2"])
            if nums:
                rb_m2 = int(nums[0])
        if rb_pm2 is None and sweet.get("pm2_renta") and sweet["pm2_renta"] != "N/D":
            pn = _re.findall(r"[\d,]+", sweet["pm2_renta"])
            if pn:
                rb_pm2 = int(pn[0].replace(",", ""))
    renta_baseline = {
        "m2": rb_m2, "pm2": rb_pm2,
        "units": 120 if rb_m2 else None,   # supuesto de proyecto típico para el simulador
        "occ": 90 if rb_m2 else None,
    }

    # Isócronas como GeoJSON (para que el tablero pinte los anillos)
    iso_out = {str(m): geo for m, geo in isocronas.items()}

    zone_data = {
        "name": zone_name,
        "subtitle": subtitle,
        # ZA-8 · identidad del análisis (el front la muestra en TODOS los encabezados)
        "analisis_nombre": identidad["analisis_nombre"],
        "analisis_version": identidad["analisis_version"],
        "analisis_id_str": identidad["analisis_id_str"],
        "analisis_fecha": identidad["analisis_fecha"],
        # F-C · resumen comercial (activos/agotados, medianas duales, producto estrella)
        "resumen_comercial": res_com["proyectos"],
        "oferta_stats": res_com["oferta_stats"],
        "top_estrella": res_com["top_estrella"],
        "criterio_estrella": res_com["criterio_estrella"],
        # DEM-1 · perfiles de demanda (cohorte × NSE × ingreso · conservación por bucket)
        "segmentos_dem1": dem1["segmentos"],
        "dem1_meta": dem1["meta"],
        "municipality": municipio,
        "label": _zone_label(producto_modo, kpis),
        "center": center, "zoom": 13,
        "population": demografia["population"],
        "households": demografia["households"],
        "tca": demografia["tca"],
        "personas_hogar": demografia["personas_hogar"],
        "ingreso_hogar": demografia["ingreso_hogar"],
        "ingreso_total_anual": demografia["ingreso_total_anual"],
        "nse": demografia["nse"],
        "nse_dominante": demografia["nse_dominante"],
        "tenencia": demografia["tenencia"],
        "hog_renta": demografia["hog_renta"],
        "edad_grupos": demografia["edad_grupos"],
        "recamaras": recamaras,
        "proyectos": proyectos,
        "kpis": kpis,
        "avgAbs": kpis["avg_abs"], "avgPm2": kpis["avg_pm2"],
        "inventario_precio": inv["inventario_precio"],
        "inventario_m2": inv["inventario_m2"],
        "productos": productos,
        "productos_renta": productos_renta,
        "renta_segmentos": renta_segmentos,
        "demanda_segmentos": demanda_segmentos,
        "renta_baseline": renta_baseline,
        "sensibilidad_baseline": _build_sensibilidad_baseline(productos, segments),
        "comercio": comercio,
        "ingreso_anual": comercio.get("ingreso_anual"),
        "renta_low": comercio.get("renta_low"),
        "renta_high": comercio.get("renta_high"),
        "demanda": comercio.get("demanda", {}),
        # Metadatos de la zona de análisis (sección Zona de análisis)
        "_zona_analisis": {
            "perfil_iso": profile["key"], "perfil_label": profile["label"],
            "isocronas": iso_out,
            "predio_m2": req.predio_m2, "uso_comercial": req.uso_comercial,
            "norm": req.norm.model_dump() if req.norm else None,
            "perception": {
                "n_total": perception["n_total"], "n_zona": perception["n_zona"],
                "media": round(perception["media"]) if perception.get("media") else None,
                "sd": round(perception["sd"]) if perception.get("sd") else None,
                "cv": round(perception["cv"], 4) if perception.get("cv") is not None else None,
                "barrera": perception["barrera"], "metodo": perception["metodo"],
                "cobertura_pct": perception["cobertura_pct"], "motivo": perception["motivo"],
                "sin_valor_percibido": perception.get("sin_valor_percibido", False),
                "nota_valor": perception.get("nota_valor"),
                "valor_zona": perception.get("valor_zona"),
                "proyectos": perception["proyectos"],
                "cluster_names": perception.get("cluster_names"),
                "ajuste_inventario": perception.get("ajuste_inventario"),
                "zona_poligono": perception.get("zona_poligono"),
                "mercados": perception.get("mercados"),
                "competidores": perception.get("competidores"),
            },
            "nse_barrier": _nse_barrier_info(agebs, agebs_geo),
            # ZA-6 · delimitación de percepción de valor + mercado meta (bandas robustas)
            "percepcion_detalle": pd_detalle,
            "agebs_geo": agebs_geo,
        },
        "_vars": build_analysis_vars(req, profile, isocronas, perception, demografia,
                                     segments, productos, productos_renta, proyectos, kpis),
        "_typologies": _build_typologies(
            ft,
            {p.get("name") for grp in ("directos", "primarios", "secundarios")
             for p in ((perception or {}).get("competidores") or {}).get(grp, [])
             if p.get("name") and p.get("name") != "N/D"}
        ),
    }
    # ZA-8 · la identidad también vive en _vars (fuente única que leen todas las secciones)
    try:
        zone_data["_vars"].update(identidad)
    except Exception:
        pass

    # DIM_DATA (para tplDemanda)
    totals = {
        "mkt_total": sum(s["mkt_total"] for s in segments),
        "mkt_venta": sum((s.get("mkt_venta") or 0) for s in segments),
        "mkt_renta": sum((s.get("mkt_renta") or 0) for s in segments),
        "nuevas_fam_anual": round(sum(s["nuevas_fam"] for s in segments), 1),
    }
    dim_data = {
        "nse_dim": nse_dim,
        "segments": segments,
        "totals": totals,
        "mercado_meta": _build_mercado_meta(nse_dim, demografia),
        "civil": _build_civil(agebs),
        "hog_comp": _build_hog_comp(agebs),
        "di_detail": _build_di_detail(agebs),
    }

    return {
        "ok": True,
        "meta": {
            "agebs": agebs_count,
            "proyectos_vvv": len(resumen),
            "tipologias_vvv": len(ft),
            "demografia_disponible": agebs_count > 0,
        },
        "zone_data": zone_data,
        "dim_data": dim_data,
    }


def _build_mercado_meta(nse_dim, demografia):
    if not nse_dim:
        return []
    # Mapea a la estructura de filas que el template indexa por [1..7]
    labels = [("Hogares", "hogares"), ("Niños (0-9)", "ninos"),
              ("Adolescentes (10-15)", "adolescentes"), ("Jóvenes (16-19)", "jovenes"),
              ("Jóvenes adultos (20-29)", "jov_adultos"), ("Consolidados (30-54)", "consolidados"),
              ("Adultos maduros (55+)", "empty_nest")]
    by_nse = {d["NSE"]: d for d in nse_dim}
    out = []
    for label, key in labels:
        row = {"label": label}
        total = 0
        for nse_k in ["A", "B", "C+", "C", "D+", "D", "E"]:
            v = round(by_nse.get(nse_k, {}).get(key, 0)) if by_nse.get(nse_k) else 0
            col = "Cm" if nse_k == "C+" else "Dm" if nse_k == "D+" else nse_k
            row[col] = v
            total += v
        row["Total"] = total
        row["pct"] = None
        out.append(row)
    # pct sobre población total (filas etarias) usando la fila Hogares como base poblacional alterna
    pob_total = sum(r["Total"] for r in out if r["label"].startswith(("Niños", "Adolesc", "Jóvenes", "Consolid", "Adultos")))
    for r in out:
        if r["label"].startswith(("Niños", "Adolesc", "Jóvenes", "Consolid", "Adultos")) and pob_total:
            r["pct"] = round(r["Total"] / pob_total * 100, 2)
    return out


def _build_civil(agebs):
    """Estado civil por grupo de edad (% sobre población del grupo). Desde campos C107-C166."""
    if not agebs:
        return {}
    grupos = ["20 a 24", "25 a 29", "30 a 34", "35 a 39", "40 a 44",
              "45 a 49", "50 a 54", "55 a 59", "60 a 64", "65 a 69", "70 a 74", "75 y Más"]
    estados = ["soltera", "casada", "en union libre", "separad  divorcia", "separad"]
    out = {}
    pob_total_zona = 0.0
    for r in agebs:
        pob_total_zona += _num(r.get("Población total 2026")) or _num(r.get("POB1")) or 0
    for g in grupos:
        acc = {"soltera": 0.0, "casada": 0.0, "union_libre": 0.0, "separada": 0.0, "no_esp": 0.0, "total": 0.0}
        present = False
        for r in agebs:
            s = _num(r.get(f"{g} soltera"))
            c = _num(r.get(f"{g} casada"))
            u = _num(r.get(f"{g} en union libre"))
            sep = _num(r.get(f"{g} separad  divorcia"))
            if sep is None:
                sep = _num(r.get(f"{g} separad"))
            if any(x is not None for x in [s, c, u, sep]):
                present = True
                acc["soltera"] += s or 0
                acc["casada"] += c or 0
                acc["union_libre"] += u or 0
                acc["separada"] += sep or 0
        if not present:
            continue
        tot = acc["soltera"] + acc["casada"] + acc["union_libre"] + acc["separada"]
        acc["total"] = tot
        if tot <= 0:
            continue
        out[g] = {
            "pob_pct": round(tot / pob_total_zona * 100, 2) if pob_total_zona else 0,
            "soltera": round(acc["soltera"] / tot * 100, 2),
            "casada": round(acc["casada"] / tot * 100, 2),
            "union_libre": round(acc["union_libre"] / tot * 100, 2),
            "separada": round(acc["separada"] / tot * 100, 2),
            "no_esp": 0.0,
        }
    return out


def _build_sensibilidad_baseline(productos, segments):
    """Baseline del simulador de Producto: del producto featured (sweet spot) o el primero recomendado."""
    import re as _re
    base = next((p for p in productos if p.get("featured")), None)
    if base is None:
        base = next((p for p in productos if p.get("recomendado")), None)
    if base is None and productos:
        base = productos[0]
    if base is None:
        return {"m2": None, "pm2": None, "abs_base": None, "ticket_base": None}

    def _parse_int(s):
        nums = _re.findall(r"[\d,]+", str(s).replace(",", ""))
        return int(nums[0]) if nums else None
    def _parse_float(s):
        nums = _re.findall(r"[\d.]+", str(s))
        return float(nums[0]) if nums else None

    m2 = _parse_int(base.get("m2")) if base.get("m2") != "N/D" else None
    pm2 = _parse_int(base.get("pm2")) if base.get("pm2") != "N/D" else None
    ticket = _parse_float(base.get("ticket")) if base.get("ticket") != "N/D" else None
    # abs base: la MISMA absorción que muestra el producto featured (incluye evidencia/competencia),
    # para que el simulador de sensibilidad parta del valor visible. Fallback: nuevas_fam/12.
    abs_base = None
    if base.get("abs") and base["abs"] != "N/D":
        abs_base = _parse_float(base["abs"])   # "2.2 un/mes" → 2.2
    if abs_base is None and base.get("nuevas_fam"):
        abs_base = round(base["nuevas_fam"] / 12, 2)
    return {"m2": m2, "pm2": pm2, "abs_base": abs_base, "ticket_base": ticket}


def _build_di_detail(agebs):
    """Detalle demográfico real de la zona (tipo de vivienda, situación conyugal, población en
    hogares, composición), en el formato que el front consume. Reemplaza el bloque hardcodeado:
    todo proviene de los AGEBs de la isócrona actual. Ausencia → listas vacías (no se inventa)."""
    if not agebs:
        return {"tipo_vivienda": [], "situacion_conyugal": [], "situacion_conyugal_total": 0,
                "poblacion_hogares": [], "tipologia_hogar": [], "personas_hogar": {}}

    def s(col):
        v = _sum(agebs, col)
        return round(v) if v is not None else 0

    # ── Tipo de vivienda (conteo real) ──
    tv = [
        ("Casa", s("Casa")),
        ("Departamento o edificio", s("Departamento o edificio")),
        ("Vivienda en vecindad o cuartería", s("Vivienda en vecidad o cuartería ") or s("Vivienda en vecidad o cuartería")),
        ("Otro tipo de vivienda", s("Otro tipo de vivienda")),
    ]
    tv_total = sum(c for _, c in tv) or 1
    tipo_vivienda = [{"label": l, "count": c, "pct": round(c / tv_total * 100, 2)}
                     for l, c in tv if c > 0]

    # ── Situación conyugal (suma de todas las bandas de edad) ──
    edades = ["20 a 24", "25 a 29", "30 a 34", "35 a 39", "40 a 44", "45 a 49", "50 a 54",
              "55 a 59", "60 a 64", "65 a 69", "70 a 74", "75 y Más"]
    conyugal_cols = {"Soltera": ["soltera"], "Casada": ["casada"],
                     "En unión libre": ["en union libre"],
                     "Separada, divorciada o viuda": ["separad  divorcia", "separad"]}
    cony = {}
    for label, sufs in conyugal_cols.items():
        tot = 0
        for e in edades:
            for suf in sufs:
                val = s(f"{e} {suf}")
                if val:
                    tot += val
                    break   # usar el primer sufijo que exista para esa edad
        cony[label] = tot
    cony_total = sum(cony.values()) or 1
    situacion_conyugal = [{"label": l, "count": c, "pct": round(c / cony_total * 100, 2)}
                          for l, c in cony.items() if c > 0]

    # ── Composición de hogares (población en hogares) ──
    hc = _build_hog_comp(agebs)
    fam = hc.get("familiares_total") or 0
    nofam = hc.get("no_familiares_total") or 0
    th_total = (fam + nofam) or 1
    tipologia_hogar = []
    if fam:
        tipologia_hogar.append({"label": "Hogares familiares", "count": fam,
                                "pct": round(fam / th_total * 100, 2), "parent": True})
        for sub, key in [("Nucleares", "nucleares"), ("Ampliados", "ampliados"), ("Compuestos", "compuestos")]:
            c = hc.get(key) or 0
            if c:
                tipologia_hogar.append({"label": sub, "count": c,
                                        "pct": round(c / fam * 100, 2), "parent_group": "familiares"})
    if nofam:
        tipologia_hogar.append({"label": "Hogares no familiares", "count": nofam,
                                "pct": round(nofam / th_total * 100, 2), "parent": True})
        for sub, key in [("Unipersonales", "unipersonal"), ("Corresidentes", "corresidentes")]:
            c = hc.get(key) or 0
            if c:
                tipologia_hogar.append({"label": sub, "count": c,
                                        "pct": round(c / nofam * 100, 2), "parent_group": "no_familiares"})

    personas_hogar = _wavg(agebs, "Personas por hogar 2026", "Hogares totales 2026")
    # Personas por hogar familiar / no familiar = población del tipo / hogares del tipo (dato real).
    pob_fam_real = s("Población familiar total")
    pob_nofam_real = s("Población no familiar total")
    ph_fam = (pob_fam_real / fam) if (pob_fam_real and fam) else _wavg(agebs, "PXHFAMILIARES", "Hogares familiares totales")
    ph_nofam = (pob_nofam_real / nofam) if (pob_nofam_real and nofam) else _wavg(agebs, "PXHNOFAMILIARES", "Hogares  no familiares totales")
    # Población en hogares (familiares vs no familiares) — estructura paralela a tipologia_hogar
    pob_fam = pob_fam_real or (round(fam * ph_fam) if (fam and ph_fam) else 0)
    pob_nofam = pob_nofam_real or (round(nofam * ph_nofam) if (nofam and ph_nofam) else 0)
    pob_h_total = (pob_fam + pob_nofam) or 1
    poblacion_hogares = []
    if pob_fam:
        poblacion_hogares.append({"label": "Población en hogares familiares", "count": pob_fam,
                                  "pct": round(pob_fam / pob_h_total * 100, 2), "parent": True})
    if pob_nofam:
        poblacion_hogares.append({"label": "Población en hogares no familiares", "count": pob_nofam,
                                  "pct": round(pob_nofam / pob_h_total * 100, 2), "parent": True})

    return {
        "tipo_vivienda": tipo_vivienda,
        "situacion_conyugal": situacion_conyugal,
        "situacion_conyugal_total": round(cony_total) if cony_total != 1 else 0,
        "tipologia_hogar": tipologia_hogar,
        "poblacion_hogares": poblacion_hogares,
        "personas_hogar": {
            "promedio_ponderado": round(personas_hogar, 2) if personas_hogar else None,
            "familiares": round(ph_fam, 2) if ph_fam else (round(personas_hogar, 2) if personas_hogar else None),
            "no_familiares": round(ph_nofam, 2) if ph_nofam else None,
        },
    }


def _build_hog_comp(agebs):
    """Composición de hogares (C60-C67). Valores absolutos."""
    if not agebs:
        return {"familiares_total": None, "no_familiares_total": None, "nucleares": None,
                "ampliados": None, "compuestos": None, "unipersonal": None, "corresidentes": None}
    def s(col):
        v = _sum(agebs, col)
        return round(v) if v is not None else 0
    return {
        "familiares_total": s("Hogares familiares totales"),
        "nucleares": s("Hogares nucleares"),
        "ampliados": s("Hogares  ampliados"),
        "compuestos": s("Hogares compuestos"),
        "no_familiares_total": s("Hogares  no familiares totales"),
        "unipersonal": s("Hogares unipersonales"),
        "corresidentes": s("Hogares corresidentes"),
    }


def _perfiles_por_segmento(nse, m2, ticket_M):
    """
    Asigna perfiles de comprador (PROFILE_CATALOG estándar) según NSE, tamaño y ticket.
    La demanda se mide por PERFIL de comprador, no solo por bucket de precio.
    Universal para todo México.
    """
    perfiles = []
    m2 = m2 or 0
    # Por tamaño (tipología) → ciclo de vida del comprador
    if m2 and m2 < 50:
        perfiles += ["joven_profesional", "inversionista", "soltero"]
    elif m2 < 75:
        perfiles += ["dink_accesible", "joven_profesional", "inversionista"]
    elif m2 < 110:
        perfiles += ["familia_joven", "dink_premium", "foraneo_ejecutivo"]
    elif m2 < 160:
        perfiles += ["familia_consolidada", "familia_joven", "empty_nester"]
    else:
        perfiles += ["familia_consolidada", "empty_nester", "inversionista"]
    # Por NSE (poder adquisitivo) → matiz premium/accesible
    if nse in ("A", "B") and "dink_premium" not in perfiles:
        perfiles.append("dink_premium")
    if nse in ("D+", "D", "E") and "soltero" not in perfiles:
        perfiles.append("soltero")
    # Tickets muy altos → patrimonial/inversionista
    if ticket_M and ticket_M >= 12 and "inversionista" not in perfiles:
        perfiles.append("inversionista")
    # Únicos, máximo 3
    seen = []
    for p in perfiles:
        if p not in seen:
            seen.append(p)
    return seen[:3]


# Etiquetas legibles del perfil de comprador (para el campo "perfil" de la tabla de demanda)
PERFIL_LABEL = {
    "joven_profesional": "Joven profesional", "dink_accesible": "DINK accesible",
    "familia_joven": "Familia joven", "familia_consolidada": "Familia consolidada",
    "dink_premium": "DINK premium", "empty_nester": "Empty nester",
    "inversionista": "Inversionista", "foraneo_ejecutivo": "Foráneo ejecutivo",
    "soltero": "Soltero/a",
}


def derive_demanda_segmentos(segments, productos=None):
    """Tabla superior de Demanda (z.demanda_segmentos). Mide por PERFIL de comprador.
    Toma m² y absorción de los productos derivados y asigna el perfil de comprador real."""
    productos = productos or []
    prod_by_bucket = {}
    for p in productos:
        prod_by_bucket[p.get("seg_dim", "")] = p
    total_mkt = max(sum((x.get("mkt_venta") or 0) for x in segments), 1)
    out = []
    for s in segments:
        if s.get("status") in ("bajo_crecimiento",):
            continue
        key = f"{s['NSE']} · {s['bucket']}"
        prod = prod_by_bucket.get(key)
        # m² y absorción del producto correspondiente
        m2_num = None
        m2 = "N/D"
        if prod and prod.get("m2") and prod["m2"] != "N/D":
            m2 = prod["m2"].replace(" m²", "")
            try: m2_num = int(m2)
            except ValueError: m2_num = None
        abs_val = "N/D"
        if prod and prod.get("abs") and prod["abs"] != "N/D":
            abs_val = prod["abs"].replace(" un/mes", "")
        # Ticket coherente (bucket inferior val_min=0 → usar val_max*0.75)
        vmin = s.get("val_min"); vmax = s.get("val_max")
        if vmin and vmax:
            ticket_M = (vmin + vmax) / 2 / 1e6
        elif vmax:
            ticket_M = vmax * 0.75 / 1e6
        else:
            ticket_M = None
        ticket_str = f"${ticket_M:.2f}M" if ticket_M else "N/D"
        # PERFIL de comprador (no status)
        perfiles = _perfiles_por_segmento(s["NSE"], m2_num, ticket_M)
        perfil_label = " · ".join(PERFIL_LABEL.get(p, p) for p in perfiles[:2]) if perfiles else "N/D"
        out.append({
            "segmento": prod.get("tipo") if prod else f"NSE {s['NSE']}",
            "perfil": perfil_label,
            "nse": s["NSE"],
            "ticket": ticket_str,
            "m2": m2,
            "abs": abs_val,
            "mix": round(s["mkt_venta"] / total_mkt * 100) if s.get("mkt_venta") else 0,
            "perfiles": perfiles,
            "origen": s.get("origen", "demand_driven"),
        })
    return out


def derive_renta_segmentos(segments, agebs=None):
    """
    Segmentos de renta (z.renta_segmentos) derivados de los mismos segments DPO de venta.
    Cada bucket de valor de vivienda tiene su renta-equivalente mensual (rent_min/rent_max,
    ~0.4% del valor) y su mercado de renta (mkt_renta = hogares que rentan en el bucket).
    Universal: la renta se ancla a la capacidad de compra y a la propensión a rentar de la zona.
    """
    if not segments:
        return []
    # Sweet spot de renta = bucket con mayor mercado de renta (mkt_renta)
    best = max(segments, key=lambda s: s.get("mkt_renta", 0) or 0, default=None)
    out = []
    for s in segments:
        if s.get("status") == "bajo_crecimiento":
            continue
        rent_min = s.get("rent_min"); rent_max = s.get("rent_max")
        ing_min = s.get("ing_min"); ing_max = s.get("ing_max")
        # m² del segmento: lo toma del producto de venta correspondiente si existe; si no, N/D
        out.append({
            "seg": s.get("bucket", "N/D"),
            "perfil": _status_label(s.get("status", "")),
            "nse": s.get("NSE", "N/D"),
            "ing": (f"${round(ing_min/1000)}k-${round(ing_max/1000)}k"
                    if ing_min and ing_max else "N/D"),
            "renta": (f"${round(rent_min/1000,1)}k-${round(rent_max/1000,1)}k"
                      if rent_min is not None and rent_max else "N/D"),
            "rent_min": rent_min, "rent_max": rent_max,
            "rec": "N/D",   # se completa desde el producto de renta
            "m2": "N/D",    # se completa desde el producto de renta
            "hog": round(s.get("mkt_renta", 0)) if s.get("mkt_renta") else 0,
            "mkt_renta": s.get("mkt_renta", 0),
            "nuevas_fam": s.get("nuevas_fam", 0),
            "sweet": (best is not None and s is best),
        })
    return out


def derive_recamaras(agebs):
    """
    Distribución % de recámaras demandadas, derivada de la composición de hogares:
      1 rec ← unipersonales + corresidentes
      2 rec ← nucleares (parejas/familias jóvenes) · principal demanda vertical
      3 rec ← ampliados + compuestos (familias grandes)
    Devuelve [%1rec, %2rec, %3rec, %4+rec]. Si no hay data → None (N/D).
    """
    if not agebs:
        return None
    def s(col):
        v = _sum(agebs, col)
        return v if v is not None else 0
    uni = s("Hogares unipersonales") + s("Hogares corresidentes")
    nuc = s("Hogares nucleares")
    amp = s("Hogares  ampliados") + s("Hogares compuestos")
    total = uni + nuc + amp
    if total <= 0:
        return None
    # 2 rec captura nucleares + parte de unipersonales (DINK); 1 rec resto unipersonal
    r1 = uni * 0.55
    r2 = nuc * 0.70 + uni * 0.45
    r3 = nuc * 0.30 + amp * 0.70
    r4 = amp * 0.30
    tot = r1 + r2 + r3 + r4
    if tot <= 0:
        return None
    return [round(r1 / tot * 100, 1), round(r2 / tot * 100, 1),
            round(r3 / tot * 100, 1), round(r4 / tot * 100, 1)]


# ╔══════════════ SECCIÓN 3: SERVIDOR FASTAPI (main) ══════════════╗
import os
import io
import zipfile
import statistics
from typing import Optional, List, Dict, Any

import httpx
import openpyxl
from pathlib import Path
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel


PRSP_BASE = os.environ.get("PRSP_BASE", "https://payment-system-prsp.onrender.com")
HTTP_TIMEOUT = float(os.environ.get("HTTP_TIMEOUT", "180"))

# Directorio de estáticos: subcarpeta "static/" junto a este app.py.
# Se resuelve relativo al archivo para que funcione igual en local y en Render.
STATIC_DIR = Path(os.environ.get("DATARIA_STATIC_DIR", str(Path(__file__).resolve().parent / "static")))
TABLERO_FILE = STATIC_DIR / "dashboard_zona_analisis.html"

app = FastAPI(title="Dataria Orchestrator", version="2.0")

# CORS abierto para que el tablero (servido en cualquier origen http/https) consuma el API.
# Para producción, restringir allow_origins a los dominios de Dataria.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ════════════════════════════════════════════════════════════════════════════
# ARQUITECTURA MODULAR POR SECCIÓN · REGISTRO + CONTROL DE ACCESO
# ════════════════════════════════════════════════════════════════════════════
# Cada sección del tablero es un módulo lógico independiente: se puede optimizar,
# cambiar, agregar o eliminar SIN tocar las demás. El orquestador (zona_procesar)
# ensambla las secciones habilitadas. Esto prepara el terreno para las 3 secciones
# nuevas (vivienda_horizontal, lotes_urbanizados, explorador_nacional).
#
# Cada entrada declara:
#   • key          identificador estable de la sección
#   • label        nombre legible
#   • payload_keys campos del zone_data/dim_data que produce (para filtrado de acceso)
#   • access_tier  "preliminar" (visible en evaluación general) | "detalle" (requiere permiso)
#   • page_id      id de la página en el front (para habilitar/deshabilitar UI)
# La función derivadora vive en su propia función derive_* (ya modular); aquí solo
# se registra el metadato que permite ensamblar, filtrar por acceso y extender.

SECTION_REGISTRY = {
    "resumen":    {"label": "Resumen ejecutivo", "page_id": "page-resumen",
                   "access_tier": "preliminar",
                   "payload_keys": ["name", "subtitle", "kpis", "resumen_comercial",
                                    "oferta_stats", "top_estrella", "criterio_estrella"]},
    "mapa":       {"label": "Mapa de zona", "page_id": "page-mapa",
                   "access_tier": "preliminar", "payload_keys": ["center", "proyectos", "isocronas"]},
    "demografia": {"label": "Demografía / NSE", "page_id": "page-resumen",
                   "access_tier": "preliminar", "payload_keys": ["population", "households", "nse", "nse_dominante"]},
    "inventario": {"label": "Inventario / Oferta", "page_id": "page-inventario",
                   "access_tier": "detalle",
                   "payload_keys": ["inventario_precio", "inventario_m2", "proyectos",
                                    "oferta_stats", "top_estrella", "resumen_comercial"]},
    "demanda":    {"label": "Demanda por perfil", "page_id": "page-demanda",
                   "access_tier": "detalle",
                   "payload_keys": ["demanda_segmentos", "mercado_meta", "civil",
                                    "segmentos_dem1", "dem1_meta"]},
    "producto":   {"label": "Producto óptimo (DPO)", "page_id": "page-producto",
                   "access_tier": "detalle", "payload_keys": ["productos", "sensibilidad_baseline"]},
    "renta":      {"label": "Mercado de renta", "page_id": "page-renta",
                   "access_tier": "detalle", "payload_keys": ["productos_renta", "renta_segmentos", "renta_baseline"]},
    "comercio":   {"label": "Comercio / Tenant mix", "page_id": "page-comercio",
                   "access_tier": "detalle", "payload_keys": ["comercio"]},
    "mezcla":     {"label": "Crea tu mezcla · Venta", "page_id": "page-mezcla",
                   "access_tier": "detalle", "payload_keys": ["productos"]},
    "mezcla_renta": {"label": "Crea tu mezcla · Renta", "page_id": "page-mezclaRenta",
                   "access_tier": "detalle", "payload_keys": ["productos_renta"]},
    "monitor":    {"label": "Monitorea tu proyecto", "page_id": "page-monitor",
                   "access_tier": "detalle", "payload_keys": ["productos", "proyectos"]},
    # ── Secciones nuevas (andamiaje · se implementarán en el siguiente paso) ──
    "vivienda_horizontal": {"label": "Vivienda horizontal", "page_id": "page-vivienda-horizontal",
                   "access_tier": "detalle", "payload_keys": ["vivienda_horizontal"], "scaffold": True},
    "lotes_urbanizados": {"label": "Venta de lotes urbanizados", "page_id": "page-lotes",
                   "access_tier": "detalle", "payload_keys": ["lotes_urbanizados"], "scaffold": True},
    "explorador_nacional": {"label": "Explorador nacional", "page_id": "page-explorador",
                   "access_tier": "preliminar", "payload_keys": ["explorador_nacional"], "scaffold": True},
}


def section_keys(include_scaffold: bool = False) -> List[str]:
    """Lista de secciones registradas. Por defecto excluye las de andamiaje (aún no implementadas)."""
    return [k for k, v in SECTION_REGISTRY.items() if include_scaffold or not v.get("scaffold")]


# ── Control de acceso por cuenta (preparación del modelo de operación) ──
# Cada cuenta declara: secciones habilitadas y zonas autorizadas para DETALLE.
# El acceso GENERAL (preliminar) está disponible aunque el DETALLE esté restringido:
# así un usuario puede hacer evaluaciones preliminares de cualquier zona, pero solo
# ve el detalle completo de las zonas que tiene autorizadas.

ACCESS_PRELIMINAR = "preliminar"   # KPIs generales, mapa, demografía, NSE dominante
ACCESS_DETALLE = "detalle"         # producto, demanda, renta, comercio, mezcla, monitor


def filter_payload_by_access(payload: Dict[str, Any], permisos: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Filtra el payload de una zona según los permisos de la cuenta.
    permisos = {
        "secciones": ["resumen","mapa",...] | "*",   # secciones habilitadas
        "zonas_detalle": ["ZMM_VallePoniente",...] | "*",  # zonas con acceso a DETALLE
        "zona_actual": "ZMM_Centro",
    }
    Si la zona actual NO está autorizada para detalle, se entregan SOLO las secciones de
    tier 'preliminar' (evaluación general), y las de 'detalle' se sustituyen por un marcador
    de acceso restringido (sin exponer datos). Si permisos es None → acceso completo (admin/local).
    """
    if not permisos:
        return payload   # sin restricciones (modo local/admin)

    secciones_hab = permisos.get("secciones", "*")
    zonas_detalle = permisos.get("zonas_detalle", "*")
    zona_actual = permisos.get("zona_actual")

    zona_autorizada_detalle = (zonas_detalle == "*" or
                               (zona_actual and zona_actual in (zonas_detalle or [])))

    out = dict(payload)
    z = dict(payload.get("zone_data", {}))
    restringidas = []

    for key, meta in SECTION_REGISTRY.items():
        if meta.get("scaffold"):
            continue
        # ¿La sección está habilitada para esta cuenta?
        sec_habilitada = (secciones_hab == "*" or key in (secciones_hab or []))
        # ¿Requiere detalle y la zona no está autorizada?
        es_detalle = meta["access_tier"] == ACCESS_DETALLE
        bloquear = (not sec_habilitada) or (es_detalle and not zona_autorizada_detalle)
        if bloquear:
            for pk in meta["payload_keys"]:
                if pk in z:
                    z[pk] = None   # se oculta el dato (el front muestra "acceso restringido")
            restringidas.append(key)

    out["zone_data"] = z
    out["_access"] = {
        "zona_autorizada_detalle": bool(zona_autorizada_detalle),
        "secciones_restringidas": sorted(set(restringidas)),
        "modo": "detalle" if zona_autorizada_detalle else "preliminar",
    }
    return out



class Normatividad(BaseModel):
    usos: Optional[str] = None
    densidad: Optional[float] = None
    altura: Optional[float] = None
    cus: Optional[float] = None
    cos: Optional[float] = None
    cav: Optional[float] = None
    estac: Optional[str] = None


class ZonaRequest(BaseModel):
    lat: float
    lng: float
    # ZA-8 · Nombre que el usuario da al análisis (encabeza la identidad universal
    # 'nombre · colonia · municipio · estado · vAAAAMMDDHHMM' en todos los tableros).
    analisis_nombre: Optional[str] = None
    predio_m2: Optional[float] = None
    uso_comercial: bool = False
    zone_name: Optional[str] = None
    colonia: Optional[str] = None
    municipio: Optional[str] = None
    estado: Optional[str] = None
    pais: Optional[str] = None
    norm: Optional[Normatividad] = None
    # Modo de producto a analizar. El front (selector superior) define cuál.
    # "vivienda_vertical" (default, no rompe nada) reusa el pipeline existente.
    # "vivienda_horizontal" usa la capa vh_venta y el diseño de producto de casa.
    # Los demás (lotes, industrial, logistica, oficinas, hotel) quedan declarados
    # pero aún no implementados: el backend responde no_disponible para ellos.
    producto: str = "vivienda_vertical"
    # ISO-MULTI · fuente de isócronas: predik (default) | valhalla | ors | tomtom.
    # None → ISO_FUENTE_DEFAULT. La lógica de negocio no cambia con la fuente.
    iso_fuente: Optional[str] = None
    # CAPTABLE · SIEMPRE incluido (directiva Héctor 8 jul 2026): el mercado natural y el
    # captable son variables SEPARADAS que las siguientes secciones procesan distinto,
    # pero este paso siempre considera ambos. (El flag permite apagarlo solo para pruebas.)
    incluir_captable: bool = True
    captable_min: Optional[int] = None   # None → principal+10, tope CAPTABLE_MIN_TOPE
    # Unidades que el proyecto nuevo planea construir (campo del front). Topa el pronóstico de
    # venta a 12/18/24 meses: no se puede vender más de lo que se construye. None → sin tope.
    unidades_proyecto: Optional[int] = None

# Modos de producto reconocidos. activo=True → pipeline implementado; False → placeholder.
PRODUCTO_MODES = {
    "vivienda_vertical":   {"label": "Vivienda vertical",   "activo": True,  "oferta_layer": "vv_venta"},
    "vivienda_horizontal": {"label": "Vivienda horizontal", "activo": True,  "oferta_layer": "vh_venta"},
    "lotes":               {"label": "Lotes",               "activo": False, "oferta_layer": None},
    "industrial":          {"label": "Industrial",          "activo": False, "oferta_layer": None},
    "logistica":           {"label": "Logística",           "activo": False, "oferta_layer": None},
    "oficinas":            {"label": "Oficinas",            "activo": False, "oferta_layer": None},
    "hotel":               {"label": "Hotel",               "activo": False, "oferta_layer": None},
}


# ════════════ ARQ-MODULAR · Caché de análisis + secciones independientes ════════════
# Regla de Héctor: cada sección debe poder ser LLAMADA Y MOSTRADA sola desde el front,
# sin depender de otra sección (usuarios con acceso a 1, 2 o todas las secciones y a
# zonas distintas). El pipeline pesado corre UNA vez (zona_procesar) y deja el análisis
# en caché; /api/zona/seccion sirve cualquier sección individual desde ahí. La caché es
# en memoria (mismo patrón que _CUENTAS); al llegar la DB (F-E) se persiste sin cambiar
# el contrato del endpoint.
ANALYSIS_CACHE: Dict[str, Dict[str, Any]] = {}
ANALYSIS_CACHE_MAX = int(os.environ.get("DATARIA_CACHE_MAX", "40"))


def _analysis_cache_put(key: Optional[str], payload: Dict[str, Any]) -> None:
    if not key:
        return
    ANALYSIS_CACHE[key] = {"payload": payload, "ts": _time.time()}
    while len(ANALYSIS_CACHE) > ANALYSIS_CACHE_MAX:
        oldest = min(ANALYSIS_CACHE, key=lambda k: ANALYSIS_CACHE[k]["ts"])
        ANALYSIS_CACHE.pop(oldest, None)


def _analysis_cache_get(key: Optional[str]) -> Optional[Dict[str, Any]]:
    entry = ANALYSIS_CACHE.get(key) if key else None
    return entry["payload"] if entry else None


class SeccionRequest(BaseModel):
    analisis_key: str          # el analisis_id_str que devolvió /api/zona/procesar
    seccion: str               # clave de SECTION_REGISTRY (resumen, mapa, demanda, ...)


class IsoCompararRequest(BaseModel):
    """ISO-MULTI · Comparación A/B de fuentes de isócrona para el MISMO pin/tiempo."""
    lat: float
    lng: float
    minutes: int = 8
    fuentes: Optional[List[str]] = None   # default: predik + valhalla (+ors/tomtom si hay key)


@app.post("/api/zona/isocrona_comparar")
async def isocrona_comparar(req: IsoCompararRequest):
    """Mide qué tanto se parecen las fuentes: área km², vértices, ms de respuesta e
    IoU contra la referencia (Predik si responde; si no, la primera que responda).
    IoU: 1.0 idénticas · ≥0.75 muy similares · <0.5 difieren de forma relevante."""
    fuentes = req.fuentes or (["predik", "valhalla"]
                              + (["ors"] if ORS_KEY else [])
                              + (["tomtom"] if TOMTOM_KEY else []))
    out: Dict[str, Any] = {}
    async with httpx.AsyncClient(timeout=60) as client:
        for f in fuentes:
            fn = ISO_PROVIDERS.get(f)
            if not fn:
                out[f] = {"ok": False, "error": "fuente desconocida"}
                continue
            t0 = _time.time()
            try:
                geo = await fn(client, req.lat, req.lng, req.minutes)
                ring = geo["coordinates"][0]
                out[f] = {"ok": True, "ms": int((_time.time() - t0) * 1000),
                          "n_vertices": len(ring), "area_km2": _area_ring_km2(ring),
                          "poligono": geo}
            except Exception as e:
                out[f] = {"ok": False, "ms": int((_time.time() - t0) * 1000),
                          "error": str(e)[:140]}
    ref = "predik" if out.get("predik", {}).get("ok") else \
        next((f for f in fuentes if out.get(f, {}).get("ok")), None)
    if ref:
        ring_ref = out[ref]["poligono"]["coordinates"][0]
        for f, d in out.items():
            if d.get("ok") and f != ref:
                d["iou_vs_referencia"] = _iou_rings(ring_ref, d["poligono"]["coordinates"][0])
                a_ref = out[ref].get("area_km2") or 0
                if a_ref and d.get("area_km2") is not None:
                    d["area_vs_referencia_pct"] = round(d["area_km2"] / a_ref * 100, 1)
    return {"ok": True, "minutes": req.minutes, "referencia": ref, "fuentes": out,
            "nota": ("IoU 1.0 = idénticas · ≥0.75 muy similares · <0.5 difieren de forma "
                     "relevante. La regla de negocio no cambia con la fuente; solo el "
                     "polígono de alcance físico.")}


class EstrellaFiltroRequest(BaseModel):
    """INV-4 · Corredor a la medida: el usuario define rangos manuales y el backend
    calcula el producto estrella + estadística robusta de ESE corredor."""
    analisis_key: str
    precio_min_M: Optional[float] = None
    precio_max_M: Optional[float] = None
    m2_min: Optional[float] = None
    m2_max: Optional[float] = None
    rec: Optional[int] = None


@app.post("/api/zona/estrella_filtro")
def estrella_filtro(req: EstrellaFiltroRequest):
    """INV-4 · Estrella y métricas robustas del corredor definido por el usuario,
    comparadas contra la estrella de la zona. Calcula SIEMPRE en backend sobre las
    tipologías del análisis en caché (el front solo captura rangos y muestra)."""
    payload = _analysis_cache_get(req.analisis_key)
    if not payload:
        return {"ok": False, "error": "Análisis no encontrado en caché · reprocesa la zona"}
    zd = payload.get("zone_data") or {}
    typ = zd.get("_typologies") or {}
    rows = []
    for proyecto, ts in typ.items():
        for t in ts:
            rows.append({
                "proyecto": proyecto,
                "precio": t.get("precio_ud"),
                "pm2": t.get("precio_m2"),
                "m2": t.get("area_total") or t.get("area_priv"),
                "rec": t.get("rec"),
                "vendidas": int(t.get("unid_vend") or 0),
                "disp": int(t.get("unid_disp") or 0),
                "abs": t.get("abs"),
            })
    def _pasa(t):
        if req.precio_min_M is not None and not (t["precio"] and t["precio"] / 1e6 >= req.precio_min_M):
            return False
        if req.precio_max_M is not None and not (t["precio"] and t["precio"] / 1e6 <= req.precio_max_M):
            return False
        if req.m2_min is not None and not (t["m2"] and t["m2"] >= req.m2_min):
            return False
        if req.m2_max is not None and not (t["m2"] and t["m2"] <= req.m2_max):
            return False
        if req.rec is not None and t.get("rec") != req.rec:
            return False
        return True
    corredor = [t for t in rows if _pasa(t)]
    est_zona = ((zd.get("top_estrella") or {}).get("zona") or [None])[0]
    return {
        "ok": True,
        "n_tipologias": len(corredor),
        "n_proyectos": len({t["proyecto"] for t in corredor}),
        "estrella_corredor": _estrella_de(corredor),
        "estrella_zona": est_zona,
        "stats": {
            "pm2": _stats_robustas([t["pm2"] for t in corredor]),
            "precio_M": _stats_robustas([t["precio"] / 1e6 for t in corredor if t.get("precio")]),
            "m2": _stats_robustas([t["m2"] for t in corredor]),
            "abs": _stats_robustas([t["abs"] for t in corredor if t.get("abs")]),
        },
        "nota_p3": "Los % de plusvalía por periodo se activan con la serie temporal (ticket P3).",
    }


@app.post("/api/zona/seccion")
def zona_seccion(req: SeccionRequest):
    """Sirve UNA sección de un análisis ya procesado (independiente del resto).
    Contrato estable para el modelo de permisos por sección/zona (la autenticación
    real se enchufa en F-E sin cambiar este contrato)."""
    cfg = SECTION_REGISTRY.get(req.seccion)
    if not cfg:
        return {"ok": False, "error": f"Sección desconocida: {req.seccion}",
                "secciones_validas": section_keys(True)}
    payload = _analysis_cache_get(req.analisis_key)
    if not payload:
        return {"ok": False, "error": "Análisis no encontrado en caché · reprocesa la zona",
                "analisis_key": req.analisis_key,
                "en_cache": sorted(ANALYSIS_CACHE.keys())[:10]}
    zd = payload.get("zone_data") or {}
    dd = payload.get("dim_data") or {}
    data: Dict[str, Any] = {}
    for k in cfg["payload_keys"]:
        if k in zd:
            data[k] = zd[k]
        elif k in dd:
            data[k] = dd[k]
        elif k in payload:
            data[k] = payload[k]
    # Contexto transversal mínimo (identidad ZA-8 SIEMPRE presente en encabezados)
    for k in ("name", "subtitle", "analisis_nombre", "analisis_version",
              "analisis_id_str", "analisis_fecha", "center"):
        data.setdefault(k, zd.get(k, payload.get(k)))
    data["producto"] = payload.get("producto")
    # _vars: fuente única que los templates consumen (TODO F-E: filtrar por access_tier)
    data["_vars"] = zd.get("_vars")
    return {"ok": True, "seccion": req.seccion, "label": cfg["label"],
            "access_tier": cfg["access_tier"], "analisis_key": req.analisis_key,
            "data": data}


# ──────────────────────── Clientes de servicios ────────────────────────
async def fetch_isochrone(client: httpx.AsyncClient, lat: float, lng: float, minutes: int) -> Dict:
    r = await client.post(f"{PRSP_BASE}/api/predik/isochrone", json={
        "latitude": lat, "longitude": lng, "minutes": minutes, "transport_type": "driving"
    })
    r.raise_for_status()
    geo = r.json()
    if geo.get("type") != "Polygon" or "coordinates" not in geo:
        raise HTTPException(502, f"Isócrona inválida para {minutes} min")
    return geo


# ════════════ ISO-MULTI · Isócronas multi-proveedor (Predik intacto + gratuitas) ════════════
# Petición de Héctor (Predik devolviendo 403): mantener la MISMA lógica de negocio y poder
# COMPARAR fuentes. Todos los adapters normalizan al MISMO contrato que Predik:
# {"type":"Polygon","coordinates":[[[lng,lat],...]]} — el resto del pipeline no cambia.
#   • predik   → el actual (default; la regla de negocio no se toca).
#   • valhalla → motor OSM en servidor público FOSSGIS (GRATUITO, SIN api key).
#   • ors      → OpenRouteService (gratuito con key · env DATARIA_ORS_KEY).
#   • tomtom   → Reachable Range (gratuito con key · env DATARIA_TOMTOM_KEY).
# Fallback automático: si Predik falla y DATARIA_ISO_FALLBACK está definido (default
# valhalla), la isócrona sale del fallback y el payload lo DECLARA (transparencia).
ISO_FUENTE_DEFAULT = os.environ.get("DATARIA_ISO_FUENTE", "predik")
ISO_FALLBACK = os.environ.get("DATARIA_ISO_FALLBACK", "valhalla")
VALHALLA_BASE = os.environ.get("DATARIA_VALHALLA", "https://valhalla1.openstreetmap.de")
ORS_BASE = os.environ.get("DATARIA_ORS", "https://api.openrouteservice.org")
ORS_KEY = os.environ.get("DATARIA_ORS_KEY", "")
TOMTOM_KEY = os.environ.get("DATARIA_TOMTOM_KEY", "")


def _ring_mayor_de_geom(geom: Dict) -> List[List[float]]:
    """Extrae el anillo exterior MAYOR de una geometría Polygon/MultiPolygon GeoJSON."""
    if not geom:
        return []
    t = geom.get("type")
    if t == "Polygon":
        anillos = [geom.get("coordinates", [[]])[0]]
    elif t == "MultiPolygon":
        anillos = [p[0] for p in geom.get("coordinates", []) if p]
    else:
        return []
    anillos = [a for a in anillos if a and len(a) >= 4]
    return max(anillos, key=len) if anillos else []


def _poly_norm(ring: List[List[float]]) -> Dict:
    """Normaliza al contrato Predik. Anillo corto → error (integridad, no inventar)."""
    if not ring or len(ring) < 4:
        raise ValueError("isócrona sin polígono utilizable")
    if ring[0] != ring[-1]:
        ring = ring + [ring[0]]
    return {"type": "Polygon", "coordinates": [[[float(x), float(y)] for x, y in ring]]}


def _parse_valhalla_iso(j: Dict, minutes: int) -> Dict:
    feats = (j or {}).get("features") or []
    cand = [f for f in feats
            if (f.get("properties") or {}).get("contour") in (minutes, float(minutes))] or feats
    if not cand:
        raise ValueError("Valhalla sin features")
    return _poly_norm(_ring_mayor_de_geom(cand[0].get("geometry") or {}))


def _parse_ors_iso(j: Dict) -> Dict:
    feats = (j or {}).get("features") or []
    if not feats:
        raise ValueError("ORS sin features")
    return _poly_norm(_ring_mayor_de_geom(feats[0].get("geometry") or {}))


def _parse_tomtom_iso(j: Dict) -> Dict:
    b = ((j or {}).get("reachableRange") or {}).get("boundary") or []
    ring = [[p.get("longitude"), p.get("latitude")] for p in b
            if p.get("longitude") is not None]
    return _poly_norm(ring)


async def fetch_isochrone_valhalla(client: httpx.AsyncClient, lat: float, lng: float, minutes: int) -> Dict:
    # POST con JSON COMPACTO: el GET con json en query codificaba espacios como '+' y
    # Valhalla respondía 400 (diagnosticado y verificado en producción · 8 jul 2026).
    r = await client.post(f"{VALHALLA_BASE}/isochrone",
                          json={"locations": [{"lat": lat, "lon": lng}], "costing": "auto",
                                "contours": [{"time": minutes}], "polygons": True,
                                "denoise": 0.3},
                          headers={"User-Agent": "Dataria/2.0 (contacto@prosperia.mx)"})
    r.raise_for_status()
    return _parse_valhalla_iso(r.json(), minutes)


async def fetch_isochrone_ors(client: httpx.AsyncClient, lat: float, lng: float, minutes: int) -> Dict:
    if not ORS_KEY:
        raise RuntimeError("ORS sin api key (env DATARIA_ORS_KEY)")
    r = await client.post(f"{ORS_BASE}/v2/isochrones/driving-car",
                          headers={"Authorization": ORS_KEY},
                          json={"locations": [[lng, lat]], "range": [minutes * 60],
                                "range_type": "time"})
    r.raise_for_status()
    return _parse_ors_iso(r.json())


async def fetch_isochrone_tomtom(client: httpx.AsyncClient, lat: float, lng: float, minutes: int) -> Dict:
    if not TOMTOM_KEY:
        raise RuntimeError("TomTom sin api key (env DATARIA_TOMTOM_KEY)")
    r = await client.get(
        f"https://api.tomtom.com/routing/1/calculateReachableRange/{lat},{lng}/json",
        params={"timeBudgetInSec": minutes * 60, "key": TOMTOM_KEY, "travelMode": "car"})
    r.raise_for_status()
    return _parse_tomtom_iso(r.json())


ISO_PROVIDERS = {
    "predik": fetch_isochrone,
    "valhalla": fetch_isochrone_valhalla,
    "ors": fetch_isochrone_ors,
    "tomtom": fetch_isochrone_tomtom,
}


# DIRECTIVA DE HÉCTOR (8 jul 2026): la fuente NO la elige el usuario — es un proceso del
# BACKEND que usa la que esté funcionando, en orden de calidad. La invariancia del
# resultado ante la fuente quedó DEMOSTRADA (VERIF-FUENTES: payload y HTML idénticos a
# geometría igual). Cadena default: valhalla (mejor calidad gratuita disponible) → predik
# (cuando restauren su credencial) → ors/tomtom si hay keys. Configurable sin código.
ISO_CADENA = [f.strip().lower() for f in os.environ.get(
    "DATARIA_ISO_CADENA", "valhalla,predik").split(",") if f.strip()]


async def fetch_isochrone_fuente(client: httpx.AsyncClient, lat: float, lng: float,
                                 minutes: int, fuente: Optional[str] = None):
    """(geojson, fuente_usada, nota). Sin fuente explícita → CADENA del backend en orden
    de calidad: primera que responda gana, siempre declarada. Con fuente explícita
    (comparador/pruebas API) → esa, con fallback a la cadena si falla."""
    intentos: List[str] = []
    if fuente:
        f = str(fuente).lower()
        if f in ISO_PROVIDERS:
            intentos.append(f)
    for f in ISO_CADENA:
        if f in ISO_PROVIDERS and f not in intentos:
            intentos.append(f)
    for f in ("ors", "tomtom"):
        if f not in intentos and ((f == "ors" and ORS_KEY) or (f == "tomtom" and TOMTOM_KEY)):
            intentos.append(f)
    if not intentos:
        intentos = ["predik"]
    errores = []
    for i, f in enumerate(intentos):
        # ANTI-502: 2 intentos por fuente con pausa breve — cubre fallos TRANSITORIOS
        # (Valhalla público con carga puntual, Predik/PRSP despertando en Render). El 502
        # solo puede ocurrir si TODAS las fuentes fallan DOS veces cada una.
        for intento in range(2):
            try:
                geo = await ISO_PROVIDERS[f](client, lat, lng, minutes)
                nota = None
                if i > 0:
                    nota = (f"{' · '.join(errores)[:120]} · isócrona generada por {f}")
                return geo, f, nota
            except Exception as e:
                if intento == 0 and not isinstance(e, RuntimeError):
                    await _aio.sleep(0.8)   # RuntimeError = sin api key: reintentar no ayuda
                    continue
                errores.append(f"{f}: {str(e)[:60]}")
                break
    raise RuntimeError(" · ".join(errores)[:220])


# ════════════ MOVILIDAD · reemplazo de las capas de tráfico de Predik ════════════
# Directiva de Héctor (8 jul 2026): de Predik se esperaban origen-destino, isócronas y
# tráfico VEHICULAR y PEATONAL. OD e isócronas ya están reemplazados (censo INEGI + cadena
# multi-proveedor). Aquí las dos capas de tráfico, con integridad declarada:
#   • PEAT-1 · Alcance peatonal: isócrona pedestrian de Valhalla (red vial OSM REAL, sin
#     key · POST verificado en vivo 8 jul 2026: 200 OK, Polygon) + masa demográfica REAL
#     (AGEBs del KMZ del DI cuyo centroide cae dentro). El FLUJO peatonal observado no
#     existe en fuente gratuita (core de pago de Placer/Predik) → "proximamente"
#     declarado; se modelará con atractores DENUE y se calibrará si Predik regresa.
#   • TRAF-1 · Tráfico vehicular: TomTom Traffic Flow (tier gratuito) en el pin + 4
#     puntos cardinales (~1 km); MEDIANAS robustas de velocidad actual vs flujo libre.
#     Sin DATARIA_TOMTOM_KEY → "proximamente" (nunca inventar).
PEAT_MINUTOS = int(os.environ.get("DATARIA_PEAT_MIN", "10"))
_TRAF_OFFSET_KM = 1.0


async def fetch_isochrone_peatonal(client: httpx.AsyncClient, lat: float, lng: float,
                                   minutes: int = PEAT_MINUTOS) -> Dict:
    r = await client.post(f"{VALHALLA_BASE}/isochrone",
                          json={"locations": [{"lat": lat, "lon": lng}],
                                "costing": "pedestrian",
                                "contours": [{"time": minutes}], "polygons": True,
                                "denoise": 0.3},
                          headers={"User-Agent": "Dataria/2.0 (contacto@prosperia.mx)"})
    r.raise_for_status()
    return _parse_valhalla_iso(r.json(), minutes)


async def _tomtom_flow_punto(client: httpx.AsyncClient, lat: float, lng: float) -> Dict:
    r = await client.get(
        "https://api.tomtom.com/traffic/services/4/flowSegmentData/absolute/10/json",
        params={"point": f"{lat},{lng}", "key": TOMTOM_KEY, "unit": "KMPH"})
    r.raise_for_status()
    d = (r.json() or {}).get("flowSegmentData") or {}
    cs, ff = d.get("currentSpeed"), d.get("freeFlowSpeed")
    if cs is None or not ff:
        raise ValueError("flujo sin velocidades")
    return {"actual": float(cs), "libre": float(ff),
            "confianza": d.get("confidence"), "frc": d.get("frc")}


async def derive_movilidad(client: httpx.AsyncClient, lat: float, lng: float,
                           agebs_geo: Optional[List[Dict]]) -> Dict[str, Any]:
    """Capas de movilidad del payload. TODO cálculo aquí (regla 2); el front solo pinta.
    Cada capa declara estado: ok · proximamente (falta ruta de dato) · error."""
    out: Dict[str, Any] = {"peatonal": None, "vehicular": None}
    # ── PEAT-1 · alcance peatonal + masa real dentro ──
    try:
        geo = await fetch_isochrone_peatonal(client, lat, lng)
        ring = geo["coordinates"][0]
        peat: Dict[str, Any] = {
            "estado": "ok", "minutos": PEAT_MINUTOS, "fuente": "valhalla_pedestrian_osm",
            "area_km2": _area_ring_km2(ring), "poligono": ring,
            "hogares": None, "poblacion": None, "n_agebs": 0, "masa_fuente": None,
            "flujo_peatonal": {"estado": "proximamente",
                               "nota": ("Flujo peatonal observado no disponible en fuente "
                                        "gratuita; se modelará con atractores DENUE "
                                        "(token pendiente) y ancla censal.")},
        }
        dentro = [g for g in (agebs_geo or [])
                  if g.get("lng") is not None and g.get("lat") is not None
                  and _point_in_ring(g["lng"], g["lat"], ring)]
        peat["n_agebs"] = len(dentro)
        if dentro:
            hogs = [_attr_num(g.get("attrs"), _RE_MASA_HOG) for g in dentro]
            pobs = [_attr_num(g.get("attrs"), _RE_MASA_POB, excl=_RE_EXTRANJERO)
                    for g in dentro]
            hogs = [h for h in hogs if h]
            pobs = [p for p in pobs if p]
            # Integridad: masa solo si la MAYORÍA de AGEBs dentro la publica; si no, N/D
            if len(hogs) >= len(dentro) * 0.5:
                peat["hogares"] = round(sum(hogs))
                peat["masa_fuente"] = "hogares_kmz"
            if len(pobs) >= len(dentro) * 0.5:
                peat["poblacion"] = round(sum(pobs))
                peat["masa_fuente"] = peat["masa_fuente"] or "poblacion_kmz"
        out["peatonal"] = peat
    except Exception as e:
        out["peatonal"] = {"estado": "error", "detalle": str(e)[:120]}
    # ── TRAF-1 · congestión vehicular (key-gated · nunca inventar) ──
    if not TOMTOM_KEY:
        out["vehicular"] = {"estado": "proximamente",
                            "nota": ("Tráfico vehicular en tiempo real vía TomTom Traffic "
                                     "(tier gratuito). Se activa al definir "
                                     "DATARIA_TOMTOM_KEY en Render.")}
        return out
    try:
        dlat = _TRAF_OFFSET_KM / 111.32
        dlng = _TRAF_OFFSET_KM / (111.32 * math.cos(math.radians(lat)))
        puntos = [(lat, lng), (lat + dlat, lng), (lat - dlat, lng),
                  (lat, lng + dlng), (lat, lng - dlng)]
        flujos = []
        for pla, pln in puntos:
            try:
                flujos.append(await _tomtom_flow_punto(client, pla, pln))
            except Exception:
                continue   # punto sin vialidad medida cerca: se omite, no se inventa
        if not flujos:
            out["vehicular"] = {"estado": "error",
                                "detalle": "TomTom sin segmentos medidos en la zona"}
            return out
        ratios = [f["actual"] / f["libre"] for f in flujos if f["libre"]]
        out["vehicular"] = {
            "estado": "ok", "fuente": "tomtom_flow", "n_puntos": len(flujos),
            "velocidad_kmh": round(statistics.median(f["actual"] for f in flujos), 1),
            "velocidad_libre_kmh": round(statistics.median(f["libre"] for f in flujos), 1),
            # 1.0 = circulación libre · 0.5 = al 50% de la velocidad de flujo libre
            "indice_fluidez": round(statistics.median(ratios), 2) if ratios else None,
            "nota": ("Medianas robustas sobre pin + 4 puntos cardinales (~1 km). "
                     "Instantánea al momento del análisis, no promedio histórico."),
        }
    except Exception as e:
        out["vehicular"] = {"estado": "error", "detalle": str(e)[:120]}
    return out


def _area_ring_km2(ring: List[List[float]]) -> Optional[float]:
    """Área del anillo (km²) por shoelace equirectangular (dato geométrico)."""
    if not ring or len(ring) < 4:
        return None
    lat0 = math.radians(sum(p[1] for p in ring) / len(ring))
    km = [((p[0] - ring[0][0]) * 111.32 * math.cos(lat0),
           (p[1] - ring[0][1]) * 111.32) for p in ring]
    s = 0.0
    for i in range(len(km) - 1):
        s += km[i][0] * km[i + 1][1] - km[i + 1][0] * km[i][1]
    return round(abs(s) / 2.0, 2)


def _iou_rings(a: List[List[float]], b: List[List[float]], n: int = 46) -> Optional[float]:
    """Intersección/Unión (IoU) de dos anillos por muestreo de rejilla (sin dependencias).
    1.0 = idénticas · ≥0.75 muy similares · <0.5 difieren de forma relevante."""
    if not a or not b or len(a) < 4 or len(b) < 4:
        return None
    xs = [p[0] for p in a] + [p[0] for p in b]
    ys = [p[1] for p in a] + [p[1] for p in b]
    x0, x1, y0, y1 = min(xs), max(xs), min(ys), max(ys)
    if x1 <= x0 or y1 <= y0:
        return None
    inter = uni = 0
    for i in range(n):
        for j in range(n):
            x = x0 + (x1 - x0) * (i + 0.5) / n
            y = y0 + (y1 - y0) * (j + 0.5) / n
            ia = _point_in_ring(x, y, a)
            ib = _point_in_ring(x, y, b)
            if ia and ib:
                inter += 1
            if ia or ib:
                uni += 1
    return round(inter / uni, 3) if uni else None


async def fetch_vvv(client: httpx.AsyncClient, ring: List[List[float]], tipo: str = "vv_venta") -> Dict:
    body = {
        "tipo": tipo,
        "geometry": {"rings": [ring], "spatialReference": {"wkid": 4326}},
        "filters": {},
        "returnGeometry": True,
        "resultRecordCount": 500,
    }
    r = await client.post(f"{PRSP_BASE}/api/v1/vvv/query", json=body)
    r.raise_for_status()
    return r.json()


async def fetch_descarga_di(client: httpx.AsyncClient, ring: List[List[float]], active_map: str = "NSE") -> Optional[bytes]:
    body = {
        "activeMap": active_map,
        "cliente": "Dataria",
        "proyecto": "ZonaAnalisis",
        "geometry": {"rings": [ring], "spatialReference": {"wkid": 4326}},
    }
    r = await client.post(f"{PRSP_BASE}/api/descargas/di/export", json=body)
    if r.status_code != 200:
        # 413 (demasiados registros) u otro → devolvemos None; demografía quedará N/D
        return None
    if not r.content.startswith(b"PK"):
        return None
    return r.content


def parse_di_geometria(zip_bytes: bytes) -> List[Dict[str, Any]]:
    """Extrae la GEORREFERENCIACIÓN de cada AGEB del KMZ que viene en la respuesta de
    DescargaDI (activeMap='NSE'). Devuelve, por AGEB: centroide (lng,lat), su NSE
    ordinal (1=A … creciente hacia NSE bajo) y su NSE textual.

    Fuente: el mismo ZIP que ya se consulta para la demografía (no hay red adicional).
    El KMZ trae un Placemark por AGEB con id 'NSE_<ordinal>_<idx>', styleUrl '#NSE_<ordinal>'
    y la geometría poligonal real. El XLSX (atributos) y el KMZ (geometría) NO vienen en
    el mismo orden y el KMZ no trae CVEGEO, por eso la geometría se usa como capa propia
    (geometría + NSE), no se fusiona fila a fila con el XLSX. Para la barrera de mercado
    socioeconómica esto es suficiente: lo que importa es dónde está cada nivel de NSE.

    Integridad: si no hay KMZ o no se puede parsear, devuelve [] (sin inventar posiciones).
    """
    try:
        z = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except Exception:
        return []
    kmz_name = next((n for n in z.namelist() if n.endswith(".kmz")), None)
    if not kmz_name:
        return []
    try:
        kml = zipfile.ZipFile(io.BytesIO(z.read(kmz_name))).read("doc.kml").decode("utf-8", errors="ignore")
    except Exception:
        return []

    out = []
    for m in re.finditer(r'<Placemark id="([^"]+)">(.*?)</Placemark>', kml, re.S):
        pid, body = m.group(1), m.group(2)
        mo = re.match(r'NSE_(\d+)_', pid)
        nse_ord = int(mo.group(1)) if mo else None
        mt = re.search(r'ingreso</td>\s*<td>([^<]+)</td>', body)
        nse_txt = mt.group(1).strip() if mt else None
        # Bloques de coordenadas POR SEPARADO (un placemark puede traer multipolígono;
        # mezclar bloques produciría un anillo cruzado inválido).
        bloques = []
        xs, ys = [], []
        for co in re.findall(r"<coordinates>(.*?)</coordinates>", body, re.S):
            blq = []
            for tok in co.strip().split():
                p = tok.split(",")
                if len(p) >= 2:
                    try:
                        bx, by = float(p[0]), float(p[1])
                        blq.append((bx, by))
                        xs.append(bx); ys.append(by)
                    except ValueError:
                        pass
            if len(blq) >= 4:
                bloques.append(blq)
        if not xs:
            continue
        cent_lng = sum(xs) / len(xs)
        cent_lat = sum(ys) / len(ys)
        # F-B/MAPA-1 · ANILLO del polígono (bloque mayor, decimado a ≤60 vértices) para
        # pintar la capa coroplética de NSE en el mapa. Aditivo: los consumidores previos
        # (proximidad, barrera) siguen usando solo nse_rank/nse_txt/lng/lat.
        ring = []
        area_km2 = None
        if bloques:
            mayor = max(bloques, key=len)
            paso = max(1, math.ceil(len(mayor) / 60))
            ring = [[round(x, 5), round(y, 5)] for x, y in mayor[::paso]]
            if ring and ring[0] != ring[-1]:
                ring.append(ring[0])
            # Área aproximada en km² (shoelace equirectangular · dato geométrico)
            lat0 = math.radians(cent_lat)
            km = [((x - cent_lng) * 111.32 * math.cos(lat0), (y - cent_lat) * 111.32)
                  for x, y in mayor]
            s = 0.0
            for i in range(len(km) - 1):
                s += km[i][0] * km[i + 1][1] - km[i + 1][0] * km[i][1]
            area_km2 = round(abs(s) / 2.0, 4)
        # Atributos GENÉRICOS de la descripción del placemark (pares <td>etiqueta|valor</td>):
        # lo que la base publique por AGEB (población, hogares, etc.) queda disponible para
        # capas futuras (densidad, flotante) SIN reescribir el parser. Cap a 25 pares.
        attrs = {}
        for ma in re.finditer(r'<td>([^<]{1,60})</td>\s*<td>([^<]{0,80})</td>', body):
            if len(attrs) >= 25:
                break
            attrs[ma.group(1).strip()] = ma.group(2).strip()
        out.append({
            "nse_rank": nse_ord,        # ordinal del NSE (señal para el clustering)
            "nse_txt": nse_txt,         # NSE textual (A..E / Industrial)
            "lng": round(cent_lng, 6),
            "lat": round(cent_lat, 6),
            "n_vertices": len(xs),
            "ring": ring,               # anillo [[lng,lat],...] para capa del mapa (F-B)
            "area_km2": area_km2,       # área del AGEB (para densidad cuando haya campos)
            "attrs": attrs,             # atributos publicados por la base (genérico)
        })
    return out


def parse_di_xlsx(zip_bytes: bytes) -> List[Dict[str, Any]]:
    """Extrae los AGEBs del ZIP de DescargaDI como lista de dicts {header: value}."""
    z = zipfile.ZipFile(io.BytesIO(zip_bytes))
    xlsx_name = next((n for n in z.namelist() if n.endswith(".xlsx")), None)
    if not xlsx_name:
        return []
    wb = openpyxl.load_workbook(io.BytesIO(z.read(xlsx_name)), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        rec = {}
        empty = True
        for ci, h in enumerate(headers, 1):
            if h is None:
                continue
            v = ws.cell(r, ci).value
            rec[str(h).strip()] = v
            if v not in (None, ""):
                empty = False
        if not empty:
            rows.append(rec)
    return rows


# ──────────────────────── Endpoint principal ────────────────────────
@app.get("/")
async def root():
    """Raíz informativa del API. El tablero se sirve en /tablero."""
    return {
        "ok": True,
        "service": "dataria-orchestrator",
        "version": "2.0",
        "tablero": "/tablero",
        "endpoints": ["/health", "/api/zona/poligono", "/api/zona/procesar",
                      "/api/zona/analyze", "/api/zona/procesar_auth",
                      "/api/secciones", "/api/cuentas"],
    }


@app.get("/tablero")
async def tablero():
    """Sirve el tablero (dashboard_zona_analisis.html) desde static/.

    Integridad: si el archivo no existe, devolver 404 explícito en vez de
    reventar el servidor. Así el despliegue del API no depende de que el
    estático esté presente.
    """
    if not TABLERO_FILE.is_file():
        return JSONResponse(
            status_code=404,
            content={
                "ok": False,
                "error": "Tablero no encontrado",
                "detalle": "Falta static/dashboard_zona_analisis.html en el despliegue.",
                "ruta_esperada": str(TABLERO_FILE),
            },
        )
    # no-cache para que cada despliegue del HTML se sirva fresco
    return FileResponse(
        str(TABLERO_FILE),
        media_type="text/html; charset=utf-8",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


@app.get("/health")
async def health():
    return {"ok": True, "service": "dataria-orchestrator", "version": "2.0", "stages": ["poligono", "procesar", "analyze"]}


@app.post("/api/zona/analyze")
async def analyze(req: ZonaRequest):
    profile = isochrone_profile(req.predio_m2, req.uso_comercial)

    async with httpx.AsyncClient(timeout=HTTP_TIMEOUT) as client:
        # 1 · Isócronas (todas las del perfil) · ISO-MULTI con fuente/fallback declarados
        isocronas: Dict[int, Dict] = {}
        iso_fuente_usada, iso_nota = None, None
        try:
            for m in profile["minutos"]:
                geo, iso_fuente_usada, nota = await fetch_isochrone_fuente(
                    client, req.lat, req.lng, m, iso_fuente_usada or req.iso_fuente)
                iso_nota = iso_nota or nota
                isocronas[m] = geo
        except Exception as e:
            raise HTTPException(502, f"Isócronas no disponibles (todas las fuentes): {e}")

        # Anillo base (azul · 8 min) = zona PRIMARIA. Anillo mayor (verde · principal)
        # = zona SECUNDARIA. La oferta y la demanda se traen del anillo MAYOR (azul+verde
        # combinadas), y luego se clasifica cada proyecto en su set (directo/primario/secundario).
        base_min = 8 if 8 in isocronas else profile["principal"]
        base_ring = isocronas[base_min]["coordinates"][0]
        outer_min = profile["principal"]                       # mayor isócrona del perfil (verde)
        outer_ring = isocronas[outer_min]["coordinates"][0] if outer_min in isocronas else base_ring

        # 2 · Oferta (VVV) venta + renta · universo = anillo MAYOR (azul+verde)
        try:
            vvv_venta = await fetch_vvv(client, outer_ring, "vv_venta")
            vvv_renta = await fetch_vvv(client, outer_ring, "vv_renta")
        except httpx.HTTPError as e:
            raise HTTPException(502, f"VVV no disponible: {e}")

        # 3 · Demografía (DescargaDI · NSE) · universo = anillo MAYOR (demanda azul+verde)
        di_bytes = await fetch_descarga_di(client, outer_ring, "NSE")

    agebs = parse_di_xlsx(di_bytes) if di_bytes else []
    agebs_geo = parse_di_geometria(di_bytes) if di_bytes else []   # geometría por AGEB (proximidad)

    # ── Derivación VVV: proyectos resumen + tipologías ft ──
    resumen = vvv_venta.get("datasets", {}).get("resumen", [])
    ft = vvv_venta.get("datasets", {}).get("ft", [])
    pagos = vvv_venta.get("datasets", {}).get("pagos", [])
    resumen_renta = vvv_renta.get("datasets", {}).get("resumen", [])
    ft_renta = vvv_renta.get("datasets", {}).get("ft", [])

    # Percepción de valor + ajuste de zona de influencia (detección de mercados)
    perception = value_perception_adjust(resumen, base_ring, req.lng, req.lat)

    # ── Derivación DIGO/DPO ──
    demografia = derive_demografia(agebs, ft, agebs_geo, req.lng, req.lat)         # population, households, tca, nse, tenencia...
    nse_dim = derive_nse_dim(agebs)                   # nse_dim[] para DIM_DATA
    segments = derive_segments(agebs, ft, "vertical", agebs_geo, req.lng, req.lat, ft_renta)   # buckets de demanda
    productos = derive_productos_venta(ft, segments, req.unidades_proyecto, demografia.get("personas_hogar"), _build_di_detail(agebs))  # productos venta
    productos_renta = derive_productos_renta(ft_renta, segments)  # productos renta
    comercio = derive_comercio(agebs, _nse_dominante_agebs(agebs, agebs_geo, req.lng, req.lat)[1], ft)  # potencial comercio/retail (NSE dinámico)

    # ── Ensamblaje del payload (esquema que consume el template) ──
    payload = assemble_zone_payload(
        req=req, profile=profile, isocronas=isocronas,
        resumen=resumen, ft=ft, pagos=pagos,
        resumen_renta=resumen_renta,
        demografia=demografia, nse_dim=nse_dim, segments=segments,
        productos=productos, productos_renta=productos_renta,
        comercio=comercio, perception=perception, agebs_count=len(agebs),
        agebs=agebs, ft_renta=ft_renta,
    )
    return payload


# ════════════════════════════════════════════════════════════════
# FLUJO EN DOS ETAPAS
#   Etapa 1: /api/zona/poligono  → SOLO Predik. Rápido. Devuelve la
#            isócrona para que el tablero pinte el polígono de inmediato.
#   Etapa 2: /api/zona/procesar  → VVV + DescargaDI + derivaciones.
#            Llena el resto de secciones. Tolerante a fallos por bloque.
# ════════════════════════════════════════════════════════════════

@app.post("/api/zona/poligono")
async def zona_poligono(req: ZonaRequest):
    """ETAPA 1 (rápida): devuelve solo las isócronas. No depende de VVV ni DescargaDI."""
    profile = isochrone_profile(req.predio_m2, req.uso_comercial)
    async with httpx.AsyncClient(timeout=HTTP_TIMEOUT) as client:
        isocronas: Dict[int, Dict] = {}
        iso_fuente_usada, iso_nota = None, None
        try:
            for m in profile["minutos"]:
                geo, iso_fuente_usada, nota = await fetch_isochrone_fuente(
                    client, req.lat, req.lng, m, iso_fuente_usada or req.iso_fuente)
                iso_nota = iso_nota or nota
                isocronas[m] = geo
        except Exception as e:
            raise HTTPException(502, f"Isócronas no disponibles (todas las fuentes): {e}")
    return {
        "ok": True,
        "iso_fuente": iso_fuente_usada,
        "iso_nota": iso_nota,
        "stage": "poligono",
        "perfil_iso": profile["key"],
        "perfil_label": profile["label"],
        "minutos": profile["minutos"],
        "principal": profile["principal"],
        "isocronas": {str(m): geo for m, geo in isocronas.items()},
        "center": [req.lat, req.lng],
    }


@app.post("/api/zona/procesar")
async def zona_procesar(req: ZonaRequest):
    """
    ETAPA 2 (pesada): toma el mismo input, recalcula la isócrona base y
    procesa VVV + DescargaDI + derivaciones. Devuelve el payload completo.
    Cada bloque está aislado: si uno falla, los demás se entregan igual y
    se reporta el error en `errors` sin tumbar toda la respuesta.

    RUTEO POR MODO DE PRODUCTO (req.producto):
    El pipeline demográfico (isócronas, DI, percepción, competidores, segmentos,
    comercio) es COMÚN a todos los modos y se ejecuta una sola vez. Solo cambian
    dos puntos según el modo: (1) la capa de oferta que se consulta, (2) la función
    que diseña el producto. Esto comparte memoria entre modos: no se cargan dos
    pipelines, solo se enruta la oferta y el diseño de producto.
    """
    modo = req.producto if req.producto in PRODUCTO_MODES else "vivienda_vertical"
    modo_cfg = PRODUCTO_MODES[modo]
    # Modos declarados pero aún no implementados → respuesta explícita (no se procesa)
    if not modo_cfg["activo"]:
        return {
            "ok": False, "stage": "procesar",
            "producto": modo, "producto_label": modo_cfg["label"],
            "no_disponible": True,
            "mensaje": f"El análisis de {modo_cfg['label']} aún no está disponible. "
                       f"Modos activos: vivienda vertical y vivienda horizontal.",
            "modos_activos": [k for k, v in PRODUCTO_MODES.items() if v["activo"]],
            "errors": {},
        }
    oferta_layer = modo_cfg["oferta_layer"]   # "vv_venta" o "vh_venta"

    profile = isochrone_profile(req.predio_m2, req.uso_comercial)
    errors = {}

    iso_fuente_usada, iso_nota = None, None
    async with httpx.AsyncClient(timeout=HTTP_TIMEOUT) as client:
        # Isócrona base (8 min o principal) — necesaria para acotar VVV/DI · ISO-MULTI
        try:
            base_min = 8 if 8 in profile["minutos"] else profile["principal"]
            base_iso, iso_fuente_usada, iso_nota = await fetch_isochrone_fuente(
                client, req.lat, req.lng, base_min, req.iso_fuente)
            isocronas = {base_min: base_iso}
            # Las demás isócronas del perfil (para pintar anillos), tolerante a fallo;
            # misma fuente que la base (consistencia entre anillos).
            for m in profile["minutos"]:
                if m not in isocronas:
                    try:
                        geo, _, _ = await fetch_isochrone_fuente(
                            client, req.lat, req.lng, m, iso_fuente_usada)
                        isocronas[m] = geo
                    except Exception:
                        pass
        except Exception as e:
            raise HTTPException(502, f"Isócronas no disponibles (todas las fuentes): {e}")
        base_ring = isocronas[base_min]["coordinates"][0]
        # Anillo MAYOR (verde · zona secundaria). Oferta y demanda se traen de azul+verde
        # combinadas; luego se clasifica cada proyecto en su set (directo/primario/secundario).
        outer_min = profile["principal"]
        outer_ring = isocronas[outer_min]["coordinates"][0] if outer_min in isocronas else base_ring

        # VVV (oferta) — universo = anillo MAYOR. La capa depende del modo:
        # vivienda_vertical → vv_venta (+vv_renta); vivienda_horizontal → vh_venta (sin renta).
        vvv_venta, vvv_renta = {}, {}
        try:
            vvv_venta = await fetch_vvv(client, outer_ring, oferta_layer)
        except Exception as e:
            errors["vvv_venta"] = str(e)
        # Renta solo aplica a vivienda vertical (no hay capa de renta horizontal)
        if modo == "vivienda_vertical":
            try:
                vvv_renta = await fetch_vvv(client, outer_ring, "vv_renta")
            except Exception as e:
                errors["vvv_renta"] = str(e)

        # DescargaDI (demografía) — universo = anillo MAYOR (demanda azul+verde)
        di_bytes = None
        try:
            di_bytes = await fetch_descarga_di(client, outer_ring, "NSE")
        except Exception as e:
            errors["descarga_di"] = str(e)

        # CAPTABLE (opt-in) · anillo amplio + DI de la corona para el modelo Huff v1
        captable_raw = None
        if req.incluir_captable:
            try:
                capt_min = req.captable_min or min(CAPTABLE_MIN_TOPE,
                                                   (profile.get("principal") or 8) + 10)
                capt_geo, capt_fuente, _cn = await fetch_isochrone_fuente(
                    client, req.lat, req.lng, capt_min, iso_fuente_usada)
                di_capt = await fetch_descarga_di(client, capt_geo["coordinates"][0], "NSE")
                captable_raw = {"ring": capt_geo["coordinates"][0], "min": capt_min,
                                "fuente": capt_fuente, "di": di_capt}
            except Exception as e:
                errors["mercado_captable"] = str(e)[:120]

    # FILTRO ESPACIAL: garantizar que solo entren proyectos/tipologías DENTRO del anillo mayor
    vvv_venta = filter_vvv_by_polygon(vvv_venta, outer_ring)
    vvv_renta = filter_vvv_by_polygon(vvv_renta, outer_ring)
    if vvv_venta.get("_spatial_filter"):
        sf = vvv_venta["_spatial_filter"]
        if sf["resumen_total"] > sf["resumen_in"] or sf["ft_total"] > sf["ft_in"]:
            errors["filtro_espacial"] = (f"VVV devolvió fuera de zona: "
                f"resumen {sf['resumen_total']}→{sf['resumen_in']}, ft {sf['ft_total']}→{sf['ft_in']}")

    agebs = parse_di_xlsx(di_bytes) if di_bytes else []
    # Georreferenciación de AGEBs (centroide + NSE) desde el KMZ del MISMO ZIP (sin red extra).
    # Capa propia para detectar la barrera de mercado socioeconómica (NSE/posición).
    try:
        agebs_geo = parse_di_geometria(di_bytes) if di_bytes else []
    except Exception as e:
        agebs_geo = []
        errors["agebs_geo"] = str(e)
    resumen = vvv_venta.get("datasets", {}).get("resumen", [])
    ft = vvv_venta.get("datasets", {}).get("ft", [])
    pagos = vvv_venta.get("datasets", {}).get("pagos", [])
    resumen_renta = vvv_renta.get("datasets", {}).get("resumen", [])
    ft_renta = vvv_renta.get("datasets", {}).get("ft", [])

    # Cada derivación aislada
    def _safe(fn, label, default):
        try:
            return fn()
        except Exception as e:
            errors[label] = str(e)
            return default

    perception = _safe(lambda: value_perception_adjust(resumen, base_ring, req.lng, req.lat, agebs_geo), "perception",
                       {"n_total": 0, "n_zona": 0, "media": None, "sd": None, "cv": None,
                        "barrera": False, "metodo": "isocrona", "cobertura_pct": 0,
                        "motivo": "No disponible", "proyectos": [], "cluster_names": None})
    # ZONA-MUESTRA · SOPORTE MÍNIMO (causa raíz del defecto reportado el 8 jul 2026): la
    # detección de mercados recorta una isócrona GENEROSA al mercado del pin, pero no puede
    # crecer una isócrona corta — con muestra chica el clúster degenera y la zona sale mal
    # (ZMM con Valhalla: 11 proyectos → hull de 7 vértices). Si el anillo base trae menos de
    # ZONA_MUESTRA_MIN proyectos y el perfil tiene anillo mayor, la percepción se evalúa
    # sobre el anillo MAYOR y se DECLARA la ampliación. Con fuentes generosas (Predik en las
    # anclas: ~40 en 8 min) esta regla NO se activa: el comportamiento validado no cambia.
    try:
        if (perception.get("n_zona") or 0) < ZONA_MUESTRA_MIN and outer_min != base_min:
            _per2 = value_perception_adjust(resumen, outer_ring, req.lng, req.lat, agebs_geo)
            if (_per2.get("n_zona") or 0) > (perception.get("n_zona") or 0):
                _per2["zona_base_ampliada"] = {
                    "de_min": base_min, "a_min": outer_min,
                    "n_antes": perception.get("n_zona") or 0,
                    "n_despues": _per2.get("n_zona"),
                    "umbral": ZONA_MUESTRA_MIN,
                }
                _per2["motivo"] = (((_per2.get("motivo") or "") +
                    f" · anillo base ampliado {base_min}→{outer_min} min por muestra "
                    f"insuficiente (n={perception.get('n_zona') or 0} < {ZONA_MUESTRA_MIN})")
                    .strip(" ·"))
                perception = _per2
    except Exception as e:
        errors["zona_muestra"] = str(e)[:120]

    # ── CLASIFICACIÓN DE COMPETIDORES EN 3 ZONAS ──
    # El inventario/KPIs/productos usan el universo COMPLETO (azul+verde combinadas).
    # El morado define el set de competidores DIRECTOS para el análisis competitivo.
    # IMPORTANTE: la barrera de NSE recorta la zona de DEMANDA (qué AGEBs generan demanda),
    # no el set de oferta competidora. Si el método es barrera_nse, los competidores se
    # clasifican por isócrona (no por el polígono de demanda, que cubriría toda la oferta).
    zona_para_competidores = (None if perception.get("metodo") == "barrera_nse"
                              else perception.get("zona_poligono"))
    competidores = _safe(
        lambda: clasificar_competidores(resumen, base_ring, outer_ring, zona_para_competidores),
        "competidores",
        {"directos": [], "primarios": [], "secundarios": [],
         "n_directos": 0, "n_primarios": 0, "n_secundarios": 0})
    perception["competidores"] = competidores
    # REGLA DE COMPARABLES (Héctor · 8 jul 2026): un competidor es DIRECTO solo si COMPARTE
    # (1) isócrona primaria (llega el mismo cliente), (2) banda de PERCEPCIÓN DE VALOR
    # (mediana±MAD robusta de la zona) y (3) bloque NSE del entorno cuando hay geometría.
    # Estar dentro de la isócrona con OTRA percepción/NSE = SECUNDARIO (clientes compartidos,
    # no competidor directo). Prioridad: percepción/NSE MANDAN sobre la geometría.
    try:
        _st = perception.get("stats_robustas") or {}
        _med, _mad = _st.get("mediana"), _st.get("mad")
        if _med and competidores:
            _li = _med - 2.0 * (_mad or _med * 0.15)
            _ls = _med + 2.0 * (_mad or _med * 0.15)
            # NSE del entorno del pin (bloque dominante) para el criterio 3
            _pin_nse = None
            _gp = [g for g in (agebs_geo or []) if g.get("nse_rank") is not None
                   and g.get("lng") is not None]
            if _gp:
                _pin_nse = min(_gp, key=lambda q: (q["lng"] - req.lng) ** 2 +
                               (q["lat"] - req.lat) ** 2)["nse_rank"]
            def _es_directo(c):
                pm2 = c.get("pm2")
                if pm2 is None or not (_li <= pm2 <= _ls):
                    return False
                if _pin_nse is not None and _gp and c.get("lng") is not None:
                    n = min(_gp, key=lambda q: (q["lng"] - c["lng"]) ** 2 +
                            (q["lat"] - c["lat"]) ** 2)["nse_rank"]
                    if abs(n - _pin_nse) > 1:   # bloque NSE distinto → no comparable
                        return False
                return True
            nuevos_dir, degradados = [], []
            for c in (competidores.get("directos") or []):
                (nuevos_dir if _es_directo(c) else degradados).append(c)
            promovidos = []
            for c in (competidores.get("primarios") or []):
                if _es_directo(c):
                    promovidos.append(c)
            competidores["primarios"] = [c for c in (competidores.get("primarios") or [])
                                         if c not in promovidos]
            for c in degradados:
                c["set_competidor"] = "secundario"
                c["nota_set"] = "distinta percepción de valor/NSE (cliente compartido)"
            for c in promovidos:
                c["set_competidor"] = "directo"
            competidores["directos"] = nuevos_dir + promovidos
            competidores["secundarios"] = (competidores.get("secundarios") or []) + degradados
            competidores["n_directos"] = len(competidores["directos"])
            competidores["n_primarios"] = len(competidores["primarios"])
            competidores["n_secundarios"] = len(competidores["secundarios"])
            # SECUNDARIOS divididos por VALOR PERCIBIDO (Héctor · 8 jul): superior = contra
            # quién NO podemos compararnos por ser mejor; inferior = producto de percepción
            # inferior que no debe distraer. Corte por banda; dentro de banda → vs mediana.
            for c in competidores["secundarios"]:
                pm2 = c.get("pm2")
                if pm2 is None:
                    c["rango_percepcion"] = "N/D"
                elif pm2 > _ls or (pm2 >= _med and _li <= pm2 <= _ls):
                    c["rango_percepcion"] = "superior"
                else:
                    c["rango_percepcion"] = "inferior"
            competidores["n_sec_superior"] = sum(
                1 for c in competidores["secundarios"] if c.get("rango_percepcion") == "superior")
            competidores["n_sec_inferior"] = sum(
                1 for c in competidores["secundarios"] if c.get("rango_percepcion") == "inferior")
            competidores["criterio_directos"] = (
                f"isócrona primaria + banda de percepción ${round(_li):,}–${round(_ls):,}/m² "
                f"+ bloque NSE del entorno")
    except Exception as e:
        errors["comparables_banda"] = str(e)[:120]
    # Nota de zona (sin recortar inventario): describe el set directo
    if perception.get("barrera"):
        if perception.get("metodo") == "barrera_nse":
            perception["ajuste_inventario"] = (
                f"Zona de demanda recortada por barrera socioeconómica (NSE). "
                f"Competidores por cercanía: primarios (8 min) {competidores['n_primarios']} · "
                f"secundarios ({outer_min} min) {competidores['n_secundarios']}.")
        else:
            perception["ajuste_inventario"] = (
                f"Competidores directos (mercado del predio): {competidores['n_directos']} proyectos. "
                f"Primarios (8 min): {competidores['n_primarios']} · "
                f"Secundarios ({outer_min} min): {competidores['n_secundarios']}. "
                f"Inventario y KPIs usan el universo completo (zona primaria + secundaria).")

    demografia = _safe(lambda: derive_demografia(agebs, ft, agebs_geo, req.lng, req.lat), "demografia", derive_demografia([]))
    nse_dim = _safe(lambda: derive_nse_dim(agebs), "nse_dim", [])
    # DEMANDA según modo: horizontal pondera por proporción de casa; vertical usa la genérica.
    tipo_viv = "horizontal" if modo == "vivienda_horizontal" else "vertical"
    segments = _safe(lambda: derive_segments(agebs, ft, tipo_viv, agebs_geo, req.lng, req.lat, ft_renta), "segments", [])
    # PRODUCTO según modo. Vertical: depa (área privativa). Horizontal: casa (terreno+construcción).
    # La renta solo existe en vertical.
    if modo == "vivienda_horizontal":
        productos = _safe(lambda: derive_productos_horizontal(ft, segments, req.unidades_proyecto), "productos", [])
        productos_renta = []
    else:
        productos = _safe(lambda: derive_productos_venta(ft, segments, req.unidades_proyecto, demografia.get("personas_hogar"), _build_di_detail(agebs)), "productos", [])
        productos_renta = _safe(lambda: derive_productos_renta(ft_renta, segments), "productos_renta", [])
    comercio = _safe(lambda: derive_comercio(agebs, _nse_dominante_agebs(agebs, agebs_geo, req.lng, req.lat)[1], ft), "comercio", {})

    # ── VALOR DE ZONA EN CASCADA (regla de negocio: la zona SIEMPRE tiene un valor) ──
    # directos → primarios → secundarios → percepción de valor por NSE/capacidad de pago.
    valor_zona = _safe(lambda: _valor_zona_cascada(competidores, perception, demografia, segments, modo),
                       "valor_zona", None)
    if valor_zona is not None:
        # Nivel 4 (sin comparables): completar pm²/ticket con el PRODUCTO RECOMENDADO, que ya
        # viene anclado a la capacidad de pago real de la demanda (percepción de valor del NSE).
        prod_ref = next((p for p in (productos or []) if p.get("featured")),
                        next((p for p in (productos or []) if p.get("recomendado")),
                             (productos[0] if productos else None)))
        if valor_zona.get("fuente") == "percepcion_valor_nse" and not valor_zona.get("pm2"):
            if prod_ref:
                valor_zona["pm2"] = prod_ref.get("pm2_num")
                m2r = prod_ref.get("m2_num")
                valor_zona["m2_ref"] = m2r
                valor_zona["ticket_ref_M"] = (round(prod_ref["ticket_num"] / 1e6, 2)
                                              if prod_ref.get("ticket_num") else None)
        # Niveles 1-3 con pm² pero sin m² de comparables: usar el m² del producto recomendado
        # para expresar un ticket de referencia coherente (pm² observado × tamaño típico).
        elif valor_zona.get("_need_m2") and prod_ref and prod_ref.get("m2_num"):
            valor_zona["m2_ref"] = prod_ref["m2_num"]
            valor_zona["ticket_ref_M"] = round(valor_zona["pm2"] * prod_ref["m2_num"] / 1e6, 2)
        # Exponer el valor de zona en perception para que la sección "zona de análisis" lo muestre.
        perception["valor_zona"] = valor_zona
        # Si la percepción no tenía media (sin clúster), usar el pm² de la cascada como valor de zona.
        if not perception.get("media") and valor_zona.get("pm2"):
            perception["media"] = valor_zona["pm2"]
            perception["valor_zona_fuente"] = valor_zona["fuente"]

    # cluster_names es un set (no serializable) → convertir a lista para el payload
    if isinstance(perception.get("cluster_names"), set):
        perception["cluster_names"] = sorted(perception["cluster_names"])

    payload = assemble_zone_payload(
        req=req, profile=profile, isocronas=isocronas,
        resumen=resumen, ft=ft, pagos=pagos, resumen_renta=resumen_renta,
        demografia=demografia, nse_dim=nse_dim, segments=segments,
        productos=productos, productos_renta=productos_renta,
        comercio=comercio, perception=perception, agebs_count=len(agebs),
        agebs=agebs, ft_renta=ft_renta, agebs_geo=agebs_geo,
        producto_modo=modo,
    )
    payload["stage"] = "procesar"
    payload["producto"] = modo
    payload["producto_label"] = modo_cfg["label"]
    # ISO-MULTI · declarar SIEMPRE con qué fuente se construyó la zona (transparencia)
    try:
        payload["zone_data"]["iso_fuente_usada"] = iso_fuente_usada
        payload["zone_data"]["iso_nota"] = iso_nota
        if iso_nota:
            errors["iso_fallback"] = iso_nota
    except Exception:
        pass
    # CAPTABLE · derivar y exponer (opt-in; integridad: errores ya declarados)
    # v2 OD-SINTÉTICO: ancla censal si está disponible; si no, degrada a Huff v1 declarado.
    try:
        mercado_captable = None
        if captable_raw:
            geo_capt = parse_di_geometria(captable_raw["di"]) if captable_raw.get("di") else []
            mercado_captable = await derive_mercado_captable_v2(
                geo_capt, outer_ring, req.lng, req.lat,
                captable_raw["min"], captable_raw["fuente"], captable_raw["ring"],
                demografia.get("municipio"), demografia.get("estado"))
        # CAPTABLE-MASA (Héctor #2 · 8 jul 2026): población/hogares REALES de la corona desde
        # el XLSX del DI captable (el KMZ a veces no publica masa) + viajes diarios con
        # destino al municipio del pin (ancla censal OD) cuando el origen trae municipio.
        try:
            if mercado_captable and mercado_captable.get("origenes") and captable_raw:
                _agc = parse_di_xlsx(captable_raw.get("di")) if captable_raw.get("di") else []
                _hg, _n = {}, {}
                for r_ in _agc:
                    _nse = ageb_nse(r_)
                    _h = _num(r_.get("Hogares totales 2026"))
                    if _nse and _h:
                        _hg[_nse] = _hg.get(_nse, 0) + _h
                        _n[_nse] = _n.get(_nse, 0) + 1
                _sal = {}
                try:
                    _od = await _od_cargar_estado(req.estado or "")
                    if _od.get("ok") and _od.get("flujos"):
                        _sal = _od_salientes(_od["flujos"],
                                             demografia.get("municipio") or req.municipio or "")
                except Exception:
                    _sal = {}
                for o in mercado_captable["origenes"]:
                    if o.get("hogares_est") is None and o.get("nse") in _hg:
                        o["hogares_di"] = round(_hg[o["nse"]])
                        o["hogares_di_nota"] = "hogares del DI captable por NSE (dato real)"
                    if o.get("municipio") and _sal:
                        for _k, _vv in _sal.items():
                            if _norm_nombre(_k) == _norm_nombre(o["municipio"]):
                                o["viajes_destino_dia"] = round(_vv)
                                break
                if _hg:
                    mercado_captable["masa_di"] = {k: round(v) for k, v in _hg.items()}
                    mercado_captable["masa_di_fuente"] = "DI captable (XLSX) · hogares 2026 por NSE"
        except Exception as e:
            errors["captable_masa"] = str(e)[:120]
        payload["zone_data"]["mercado_captable"] = mercado_captable
        if mercado_captable and mercado_captable.get("activo") \
                and isinstance(payload["zone_data"].get("dem1_meta"), dict):
            payload["zone_data"]["dem1_meta"]["nota_captable"] = \
                "Mercado captable ACTIVO (modelo Huff v1) · orígenes en Zona de Análisis"
    except Exception as e:
        errors["mercado_captable"] = str(e)[:120]
    # MOVILIDAD (PEAT-1/TRAF-1) · capas de tráfico del pin — reemplazo de las capas de
    # Predik. Cliente propio y NUNCA bloqueante: si falla, el análisis sigue completo.
    try:
        async with httpx.AsyncClient(timeout=25) as _mcli:
            payload["zone_data"]["movilidad"] = await derive_movilidad(
                _mcli, req.lat, req.lng, agebs_geo)
    except Exception as e:
        payload["zone_data"]["movilidad"] = None
        errors["movilidad"] = str(e)[:120]
    payload["errors"] = errors
    # ARQ-MODULAR · dejar el análisis en caché para servir secciones independientes
    # (/api/zona/seccion) y devolver la llave al front (= identidad ZA-8).
    try:
        _key = (payload.get("zone_data") or {}).get("analisis_id_str")
        _analysis_cache_put(_key, payload)
        payload["analisis_key"] = _key
    except Exception as e:
        errors["cache"] = str(e)[:80]
    return payload


# ════════════════════════════════════════════════════════════════════════════
# CUENTAS DE USUARIO Y CONTROL DE ACCESO (modelo de operación)
# ════════════════════════════════════════════════════════════════════════════
# Preparación del backend para generar cuentas con atributos de acceso por sección
# y por zona. Almacén en memoria (sustituible por DB/identidad real en producción).
# El objetivo operativo: restringir el acceso DETALLADO a zonas no autorizadas,
# manteniendo el acceso GENERAL (preliminar) disponible para evaluaciones.

# ════════════ OD-SINTÉTICO · diagnóstico del ancla censal ════════════
@app.get("/api/od/status")
async def od_status(estado: str, muni: Optional[str] = None):
    """Diagnóstico del ancla censal SIN necesitar al equipo: reporta qué fuente de la
    cadena respondió (local/autofetch), cuántos flujos parseó y los marginales top hacia
    el municipio dado. Permite verificar/ajustar DATARIA_OD_URL_TPL desde Safari."""
    OD_CACHE.pop(_norm_nombre(estado), None)   # forzar recarga para diagnóstico honesto
    ancla = await _od_cargar_estado(estado)
    out = {"ok": ancla.get("ok"), "estado": estado, "fuente": ancla.get("fuente"),
           "error": ancla.get("error"),
           "config": {"autofetch": OD_AUTOFETCH, "url_tpl": OD_URL_TPL or "(sin configurar)",
                      "dir_local": str(OD_DIR_LOCAL)}}
    if ancla.get("ok"):
        fl = ancla["flujos"]
        out["flujos"] = {k: len(v) for k, v in fl.items()}
        if muni:
            marg = _od_marginales_hacia(fl, muni)
            out["marginales_hacia_muni"] = {
                k: sorted(v.items(), key=lambda kv: -kv[1])[:8] for k, v in marg.items()}
    return out


# ════════════ ZA-2 · GEOCODIFICACIÓN (dirección ⇄ pin) ════════════
# Cadena de proveedores: 1º ArcGIS World Geocoder (estándar Esri; P5: cuando el equipo
# PRSP exponga el geocodificador del ArcGIS de Prosperia se apunta ahí vía la variable
# de entorno DATARIA_ARCGIS_GEOCODE sin tocar código) · 2º Nominatim/OSM (abierto,
# gratuito, buena precisión en MX). Integridad: sin resultado → lista vacía, no inventa.
ARCGIS_GEOCODE_BASE = os.environ.get(
    "DATARIA_ARCGIS_GEOCODE",
    "https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer")
NOMINATIM_BASE = os.environ.get("DATARIA_NOMINATIM", "https://nominatim.openstreetmap.org")
_GEOCODE_UA = {"User-Agent": "Dataria/2.0 (contacto@prosperia.mx)"}


@app.get("/api/zona/geocode")
async def zona_geocode(q: str):
    """Dirección escrita → candidatos {lat,lng,label,colonia,municipio,estado} (ZA-2)."""
    out = {"ok": True, "q": q, "fuente": None, "resultados": []}
    async with httpx.AsyncClient(timeout=15) as client:
        try:
            r = await client.get(f"{ARCGIS_GEOCODE_BASE}/findAddressCandidates",
                                 params={"f": "json", "singleLine": q, "countryCode": "MEX",
                                         "maxLocations": 5, "outFields": "Nbrhd,City,Region"})
            for c in (r.json() or {}).get("candidates") or []:
                loc = c.get("location") or {}
                at = c.get("attributes") or {}
                if loc.get("y") is None:
                    continue
                out["resultados"].append({"lat": loc["y"], "lng": loc["x"],
                                          "label": c.get("address"),
                                          "colonia": at.get("Nbrhd") or None,
                                          "municipio": at.get("City") or None,
                                          "estado": at.get("Region") or None})
            if out["resultados"]:
                out["fuente"] = "arcgis"
                return out
        except Exception:
            pass
        try:
            r = await client.get(f"{NOMINATIM_BASE}/search", headers=_GEOCODE_UA,
                                 params={"format": "json", "q": q, "countrycodes": "mx",
                                         "limit": 5, "addressdetails": 1})
            for c in (r.json() or []):
                ad = c.get("address") or {}
                out["resultados"].append({"lat": float(c["lat"]), "lng": float(c["lon"]),
                                          "label": c.get("display_name"),
                                          "colonia": ad.get("neighbourhood") or ad.get("suburb"),
                                          "municipio": ad.get("city") or ad.get("town") or ad.get("municipality"),
                                          "estado": ad.get("state")})
            out["fuente"] = "nominatim" if out["resultados"] else None
        except Exception as e:
            out["ok"] = False
            out["error"] = str(e)[:120]
    return out


@app.get("/api/zona/reverse")
async def zona_reverse(lat: float, lng: float):
    """Pin → dirección (autollenar colonia/municipio/estado del formulario · ZA-2)."""
    out = {"ok": True, "lat": lat, "lng": lng, "fuente": None, "direccion": None,
           "colonia": None, "municipio": None, "estado": None, "pais": None}
    async with httpx.AsyncClient(timeout=15) as client:
        try:
            r = await client.get(f"{ARCGIS_GEOCODE_BASE}/reverseGeocode",
                                 params={"f": "json", "location": f"{lng},{lat}"})
            ad = (r.json() or {}).get("address") or {}
            if ad.get("City") or ad.get("LongLabel"):
                out.update(colonia=ad.get("Neighborhood"), municipio=ad.get("City"),
                           estado=ad.get("Region"), pais=ad.get("CntryName") or "México",
                           direccion=ad.get("LongLabel") or ad.get("Match_addr"),
                           fuente="arcgis")
                return out
        except Exception:
            pass
        try:
            r = await client.get(f"{NOMINATIM_BASE}/reverse", headers=_GEOCODE_UA,
                                 params={"format": "json", "lat": lat, "lon": lng,
                                         "addressdetails": 1})
            j = r.json() or {}
            ad = j.get("address") or {}
            out.update(colonia=ad.get("neighbourhood") or ad.get("suburb"),
                       municipio=ad.get("city") or ad.get("town") or ad.get("municipality"),
                       estado=ad.get("state"), pais=ad.get("country"),
                       direccion=j.get("display_name"),
                       fuente="nominatim" if (ad.get("state") or j.get("display_name")) else None)
        except Exception as e:
            out["ok"] = False
            out["error"] = str(e)[:120]
    return out


# ════════════ INV-3 · FICHAS DE INVENTARIO (PDF hoja carta para bancos) ════════════
class FichaInventarioRequest(BaseModel):
    analisis_key: str
    nombre_proyecto: str            # lo captura el usuario (portada)
    banco: str
    desarrollador: str
    proyecto: Optional[str] = None  # opcional: limitar a UN proyecto del inventario


def _ficha_pdf_bytes(payload: Dict[str, Any], req: "FichaInventarioRequest") -> bytes:
    """Construye el PDF (carta) con guías Dataria: portada, resumen de zona, una sección
    por proyecto y UNA HOJA POR PRODUCTO. Plusvalías por periodo → 'próximamente' (P3)."""
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, PageBreak)
    INK = colors.HexColor("#0B1020")
    AZURE = colors.HexColor("#0540F2")
    PULSE = colors.HexColor("#00B564")
    HAIR = colors.HexColor("#C9CCD6")
    PAPER = colors.HexColor("#F0F1F4")
    st_t = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=22, textColor=INK, spaceAfter=6)
    st_h = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=13, textColor=AZURE, spaceAfter=4)
    st_b = ParagraphStyle("b", fontName="Helvetica", fontSize=9, textColor=INK, leading=12)
    st_s = ParagraphStyle("s", fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#5A6072"))
    zd = payload.get("zone_data") or {}
    rc = zd.get("resumen_comercial") or {}
    os_ = zd.get("oferta_stats") or {}
    typ = zd.get("_typologies") or {}
    idstr = zd.get("analisis_id_str") or ""
    fmt = lambda v, p="$", s="": (p + f"{v:,.0f}" + s) if isinstance(v, (int, float)) else "—"
    prox = "próximamente (serie temporal en integración)"

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=18 * mm, bottomMargin=16 * mm,
                            title=f"Fichas de inventario · {req.nombre_proyecto}")
    E = []
    # ── PORTADA ──
    E.append(Spacer(1, 40 * mm))
    E.append(Paragraph("FICHAS DE INVENTARIO", st_t))
    E.append(Paragraph(req.nombre_proyecto, ParagraphStyle("p2", parent=st_t, fontSize=17, textColor=AZURE)))
    E.append(Spacer(1, 8 * mm))
    port = Table([["Banco", req.banco], ["Desarrollador", req.desarrollador],
                  ["Análisis", idstr], ["Zona", f"{zd.get('name') or ''} · {zd.get('subtitle') or ''}"],
                  ["Fecha", zd.get("analisis_fecha") or ""]], colWidths=[35 * mm, 130 * mm])
    port.setStyle(TableStyle([("FONT", (0, 0), (-1, -1), "Helvetica", 10),
                              ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 10),
                              ("TEXTCOLOR", (0, 0), (0, -1), AZURE),
                              ("LINEBELOW", (0, 0), (-1, -1), 0.4, HAIR),
                              ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    E.append(port)
    E.append(Spacer(1, 60 * mm))
    E.append(Paragraph("Elaborado: Dataria Team · San Pedro Garza García, Nuevo León y "
                       "Guadalajara, Jalisco · Hecho en México", st_s))
    E.append(PageBreak())
    # ── RESUMEN DE ZONA (medianas duales + top estrella) ──
    E.append(Paragraph("Resumen del corredor · métricas robustas", st_h))
    def _st(k):
        d = os_.get(k) or {}
        return f"{fmt(d.get('mediana'))} · núcleo {fmt(d.get('p25'))}–{fmt(d.get('p75'))} (n={d.get('n', 0)})"
    res_t = Table([["Métrica", "Todo el inventario", "Solo disponible"],
                   ["$/m² (mediana)", _st("pm2_total"), _st("pm2_disponible")],
                   ["Precio M (mediana)", _st("precio_M_total"), _st("precio_M_disponible")],
                   ["m² (mediana)", _st("m2_total"), _st("m2_disponible")],
                   ["Absorción un/mes", _st("abs_total"), _st("abs_disponible")]],
                  colWidths=[38 * mm, 66 * mm, 66 * mm])
    res_t.setStyle(TableStyle([("FONT", (0, 0), (-1, -1), "Helvetica", 8),
                               ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 8),
                               ("BACKGROUND", (0, 0), (-1, 0), INK),
                               ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                               ("GRID", (0, 0), (-1, -1), 0.4, HAIR),
                               ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PAPER])]))
    E.append(res_t)
    E.append(Spacer(1, 5 * mm))
    top = (zd.get("top_estrella") or {}).get("zona") or []
    if top:
        E.append(Paragraph("Top 3 · producto estrella de la zona", st_h))
        for i, e in enumerate(top, 1):
            E.append(Paragraph(f"★ {i}. {e.get('proyecto')} · {e.get('rec')} rec · {e.get('m2')} m² · "
                               f"${e.get('precio_M')}M · {fmt(e.get('pm2'))}/m² · desplazado "
                               f"{e.get('desplazamiento_pct')}% · {e.get('abs') or 's/'} un/mes", st_b))
        E.append(Paragraph(zd.get("criterio_estrella") or "", st_s))
    E.append(Paragraph(f"Plusvalía por periodo (mes/trimestre/histórica): {prox}.", st_s))
    E.append(PageBreak())
    # ── SECCIÓN POR PROYECTO + HOJA POR PRODUCTO ──
    nombres = [req.proyecto] if (req.proyecto and req.proyecto in typ) else sorted(typ.keys())
    for nombre in nombres:
        r = rc.get(nombre) or {}
        E.append(Paragraph(nombre, st_t))
        E.append(Paragraph(f"Estatus: {(r.get('estatus') or 'sin dato').upper()} · desplazado "
                           f"{r.get('desplazamiento_pct') if r.get('desplazamiento_pct') is not None else '—'}% · "
                           f"$/m² mediana {fmt(r.get('pm2_mediana'))} (disp {fmt(r.get('pm2_mediana_disp'))}) · "
                           f"tamaño mediana {r.get('m2_mediana') or '—'} m² · "
                           f"{r.get('n_tipologias') or len(typ.get(nombre) or [])} tipologías", st_b))
        est = r.get("estrella")
        if est:
            E.append(Paragraph(f"★ Producto estrella: {est.get('rec')} rec · {est.get('m2')} m² · "
                               f"${est.get('precio_M')}M · {fmt(est.get('pm2'))}/m² · "
                               f"desplazado {est.get('desplazamiento_pct')}%", st_b))
        E.append(Paragraph(f"Desarrollador: {req.desarrollador if req.proyecto == nombre else 'próximamente (dato en integración con la base)'}", st_s))
        E.append(PageBreak())
        for t in (typ.get(nombre) or []):
            E.append(Paragraph(f"{nombre} · Tipología {t.get('tipo')}", st_h))
            tot = (t.get("unid_vend") or 0) + (t.get("unid_disp") or 0)
            despl = round((t.get("unid_vend") or 0) / tot * 100, 1) if tot else None
            filas = [
                ["Programa (recámaras)", str(t.get("rec") if t.get("rec") is not None else "—")],
                ["Área privativa", f"{t.get('area_priv')} m²" if t.get("area_priv") else "—"],
                ["Área total", f"{t.get('area_total')} m²" if t.get("area_total") else "—"],
                ["Terraza", str(t.get("area_terr"))],
                ["Cajones de estacionamiento", str(t.get("cajones"))],
                ["Precio por unidad", fmt(t.get("precio_ud"))],
                ["Precio por m²", fmt(t.get("precio_m2"), "$", "/m²")],
                ["Unidades (tot/vend/disp)", f"{t.get('unid_total') or '—'} / {t.get('unid_vend') or 0} / {t.get('unid_disp') or 0}"],
                ["% desplazado", f"{despl}%" if despl is not None else "—"],
                ["Absorción observada", f"{t.get('abs')} un/mes" if t.get("abs") else "—"],
                ["Calidad / equipamiento", "próximamente (dato en integración con la base)"],
                ["Δ precio vs periodo anterior", prox],
                ["Plusvalía mes / trimestre / histórica", prox],
            ]
            tt = Table(filas, colWidths=[62 * mm, 105 * mm])
            tt.setStyle(TableStyle([("FONT", (0, 0), (-1, -1), "Helvetica", 9),
                                    ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 9),
                                    ("TEXTCOLOR", (0, 0), (0, -1), INK),
                                    ("GRID", (0, 0), (-1, -1), 0.4, HAIR),
                                    ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, PAPER]),
                                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
            E.append(tt)
            E.append(Spacer(1, 4 * mm))
            E.append(Paragraph("Fuente: base ArcGIS Prosperia · dato vigente del periodo más "
                               "reciente expuesto por el API.", st_s))
            E.append(PageBreak())
    def _pie(canv, _doc):
        canv.saveState()
        canv.setStrokeColor(HAIR)
        canv.line(18 * mm, 12 * mm, letter[0] - 18 * mm, 12 * mm)
        canv.setFont("Helvetica", 7)
        canv.setFillColor(colors.HexColor("#5A6072"))
        canv.drawString(18 * mm, 8.5 * mm, f"Dataria · {idstr}")
        canv.drawRightString(letter[0] - 18 * mm, 8.5 * mm, f"Página {canv.getPageNumber()}")
        canv.setFillColor(PULSE)
        canv.rect(18 * mm, 13.5 * mm, 10 * mm, 0.8 * mm, stroke=0, fill=1)
        canv.restoreState()
    doc.build(E, onFirstPage=_pie, onLaterPages=_pie)
    return buf.getvalue()


@app.post("/api/zona/ficha_inventario")
def ficha_inventario(req: FichaInventarioRequest):
    """INV-3 · Genera el PDF de fichas (hoja carta) desde el análisis en caché."""
    from fastapi import Response as _Resp
    payload = _analysis_cache_get(req.analisis_key)
    if not payload:
        return {"ok": False, "error": "Análisis no encontrado en caché · reprocesa la zona"}
    try:
        pdf = _ficha_pdf_bytes(payload, req)
    except ImportError:
        return {"ok": False, "error": "reportlab no instalado en el servidor (requirements.txt)"}
    except Exception as e:
        return {"ok": False, "error": f"Error generando PDF: {str(e)[:150]}"}
    nombre = re.sub(r"[^\w\-]+", "_", req.nombre_proyecto)[:40] or "ficha"
    return _Resp(content=pdf, media_type="application/pdf",
                 headers={"Content-Disposition": f'attachment; filename="fichas_{nombre}.pdf"'})


class CuentaCreate(BaseModel):
    usuario: str
    nombre: Optional[str] = None
    # Secciones habilitadas: lista de keys de SECTION_REGISTRY, o "*" para todas
    secciones: Any = "*"
    # Zonas con acceso a DETALLE: lista de zone_name, o "*" para todas
    zonas_detalle: Any = "*"
    # Rol informativo (no controla acceso por sí solo)
    rol: Optional[str] = "analista"


class MixEvalRequest(BaseModel):
    """Evaluación de amenaza competitiva del Monitor para uno o varios productos del mix interactivo.
    El front envía los items editables + horizonte + captación + los datos de zona que ya tiene
    (typologies y segments del payload), para que el BACKEND aplique la regla de negocio."""
    item: Optional[Dict[str, Any]] = None
    items: Optional[List[Dict[str, Any]]] = None
    period: int = 24
    capture: float = 1.0
    typologies: Dict[str, List[Dict[str, Any]]] = {}
    segments: List[Dict[str, Any]] = []


# Almacén en memoria (no persistente entre reinicios; en producción → base de datos)
_CUENTAS: Dict[str, Dict[str, Any]] = {}


def _cuenta_to_permisos(cuenta: Dict[str, Any], zona_actual: Optional[str]) -> Dict[str, Any]:
    return {
        "secciones": cuenta.get("secciones", "*"),
        "zonas_detalle": cuenta.get("zonas_detalle", "*"),
        "zona_actual": zona_actual,
    }


@app.get("/api/secciones")
def listar_secciones():
    """Devuelve el registro de secciones (para que el front sepa qué habilitar/mostrar)."""
    return {
        "secciones": [
            {"key": k, "label": v["label"], "page_id": v["page_id"],
             "access_tier": v["access_tier"], "scaffold": v.get("scaffold", False)}
            for k, v in SECTION_REGISTRY.items()
        ],
        "activas": section_keys(),
    }


@app.post("/api/zona/evaluar_mix")
def evaluar_mix(req: MixEvalRequest):
    """MONITOR · evalúa la amenaza competitiva de uno o varios productos del mix (regla de negocio
    en backend). El front llama aquí (con debounce) cada vez que el usuario edita el mix o mueve
    los selectores de horizonte/captación. Acepta un item (`item`) o el mix completo (`items`).
    Devuelve estrategia, threat ratio, demanda neta y competidores directos por producto.
    """
    if req.items is not None:
        out = [amenaza_competitiva(it, req.period, req.capture, req.typologies, req.segments)
               for it in req.items]
        return {"ok": True, "amenazas": out}
    res = amenaza_competitiva(req.item or {}, req.period, req.capture, req.typologies, req.segments)
    return {"ok": True, "amenaza": res}


@app.post("/api/cuentas")
def crear_cuenta(c: CuentaCreate):
    """Crea/actualiza una cuenta con sus atributos de acceso por sección y por zona."""
    _CUENTAS[c.usuario] = {
        "usuario": c.usuario, "nombre": c.nombre or c.usuario,
        "secciones": c.secciones, "zonas_detalle": c.zonas_detalle, "rol": c.rol,
    }
    return {"ok": True, "cuenta": _CUENTAS[c.usuario]}


@app.get("/api/cuentas")
def listar_cuentas():
    return {"cuentas": list(_CUENTAS.values())}


@app.get("/api/cuentas/{usuario}")
def obtener_cuenta(usuario: str):
    c = _CUENTAS.get(usuario)
    if not c:
        raise HTTPException(404, "Cuenta no encontrada")
    return c


@app.delete("/api/cuentas/{usuario}")
def borrar_cuenta(usuario: str):
    if usuario in _CUENTAS:
        del _CUENTAS[usuario]
        return {"ok": True}
    raise HTTPException(404, "Cuenta no encontrada")


class ZonaProcesarAuth(ZonaRequest):
    """zona/procesar con identidad opcional: si se envía `usuario`, el payload se filtra
    según los permisos de la cuenta (detalle vs preliminar). Sin usuario → acceso completo."""
    usuario: Optional[str] = None


@app.post("/api/zona/procesar_auth")
async def zona_procesar_auth(req: ZonaProcesarAuth):
    """
    Igual que /api/zona/procesar pero aplica control de acceso por cuenta.
    Si la zona no está autorizada para detalle, devuelve solo la vista preliminar.
    """
    base = ZonaRequest(**{k: getattr(req, k) for k in ZonaRequest.model_fields})
    payload = await zona_procesar(base)
    if req.usuario:
        cuenta = _CUENTAS.get(req.usuario)
        if not cuenta:
            raise HTTPException(403, "Cuenta no encontrada o sin permisos")
        permisos = _cuenta_to_permisos(cuenta, req.zone_name)
        payload = filter_payload_by_access(payload, permisos)
    return payload
